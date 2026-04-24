import os
import re
import shutil
import asyncio
import json
from pathlib import Path
from datetime import datetime
from typing import Any

from dotenv import load_dotenv
import requests
from openpyxl import load_workbook
from pypdf import PdfReader
from docx import Document

from telegram import Update
from telegram.error import NetworkError
from telegram.ext import Application, MessageHandler, ContextTypes, filters


BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env", override=True)

# ========= CONFIG =========
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()

# API do StarIA (FastAPI)
STAR_API_BASE = os.getenv("STAR_API_BASE", "http://127.0.0.1:8088").strip()
STAR_USE_RAG_DEFAULT = os.getenv("STAR_USE_RAG", "true").lower() == "true"
STAR_OLLAMA_MODEL = os.getenv("STAR_OLLAMA_MODEL", "star-llama").strip()

# Raiz oficial do StarIA
DRIVE_ROOT = Path(
    os.getenv("STARIA_DRIVE_ROOT", r"G:\Drives compartilhados\STARMKT\StarIA")
).resolve()

# Pastas operacionais
CURRICULOS_DIR = DRIVE_ROOT / "curriculos"
BANCO_TALENTOS_DIR = DRIVE_ROOT / "banco_talentos"
BANCO_TALENTOS_XLSX = BANCO_TALENTOS_DIR / "banco_talentos.xlsx"

# Script de indexação manual/rebuild
INDEXER_SCRIPT = Path(os.getenv("INDEXER_SCRIPT", r"C:\AI\backend\index_once.py"))

# Segurança opcional
ALLOWED_CHAT_IDS = os.getenv("ALLOWED_CHAT_IDS", "").strip()
ALLOWED = set()
if ALLOWED_CHAT_IDS:
    for x in ALLOWED_CHAT_IDS.split(","):
        x = x.strip()
        if x:
            ALLOWED.add(int(x))


# ========= HELPERS =========
def normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "")).strip()


def safe_str(value: Any) -> str:
    if value is None:
        return ""
    return normalize_spaces(str(value))


def sanitize_sheet_value(value: Any) -> str:
    s = safe_str(value)
    if not s:
        return ""
    if s.startswith(("=", "+", "-", "@")):
        return "'" + s
    return s


def route_folder(user_text: str) -> Path | None:
    t = (user_text or "").strip().lower()

    if re.search(r"\bcurr[ií]cul", t) or "curriculos" in t or "currículos" in t:
        return CURRICULOS_DIR

    return None


def safe_filename(name: str) -> str:
    name = (name or "").strip().replace("\u0000", "")
    name = re.sub(r'[<>:"/\\|?*]+', "_", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name[:180] if len(name) > 180 else name


def now_stamp() -> str:
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


def is_probably_question(text: str) -> bool:
    if not text:
        return False

    t = text.strip().lower()

    if t.startswith("/"):
        return False

    if "?" in t:
        return True

    return bool(
        re.match(
            r"^(quais|qual|me diga|liste|listar|lista|mostre|mostrar|resuma|procure|busque|tem|existe|verifique|explique|defina|como|quem|onde|quando|por que|porque|fala|me fale|quantos|quantas)\b",
            t,
        )
    )


def looks_like_greeting(text: str) -> bool:
    if not text:
        return False

    t = text.strip().lower()

    greetings = [
        "opa",
        "epa",
        "oi",
        "olá",
        "ola",
        "hey",
        "eai",
        "e aí",
        "e ai",
        "e ae",
        "bom dia",
        "boa tarde",
        "boa noite",
        "como vai",
        "tudo bem",
        "fala aí",
        "fala ai",
        "fala ae",
        "salve",
    ]

    if t in greetings:
        return True

    return any(t.startswith(g) for g in greetings)


def call_star_api(question: str, use_rag: bool = True) -> dict:
    payload = {
        "question": question,
        "use_rag": use_rag,
    }

    print("[BOT] DEBUG PAYLOAD:", payload)

    r = requests.post(
        f"{STAR_API_BASE}/ask",
        json=payload,
        timeout=60,
    )
    r.raise_for_status()
    data = r.json()
    print("[BOT] JSON da API:", data)
    return data


def looks_like_curriculos_inventory(text: str) -> bool:
    s = (text or "").strip().lower()

    if not re.search(r"\bcurr[ií]cul", s):
        return False

    patterns = [
        r"\bquant(os|as)\b.*\bexist(em|e)?\b",
        r"\bquantidade\b",
        r"\bconta(r)?\b",
        r"\blistar?\b",
        r"\bquais\b.*\bexist(em|e)?\b",
        r"\btem\b.*\bcurr",
        r"\bver\b.*\bcurr",
        r"\bn[uú]mero\b.*\bcurr",
        r"\btotal\b.*\bcurr",
    ]
    return any(re.search(p, s) for p in patterns)


def list_curriculos_files(limit: int = 50) -> list[dict]:
    exts = {".pdf", ".docx", ".doc", ".txt", ".rtf"}
    if not CURRICULOS_DIR.exists():
        return []

    items = []
    for p in CURRICULOS_DIR.rglob("*"):
        if p.is_file() and p.suffix.lower() in exts:
            try:
                st = p.stat()
                items.append(
                    {
                        "name": p.name,
                        "path": str(p),
                        "size": st.st_size,
                        "mtime": int(st.st_mtime),
                    }
                )
            except Exception:
                items.append({"name": p.name, "path": str(p)})

    items.sort(key=lambda x: x.get("mtime", 0), reverse=True)
    return items[:limit]


# ========= BANCO DE TALENTOS =========
def is_banco_talentos_question(text: str) -> bool:
    t = (text or "").strip().lower()
    triggers = [
        "banco de talentos",
        "banco talentos",
        "como está nosso banco de talentos",
        "como está o banco de talentos",
        "quantos candidatos temos",
        "quantos candidatos há",
        "quantos candidatos existem",
        "liste o banco de talentos",
        "listar banco de talentos",
        "me mostre o banco de talentos",
        "resuma o banco de talentos",
    ]
    return any(x in t for x in triggers)


def is_add_to_talent_bank_intent(text: str) -> bool:
    t = (text or "").strip().lower()
    triggers = [
        "adicione o candidato em anexo no nosso banco de talentos",
        "adicionar candidato no banco de talentos",
        "adicione no banco de talentos",
        "coloque no banco de talentos",
        "adicione o anexo no banco de talentos",
        "adicionar no banco de talentos",
        "salve no banco de talentos",
    ]
    return any(x in t for x in triggers)


def ensure_bank_paths():
    CURRICULOS_DIR.mkdir(parents=True, exist_ok=True)
    BANCO_TALENTOS_DIR.mkdir(parents=True, exist_ok=True)
    if not BANCO_TALENTOS_XLSX.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {BANCO_TALENTOS_XLSX}")


def _load_banco_talentos_rows(limit: int = 200) -> list[dict]:
    ensure_bank_paths()

    wb = load_workbook(BANCO_TALENTOS_XLSX, data_only=True)
    ws = wb.active

    headers = [safe_str(c.value) for c in ws[1]]
    rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {}
        has_any = False

        for idx, val in enumerate(row):
            key = headers[idx] if idx < len(headers) else f"col_{idx+1}"
            item[key] = val
            if val not in (None, ""):
                has_any = True

        if has_any:
            rows.append(item)

    return rows[:limit]


def build_banco_talentos_summary() -> str:
    rows = _load_banco_talentos_rows(limit=300)

    if not rows:
        return "📂 O banco de talentos está vazio no momento."

    total = len(rows)
    lines = [f"📌 Nosso banco de talentos possui {total} candidato(s) registrado(s).\n"]

    for r in rows[:10]:
        nome = safe_str(r.get("Nome completo"))
        cargo = safe_str(r.get("Cargo pretendido"))
        nivel = safe_str(r.get("Nível"))
        loc = safe_str(r.get("Localização"))

        line = f"- {nome or 'Sem nome'}"
        if cargo:
            line += f" | Cargo: {cargo}"
        if nivel:
            line += f" | Nível: {nivel}"
        if loc:
            line += f" | Local: {loc}"

        lines.append(line)

    if total > 10:
        lines.append(f"\n... (+{total - 10} candidato(s))")

    return "\n".join(lines)


def build_header_map(ws) -> dict:
    return {
        safe_str(cell.value): idx + 1
        for idx, cell in enumerate(ws[1])
        if safe_str(cell.value)
    }


def get_next_candidate_id(ws, header_map: dict) -> str:
    id_col = header_map.get("ID")
    if not id_col:
        return "BT0001"

    max_num = 0
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=id_col).value
        s = safe_str(value)
        m = re.match(r"BT(\d+)", s, flags=re.IGNORECASE)
        if m:
            max_num = max(max_num, int(m.group(1)))

    return f"BT{max_num + 1:04d}"


def extract_level(text: str) -> str:
    if not text:
        return ""

    text = text.lower()

    patterns = [
        r"n[íi]vel\s*:\s*(j[uú]nior|pleno|s[êe]nior)",
        r"senioridade\s*:\s*(j[uú]nior|pleno|s[êe]nior)",
    ]

    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            value = match.group(1)

            if value in ["junior", "júnior"]:
                return "Júnior"
            if value == "pleno":
                return "Pleno"
            if value in ["senior", "sênior"]:
                return "Sênior"

    return ""


def extract_portfolio(text: str) -> str:
    if not text:
        return ""

    pattern = r"(portf[oó]lio|portfolio)\s*:\s*(https?://\S+|\S+\.\S+)"
    match = re.search(pattern, text, flags=re.IGNORECASE)

    if match:
        return match.group(2).strip()

    return ""


def extract_text_from_pdf(file_path: Path) -> str:
    try:
        reader = PdfReader(str(file_path))
        parts = []
        for page in reader.pages:
            txt = page.extract_text() or ""
            if txt.strip():
                parts.append(txt)
        return "\n".join(parts).strip()
    except Exception as e:
        print(f"[BOT] Falha ao ler PDF {file_path.name}: {e}")
        return ""


def extract_text_from_docx(file_path: Path) -> str:
    try:
        doc = Document(str(file_path))
        parts = []
        for p in doc.paragraphs:
            txt = p.text or ""
            if txt.strip():
                parts.append(txt)
        return "\n".join(parts).strip()
    except Exception as e:
        print(f"[BOT] Falha ao ler DOCX {file_path.name}: {e}")
        return ""


def extract_text_from_file(file_path: Path) -> str:
    ext = file_path.suffix.lower()

    if ext in {".txt", ".rtf"}:
        try:
            return file_path.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            try:
                return file_path.read_text(errors="ignore")
            except Exception as e:
                print(f"[BOT] Falha ao ler TXT/RTF {file_path.name}: {e}")
                return ""

    if ext == ".pdf":
        return extract_text_from_pdf(file_path)

    if ext == ".docx":
        return extract_text_from_docx(file_path)

    return ""


def guess_job_title_from_text(text: str) -> str:
    if not text:
        return ""

    lines = [normalize_spaces(x) for x in text.splitlines() if normalize_spaces(x)]
    top = lines[:20]

    job_patterns = [
        r"\b(analista [\w\s]+)\b",
        r"\b(assistente [\w\s]+)\b",
        r"\b(coordenador[a]? [\w\s]+)\b",
        r"\b(gerente [\w\s]+)\b",
        r"\b(especialista [\w\s]+)\b",
        r"\b(supervisor[a]? [\w\s]+)\b",
        r"\b(designer(?: gráfico)?(?: [\w\s]+)?)\b",
        r"\b(social media)\b",
        r"\b(marketing(?: [\w\s]+)?)\b",
        r"\b(produtor[a]? de eventos)\b",
        r"\b(eventos corporativos)\b",
        r"\b(planejamento de eventos)\b",
    ]

    for line in top:
        lower = line.lower()
        for pattern in job_patterns:
            m = re.search(pattern, lower, flags=re.IGNORECASE)
            if m:
                return normalize_spaces(m.group(1)).title()

    return ""


def extract_json_block(text: str) -> str:
    if not text:
        return ""

    text = text.strip().replace("```json", "").replace("```", "").strip()

    start = text.find("{")
    end = text.rfind("}")

    if start >= 0 and end > start:
        return text[start:end + 1].strip()

    return text


def extract_candidate_data_with_ai(file_path: Path) -> dict:
    text = extract_text_from_file(file_path)
    print(f"[BOT] Texto extraído de {file_path.name}: {len(text)} caracteres")

    empty = {
        "nome_completo": "",
        "idade": "",
        "localizacao": "",
        "cargo_pretendido": "",
        "habilidades": "",
        "formacoes": "",
        "email": "",
        "telefone": "",
    }

    if not text.strip():
        return empty

    prompt = f"""
Extraia do currículo as informações abaixo e devolva APENAS JSON válido.

Campos obrigatórios do JSON:
{{
  "nome_completo": "",
  "idade": "",
  "localizacao": "",
  "cargo_pretendido": "",
  "habilidades": "",
  "formacoes": "",
  "email": "",
  "telefone": ""
}}

Regras:
- Não invente informações.
- Se não encontrar um campo, deixe "".
- Use para "cargo_pretendido" o cargo desejado, headline profissional, título principal do currículo ou área/cargo mais provável explicitamente indicada no topo do documento.
- "habilidades" deve ser uma string curta, com itens separados por "; ".
- "formacoes" deve ser uma string curta, com itens separados por "; ".
- Responda somente com JSON puro, sem markdown, sem explicação.

Nome do arquivo: {file_path.name}

Texto do currículo:
{text[:15000]}
""".strip()

    try:
        payload = {
            "question": prompt,
            "use_rag": False,
            "model": STAR_OLLAMA_MODEL,
        }

        print("[BOT] Chamando extração IA em:", STAR_API_BASE)
        r = requests.post(f"{STAR_API_BASE}/ask", json=payload, timeout=180)
        r.raise_for_status()

        data = r.json()
        answer = data.get("answer", "").strip()
        json_text = extract_json_block(answer)
        parsed = json.loads(json_text)

        extracted = {
            "nome_completo": safe_str(parsed.get("nome_completo")),
            "idade": safe_str(parsed.get("idade")),
            "localizacao": safe_str(parsed.get("localizacao")),
            "cargo_pretendido": safe_str(parsed.get("cargo_pretendido")),
            "habilidades": safe_str(parsed.get("habilidades")),
            "formacoes": safe_str(parsed.get("formacoes")),
            "email": safe_str(parsed.get("email")),
            "telefone": safe_str(parsed.get("telefone")),
        }

        if not extracted["cargo_pretendido"]:
            extracted["cargo_pretendido"] = guess_job_title_from_text(text)

        return extracted

    except Exception as e:
        print("[BOT] Falha na extração IA:", repr(e))
        fallback = empty.copy()
        fallback["cargo_pretendido"] = guess_job_title_from_text(text)
        return fallback


def append_candidate_to_sheet(
    file_path: Path,
    sender: str,
    level: str,
    portfolio: str,
    extracted: dict,
):
    ensure_bank_paths()

    wb = load_workbook(BANCO_TALENTOS_XLSX)
    ws = wb.active
    header_map = build_header_map(ws)

    required_headers = [
        "ID",
        "Nome completo",
        "Idade",
        "Localização",
        "Cargo pretendido",
        "Nível",
        "Portfólio",
        "Habilidades",
        "Formações",
        "Email",
        "Telefone",
        "Caminho do currículo",
        "Nome do arquivo",
        "Data de entrada",
        "Origem",
        "Remetente do email",
        "Status",
        "Observações",
    ]

    missing = [h for h in required_headers if h not in header_map]
    if missing:
        raise RuntimeError(f"Planilha sem colunas esperadas: {missing}")

    candidate_id = get_next_candidate_id(ws, header_map)

    # insere sempre no topo, abaixo do cabeçalho
    ws.insert_rows(2, amount=1)
    next_row = 2

    values = {
        "ID": candidate_id,
        "Nome completo": extracted.get("nome_completo", ""),
        "Idade": extracted.get("idade", ""),
        "Localização": extracted.get("localizacao", ""),
        "Cargo pretendido": extracted.get("cargo_pretendido", ""),
        "Nível": level or "",
        "Portfólio": portfolio or "",
        "Habilidades": extracted.get("habilidades", ""),
        "Formações": extracted.get("formacoes", ""),
        "Email": extracted.get("email", ""),
        "Telefone": extracted.get("telefone", ""),
        "Caminho do currículo": str(file_path),
        "Nome do arquivo": file_path.name,
        "Data de entrada": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Origem": "telegram",
        "Remetente do email": sender,
        "Status": "Banco de talentos",
        "Observações": "",
    }

    for header, value in values.items():
        col = header_map[header]
        ws.cell(row=next_row, column=col, value=sanitize_sheet_value(value))

    wb.save(BANCO_TALENTOS_XLSX)
    print(f"[BOT] Banco de talentos atualizado com {candidate_id}: {file_path.name}")


async def run_indexer_and_notify(msg):
    try:
        import subprocess
        import sys

        p = subprocess.run(
            [sys.executable, str(INDEXER_SCRIPT)],
            cwd=str(INDEXER_SCRIPT.parent),
            capture_output=True,
            text=True,
            timeout=600,
        )

        if p.returncode == 0:
            await msg.reply_text("📚 Indexação concluída.")
        else:
            err = (p.stderr or p.stdout or "").strip()
            await msg.reply_text(f"⚠️ Indexação falhou:\n{err[-1200:]}")
    except Exception as e:
        await msg.reply_text(f"⚠️ Não consegui rodar a indexação: {e}")


# ========= HANDLERS =========
async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id if update.effective_chat else None
    if ALLOWED and chat_id not in ALLOWED:
        await update.message.reply_text("Acesso não autorizado para este bot.")
        return

    msg = update.message
    if not msg or not msg.document:
        return

    caption_or_text = (msg.caption or msg.text or "").strip()

    # ===== FLUXO: adicionar no banco de talentos =====
    if is_add_to_talent_bank_intent(caption_or_text):
        try:
            ensure_bank_paths()

            doc = msg.document
            file_name = safe_filename(doc.file_name or f"arquivo_{now_stamp()}")
            tmp_dir = Path(r"C:\AI\runtime\tmp_telegram")
            tmp_dir.mkdir(parents=True, exist_ok=True)

            tmp_path = tmp_dir / f"{now_stamp()}__{file_name}"
            final_path = CURRICULOS_DIR / file_name

            tg_file = await context.bot.get_file(doc.file_id)
            await tg_file.download_to_drive(custom_path=str(tmp_path))

            if final_path.exists():
                stem = Path(file_name).stem
                suf = Path(file_name).suffix
                n = 2
                while True:
                    cand = CURRICULOS_DIR / f"{stem}__{n}{suf}"
                    if not cand.exists():
                        final_path = cand
                        break
                    n += 1

            shutil.move(str(tmp_path), str(final_path))
            print("[BOT] Currículo salvo no banco:", final_path)

            level = extract_level(caption_or_text)
            portfolio = extract_portfolio(caption_or_text)

            extracted = extract_candidate_data_with_ai(final_path)

            has_any_data = any([
                extracted.get("nome_completo"),
                extracted.get("idade"),
                extracted.get("localizacao"),
                extracted.get("cargo_pretendido"),
                extracted.get("habilidades"),
                extracted.get("formacoes"),
                extracted.get("email"),
                extracted.get("telefone"),
            ])

            if not has_any_data:
                await msg.reply_text(
                    "⚠️ Recebi o arquivo e salvei em currículos, mas não consegui extrair dados suficientes para criar o candidato no banco de talentos."
                )
                return

            sender = f"Telegram chat {chat_id}"
            append_candidate_to_sheet(
                file_path=final_path,
                sender=sender,
                level=level,
                portfolio=portfolio,
                extracted=extracted,
            )

            confirm = (
                "✅ Candidato adicionado ao banco de talentos com sucesso.\n\n"
                f"👤 Nome: {extracted.get('nome_completo') or 'Não identificado'}\n"
                f"💼 Cargo: {extracted.get('cargo_pretendido') or 'Não identificado'}\n"
                f"📎 Arquivo: {final_path.name}"
            )

            await msg.reply_text(confirm[:3500])

            await run_indexer_and_notify(msg)
            return

        except Exception as e:
            print("[BOT] Erro ao adicionar no banco de talentos:", repr(e))
            await msg.reply_text(f"⚠️ Erro ao adicionar no banco de talentos: {e}")
            return

    # ===== FLUXO ANTIGO: salvar em pasta por legenda =====
    target_dir = route_folder(caption_or_text)
    if target_dir is None:
        await msg.reply_text(
            "Não entendi a pasta de destino.\n"
            "Ex: envie o arquivo com a legenda: 'Coloque na pasta de currículos'.\n\n"
            "Ou use: 'adicione o candidato em anexo no nosso banco de talentos'."
        )
        return

    target_dir.mkdir(parents=True, exist_ok=True)

    doc = msg.document
    file_name = safe_filename(doc.file_name or f"arquivo_{now_stamp()}")
    tmp_dir = Path(r"C:\AI\runtime\tmp_telegram")
    tmp_dir.mkdir(parents=True, exist_ok=True)

    tmp_path = tmp_dir / f"{now_stamp()}__{file_name}"
    final_path = target_dir / file_name

    tg_file = await context.bot.get_file(doc.file_id)
    await tg_file.download_to_drive(custom_path=str(tmp_path))

    if final_path.exists():
        stem = Path(file_name).stem
        suf = Path(file_name).suffix
        n = 2
        while True:
            cand = target_dir / f"{stem}__{n}{suf}"
            if not cand.exists():
                final_path = cand
                break
            n += 1

    shutil.move(str(tmp_path), str(final_path))

    await msg.reply_text(f"✅ Recebido e salvo em:\n{final_path}")
    await run_indexer_and_notify(msg)


async def handle_text_only(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id if update.effective_chat else None
    if ALLOWED and chat_id not in ALLOWED:
        try:
            await update.message.reply_text("Acesso não autorizado para este bot.")
        except Exception as send_err:
            print("[BOT] Falha ao enviar acesso não autorizado:", repr(send_err))
        return

    msg = update.message
    t = (msg.text or "").strip() if msg and msg.text else ""
    if not t:
        return

    print("[BOT] Mensagem recebida:", t)

    # 1) Consulta direta ao banco de talentos
    if is_banco_talentos_question(t):
        try:
            out = build_banco_talentos_summary()
            await msg.reply_text(out[:3500])
            return
        except Exception as e:
            print("[BOT] Erro ao consultar banco de talentos:", repr(e))
            await msg.reply_text(f"⚠️ Erro ao consultar banco de talentos: {e}")
            return

    # 2) Inventário factual de currículos
    if looks_like_curriculos_inventory(t):
        try:
            files = list_curriculos_files(limit=80)
            total = len(files)

            if total == 0:
                await msg.reply_text(
                    "📂 Pasta de currículos está vazia.\n"
                    f"Destino: {CURRICULOS_DIR}"
                )
                return

            names = [f["name"] for f in files if f.get("name")]
            preview = names[:25]

            out = (
                f"📌 Currículos encontrados: {total}\n"
                f"📂 Pasta: {CURRICULOS_DIR}\n\n"
                "🗂 Últimos arquivos:\n- " + "\n- ".join(preview)
            )
            if total > len(preview):
                out += f"\n... (+{total - len(preview)})"

            await msg.reply_text(out[:3500])
            return

        except Exception as e:
            print("[BOT] Erro ao listar currículos:", repr(e))
            try:
                await msg.reply_text(f"⚠️ Erro ao listar currículos: {e}")
            except Exception as send_err:
                print("[BOT] Falha ao enviar erro ao Telegram:", repr(send_err))
            return

    # 3) Comando de pasta sem arquivo
    if route_folder(t) is not None and not is_probably_question(t):
        try:
            await msg.reply_text(
                "Entendi a pasta — agora me envie o arquivo (PDF/DOCX/XLSX) junto com essa legenda.\n"
                "Ex: envie o PDF com: 'Coloque na pasta de currículos'."
            )
        except Exception as send_err:
            print("[BOT] Falha ao enviar orientação de pasta:", repr(send_err))
        return

    # 4) Perguntas OU saudações vão para o StarIA
    if is_probably_question(t) or looks_like_greeting(t):
        try:
            print("[BOT] Chamando StarIA em:", STAR_API_BASE)

            data = await asyncio.to_thread(call_star_api, t, STAR_USE_RAG_DEFAULT)

            print("[BOT] Resposta recebida da API")

            answer = (data.get("answer") or "").strip() or "(sem resposta)"
            sources = data.get("sources") or []
            files = data.get("files") or []

            out = answer

            if files:
                names = [f.get("name") for f in files if isinstance(f, dict) and f.get("name")]
                if names:
                    out += "\n\n📎 Arquivos:\n- " + "\n- ".join(names[:20])

            if sources:
                out += "\n\n🧾 Fontes:\n- " + "\n- ".join(sources[:6])

            await msg.reply_text(out[:3500])
            return

        except Exception as e:
            print("[BOT] Erro ao consultar StarIA:", repr(e))
            try:
                await msg.reply_text(f"⚠️ Erro ao consultar StarIA: {e}")
            except Exception as send_err:
                print("[BOT] Falha ao enviar mensagem de erro ao Telegram:", repr(send_err))
            return

    # 5) Fallback final
    try:
        await msg.reply_text(
            "Me envie o arquivo (PDF/DOCX/XLSX) com uma legenda indicando a ação.\n\n"
            "Exemplos:\n"
            "- 'Coloque na pasta de currículos'\n"
            "- 'adicione o candidato em anexo no nosso banco de talentos'\n\n"
            "Ou faça uma pergunta (ex: 'Como está nosso banco de talentos?')."
        )
    except Exception as send_err:
        print("[BOT] Falha ao enviar mensagem padrão:", repr(send_err))


async def on_error(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    err = context.error

    if isinstance(err, NetworkError):
        print("[BOT] NetworkError transitório (reconectando):", err)
        return

    print("[BOT] Exceção não tratada:", repr(err))


def main():
    if not TELEGRAM_BOT_TOKEN:
        raise SystemExit("Defina TELEGRAM_BOT_TOKEN no ambiente (.env)")

    app = (
        Application.builder()
        .token(TELEGRAM_BOT_TOKEN)
        .read_timeout(30)
        .write_timeout(30)
        .connect_timeout(30)
        .pool_timeout(30)
        .build()
    )

    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text_only))
    app.add_error_handler(on_error)

    print("[ONLINE] StarIA running in polling mode")
    app.run_polling(
        allowed_updates=Update.ALL_TYPES,
        drop_pending_updates=True,
    )


if __name__ == "__main__":
    main()