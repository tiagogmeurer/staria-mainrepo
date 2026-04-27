from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook, Workbook

from rh.talent_bank_workbook import DEFAULT_BANCO_TALENTOS_XLSX, safe_str


OUTPUT_DIR = Path(r"G:\Drives compartilhados\STARMKT\StarIA\banco_talentos")


def parse_score(value) -> float:
    text = safe_str(value).replace(",", ".")
    try:
        return float(text)
    except Exception:
        return 0.0


def export_top3_by_sheet(
    source_path: Path = DEFAULT_BANCO_TALENTOS_XLSX,
    output_dir: Path = OUTPUT_DIR,
) -> dict:
    source_path = Path(source_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"top3_candidatos_por_vaga_{stamp}.xlsx"

    wb_src = load_workbook(source_path, data_only=False)
    wb_out = Workbook()

    default_ws = wb_out.active
    wb_out.remove(default_ws)

    total_exported = 0
    sheets_exported = []

    for ws_src in wb_src.worksheets:
        headers = [safe_str(c.value) for c in ws_src[1]]

        if "Nota" not in headers:
            continue

        nota_idx = headers.index("Nota") + 1

        rows = []
        for row in ws_src.iter_rows(min_row=2):
            values = [cell.value for cell in row]

            if not any(safe_str(v) for v in values):
                continue

            nota = parse_score(row[nota_idx - 1].value)
            rows.append((nota, values))

        rows.sort(key=lambda x: x[0], reverse=True)
        top3 = rows[:3]

        ws_out = wb_out.create_sheet(ws_src.title[:31])
        ws_out.append(headers)

        for _, values in top3:
            ws_out.append(values)
            total_exported += 1

        # Copia larguras aproximadas
        for col_idx, col in enumerate(ws_src.iter_cols(min_row=1, max_row=1), start=1):
            letter = ws_out.cell(row=1, column=col_idx).column_letter
            src_letter = ws_src.cell(row=1, column=col_idx).column_letter
            ws_out.column_dimensions[letter].width = ws_src.column_dimensions[src_letter].width

        ws_out.freeze_panes = "A2"
        ws_out.auto_filter.ref = ws_out.dimensions
        sheets_exported.append(ws_src.title)

    wb_out.save(output_path)

    return {
        "ok": True,
        "source": str(source_path),
        "output": str(output_path),
        "sheets_exported": sheets_exported,
        "total_exported": total_exported,
    }


if __name__ == "__main__":
    result = export_top3_by_sheet()
    print(json.dumps(result, ensure_ascii=False, indent=2))