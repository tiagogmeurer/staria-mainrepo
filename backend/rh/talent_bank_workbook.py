from __future__ import annotations

import os
import re
import shutil
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


DEFAULT_DRIVE_ROOT = Path(
    os.getenv("STARIA_DRIVE_ROOT", r"G:\Drives compartilhados\STARMKT\StarIA")
)

DEFAULT_BANCO_TALENTOS_XLSX = Path(
    os.getenv(
        "STARIA_TALENTS_XLSX",
        str(DEFAULT_DRIVE_ROOT / "banco_talentos" / "banco_talentos.xlsx"),
    )
)

DEFAULT_REFINED_XLSX = Path(
    os.getenv(
        "CANDIDATOS_REFINADOS_XLSX",
        str(DEFAULT_DRIVE_ROOT / "banco_talentos" / "candidatos_refinados.xlsx"),
    )
)


CANONICAL_SHEETS = [
    "EXECUTIVO DE CONTAS",
    "ATENDIMENTO",
    "COORDENADOR DE CONTEÚDO",
    "DIRETOR DE ARTE BRANDING",
    "DIRETOR DE ARTE DIGITAL",
    "DIRETOR DE ARTE INSTITUCIONAL",
    "DIAGRAMADOR",
    "PLAN. PERFORMANCE & GROWTH",
    "PLANEJAMENTO ESTRATÉGICO",
    "MOTION DESIGNER",
    "REDATOR",
]

VISIBLE_HEADERS = [
    "ID",
    "Nota",
    "Nome completo",
    "Localização",
    "Cargo pretendido",
    "Nível",
    "Portfólio",
    "Habilidades",
    "Formação",
    "Email",
    "Telefone",
]

TECHNICAL_HEADERS = [
    "Caminho do currículo",
    "Nome do arquivo",
    "Data de entrada",
    "Origem",
    "Remetente do email",
    "Status",
    "Observações",
    "Role ID sugerido",
    "Título normalizado",
    "Top 3 roles aderentes",
    "Resumo de aderência",
    "Flags de risco",
]

SHEET_DISPLAY_TITLES = {
    "EXECUTIVO DE CONTAS": "Executivo de Contas",
    "ATENDIMENTO": "Atendimento",
    "COORDENADOR DE CONTEÚDO": "Coordenador de Conteúdo",
    "DIRETOR DE ARTE BRANDING": "Diretor de Arte Branding",
    "DIRETOR DE ARTE DIGITAL": "Diretor de Arte Digital",
    "DIRETOR DE ARTE INSTITUCIONAL": "Diretor de Arte Institucional",
    "DIAGRAMADOR": "Diagramador",
    "PLAN. PERFORMANCE & GROWTH": "Plan. Performance & Growth",
    "PLANEJAMENTO ESTRATÉGICO": "Planejamento Estratégico",
    "MOTION DESIGNER": "Motion Designer",
    "REDATOR": "Redator",
}

FALLBACK_REFINED_HEADERS = [
    "Nota",
    "Nome completo",
    "Localização",
    "Cargo pretendido",
    "Nível",
    "Portfólio",
    "Habilidades",
    "Formação",
    "Email",
    "Telefone",
]

HEADER_ALIASES = {
    "Nome ": "Nome completo",
    "Nome": "Nome completo",
    "Nome Completo": "Nome completo",
    "Nome completo": "Nome completo",
    "Localizacao": "Localização",
    "Localização": "Localização",
    "Cargo Pretendido": "Cargo pretendido",
    "Cargo pretendido": "Cargo pretendido",
    "Nivel": "Nível",
    "Nível": "Nível",
    "Habilidades/Experiência": "Habilidades",
    "Habilidades / Experiência": "Habilidades",
    "Experiência": "Habilidades",
    "Experiencia": "Habilidades",
    "Competências": "Habilidades",
    "Competencias": "Habilidades",
    "Portfolio": "Portfólio",
    "Portfólio": "Portfólio",
    "Formações": "Formação",
    "Formacao": "Formação",
    "Formação": "Formação",
    "E-mail": "Email",
    "Email": "Email",
    "TELEFONE": "Telefone",
    "Telefone": "Telefone",
    "Curriculo": "Caminho do currículo",
    "Currículo": "Caminho do currículo",
    "COORDENADOR DE CONTEUDO": "COORDENADOR DE CONTEÚDO",
    "PLAN. PERFORMANCE & GROW": "PLAN. PERFORMANCE & GROWTH",
}


def safe_str(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()

def clean_excel_value(value: Any) -> Any:
    if value is None:
        return ""

    if isinstance(value, str):
        value = ILLEGAL_CHARACTERS_RE.sub("", value)
        value = re.sub(r"\s+", " ", value).strip()

        if value.startswith(("=", "+", "-", "@")):
            value = "'" + value

        return value

    return value    


def norm(value: Any) -> str:
    text = safe_str(value).lower()
    repl = {
        "á": "a",
        "à": "a",
        "â": "a",
        "ã": "a",
        "é": "e",
        "ê": "e",
        "í": "i",
        "ó": "o",
        "ô": "o",
        "õ": "o",
        "ú": "u",
        "ç": "c",
    }
    for a, b in repl.items():
        text = text.replace(a, b)
    return text


def parse_score(value: Any) -> float:
    text = safe_str(value).replace(",", ".")
    try:
        return float(text)
    except Exception:
        return 0.0


def normalize_header(header: Any) -> str:
    h = safe_str(header)
    return HEADER_ALIASES.get(h, h)


def normalize_sheet_name(sheet_name: Any) -> str:
    raw = safe_str(sheet_name)
    mapped = HEADER_ALIASES.get(raw, raw)
    n = norm(mapped)

    for canonical in CANONICAL_SHEETS:
        if norm(canonical) == n:
            return canonical

    if "coordenador" in n and "conteudo" in n:
        return "COORDENADOR DE CONTEÚDO"

    if "planejamento" in n and ("performance" in n or "growth" in n or "grow" in n):
        return "PLAN. PERFORMANCE & GROWTH"

    return raw


def sheet_display_title(sheet_name: str) -> str:
    sheet_name = normalize_sheet_name(sheet_name)
    return SHEET_DISPLAY_TITLES.get(sheet_name, safe_str(sheet_name).title())


def normalize_role_to_sheet_name(role: str) -> str:
    r = norm(role)

    if not r:
        return "EXECUTIVO DE CONTAS"

    if (
        "diagramador" in r
        or "diagramadora" in r
        or "diagramacao" in r
        or "tabloide" in r
        or "tablóide" in r
        or "ofertas" in r
        or "encarte" in r
    ):
        return "DIAGRAMADOR"

    if "diretor" in r and "arte" in r and (
        "branding" in r or "identidade" in r or "produto" in r
    ):
        return "DIRETOR DE ARTE BRANDING"

    if "diretor" in r and "arte" in r and (
        "digital" in r or "performance" in r or "marketplace" in r
    ):
        return "DIRETOR DE ARTE DIGITAL"

    if "diretor" in r and "arte" in r:
        return "DIRETOR DE ARTE INSTITUCIONAL"

    if "performance" in r or "growth" in r or "grow" in r:
        return "PLAN. PERFORMANCE & GROWTH"

    if "estrategista" in r or "planejamento" in r:
        return "PLANEJAMENTO ESTRATÉGICO"

    if "motion" in r:
        return "MOTION DESIGNER"

    if "redator" in r or "copywriter" in r or "copy" == r:
        return "REDATOR"

    if "conteudo" in r or "conteúdo" in r or "social media" in r:
        return "COORDENADOR DE CONTEÚDO"

    if "atendimento" in r or "account manager" in r or "relacionamento" in r:
        return "ATENDIMENTO"

    if "executivo" in r or "contas" in r or "comercial" in r:
        return "EXECUTIVO DE CONTAS"

    return "EXECUTIVO DE CONTAS"


def get_headers(ws) -> list[str]:
    return [normalize_header(c.value) for c in ws[1]]


def build_header_map(ws) -> dict[str, int]:
    out = {}
    for idx, cell in enumerate(ws[1], start=1):
        h = normalize_header(cell.value)
        if h:
            out[h] = idx
    return out


def backup_workbook(path: Path) -> Path | None:
    if not path.exists():
        return None

    backup_dir = path.parent / "_backups"
    backup_dir.mkdir(parents=True, exist_ok=True)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"{path.stem}_backup_{stamp}{path.suffix}"
    shutil.copy2(path, backup_path)
    return backup_path


def copy_cell_style(src, dst) -> None:
    if src is None:
        return

    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def first_non_empty_header_cell(ws):
    for cell in ws[1]:
        if safe_str(cell.value):
            return cell
    return ws.cell(row=1, column=1)


def load_or_create_workbook(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        return load_workbook(path, data_only=False)

    wb = Workbook()
    ws = wb.active
    ws.title = CANONICAL_SHEETS[0]
    return wb


def get_refined_workbook(refined_path: Path):
    if refined_path.exists():
        return load_workbook(refined_path, data_only=False)
    return None


def get_refined_sheet_headers(template_ws) -> list[str]:
    if template_ws is None:
        return FALLBACK_REFINED_HEADERS.copy()

    headers = []
    for h in get_headers(template_ws):
        if not h:
            continue
        if h == "ID":
            continue
        headers.append(h)

    if not headers:
        headers = FALLBACK_REFINED_HEADERS.copy()

    seen = set()
    deduped = []
    for h in headers:
        if h not in seen:
            seen.add(h)
            deduped.append(h)

    return deduped


def build_bank_headers(refined_headers: list[str] | None = None) -> list[str]:
    headers = VISIBLE_HEADERS.copy()

    for h in TECHNICAL_HEADERS:
        if h not in headers:
            headers.append(h)

    return headers


def normalize_row_data(row: dict[str, Any]) -> dict[str, Any]:
    out = {}

    for key, value in row.items():
        normalized_key = normalize_header(key)
        out[normalized_key] = value

    if safe_str(out.get("Formações")) and not safe_str(out.get("Formação")):
        out["Formação"] = out.get("Formações")

    if safe_str(out.get("Nome Completo")) and not safe_str(out.get("Nome completo")):
        out["Nome completo"] = out.get("Nome Completo")

    return out


def extract_rows_from_workbook(wb) -> list[dict[str, Any]]:
    rows = []

    for ws in wb.worksheets:
        header_map = build_header_map(ws)
        if not header_map:
            continue

        headers = get_headers(ws)

        for row in ws.iter_rows(min_row=2, values_only=True):
            item = {"_source_sheet": normalize_sheet_name(ws.title)}
            has_any = False

            for idx, val in enumerate(row):
                header = headers[idx] if idx < len(headers) else f"col_{idx + 1}"
                header = normalize_header(header)

                if not header:
                    continue

                item[header] = val

                if val not in (None, ""):
                    has_any = True

            if has_any:
                rows.append(normalize_row_data(item))

    return rows


def duplicate_key_from_row(row: dict[str, Any]) -> str:
    row = normalize_row_data(row)

    email = normalize_duplicate_key(row.get("Email"))
    phone = normalize_phone(row.get("Telefone"))
    name = normalize_duplicate_key(row.get("Nome completo"))
    role = normalize_duplicate_key(row.get("Cargo pretendido"))
    resume = normalize_duplicate_key(row.get("Nome do arquivo") or row.get("Caminho do currículo"))

    if email and role:
        return f"email:{email}|role:{role}"

    if phone and len(phone) >= 8 and role:
        return f"phone:{phone}|role:{role}"

    if name and role:
        return f"name:{name}|role:{role}"

    if resume:
        return f"resume:{resume}"

    return f"row:{name}|{phone}|{email}|{role}|{resume}"


def merge_row_values(base: dict[str, Any], incoming: dict[str, Any]) -> dict[str, Any]:
    base = normalize_row_data(base)
    incoming = normalize_row_data(incoming)

    base_score = parse_score(base.get("Nota"))
    incoming_score = parse_score(incoming.get("Nota"))

    if incoming_score > base_score:
        primary = incoming.copy()
        secondary = base
    else:
        primary = base.copy()
        secondary = incoming

    for key, value in secondary.items():
        if key.startswith("_"):
            continue
        if safe_str(value) and not safe_str(primary.get(key)):
            primary[key] = value

    return primary


def merge_duplicate_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: dict[str, dict[str, Any]] = {}
    ordered_keys = []

    for row in rows:
        row = normalize_row_data(row)
        key = duplicate_key_from_row(row)

        if key not in seen:
            seen[key] = row
            ordered_keys.append(key)
            continue

        seen[key] = merge_row_values(seen[key], row)

    return [seen[k] for k in ordered_keys]


def copy_template_dimensions(target_ws, target_headers: list[str], template_ws=None) -> None:
    if template_ws is None:
        for idx, header in enumerate(target_headers, start=1):
            letter = target_ws.cell(row=1, column=idx).column_letter
            target_ws.column_dimensions[letter].width = 22
        return

    template_headers = get_headers(template_ws)

    for idx, header in enumerate(target_headers, start=1):
        target_letter = target_ws.cell(row=1, column=idx).column_letter

        if header == "ID":
            src_cell = first_non_empty_header_cell(template_ws)
            src_letter = src_cell.column_letter
        elif header in template_headers:
            src_idx = template_headers.index(header) + 1
            src_letter = template_ws.cell(row=1, column=src_idx).column_letter
        else:
            src_cell = first_non_empty_header_cell(template_ws)
            src_letter = src_cell.column_letter

        width = template_ws.column_dimensions[src_letter].width
        target_ws.column_dimensions[target_letter].width = width or 22

    for header in TECHNICAL_HEADERS:
        if header in target_headers:
            idx = target_headers.index(header) + 1
            letter = target_ws.cell(row=1, column=idx).column_letter
            target_ws.column_dimensions[letter].hidden = True
            target_ws.column_dimensions[letter].width = 18


def apply_header_styles(target_ws, target_headers: list[str], template_ws=None) -> None:
    template_headers = get_headers(template_ws) if template_ws is not None else []

    for idx, header in enumerate(target_headers, start=1):
        dst = target_ws.cell(row=1, column=idx)
        dst.value = header

        src = None

        if template_ws is not None:
            if header == "ID":
                src = first_non_empty_header_cell(template_ws)
            elif header in template_headers:
                src = template_ws.cell(row=1, column=template_headers.index(header) + 1)
            else:
                src = first_non_empty_header_cell(template_ws)

        copy_cell_style(src, dst)


def apply_body_row_style(target_ws, row_idx: int, target_headers: list[str], template_ws=None) -> None:
    if template_ws is None:
        return

    template_headers = get_headers(template_ws)
    template_row = 2 if template_ws.max_row >= 2 else 1

    for idx, header in enumerate(target_headers, start=1):
        dst = target_ws.cell(row=row_idx, column=idx)

        src = None

        if header == "ID":
            src = template_ws.cell(
                row=template_row,
                column=first_non_empty_header_cell(template_ws).column,
            )
        elif header in template_headers:
            src = template_ws.cell(row=template_row, column=template_headers.index(header) + 1)
        else:
            src = template_ws.cell(
                row=template_row,
                column=first_non_empty_header_cell(template_ws).column,
            )

        copy_cell_style(src, dst)


def clear_and_write_sheet(ws, headers: list[str], rows: list[dict[str, Any]], template_ws=None) -> None:
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)

    ws.append(headers)
    apply_header_styles(ws, headers, template_ws)

    for row_data in rows:
        row_data = normalize_row_data(row_data)
        ws.append([row_data.get(h, "") for h in headers])
        apply_body_row_style(ws, ws.max_row, headers, template_ws)

    ws.freeze_panes = "A2"
    copy_template_dimensions(ws, headers, template_ws)

    if ws.max_row >= 1:
        ws.auto_filter.ref = ws.dimensions


def get_next_candidate_id_from_rows(rows: list[dict[str, Any]]) -> str:
    max_num = 0

    for row in rows:
        value = safe_str(row.get("ID"))
        m = re.match(r"BT(\d+)", value, flags=re.IGNORECASE)
        if m:
            max_num = max(max_num, int(m.group(1)))

    return f"BT{max_num + 1:04d}"


def ensure_ids(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    max_num = 0

    for row in rows:
        value = safe_str(row.get("ID"))
        m = re.match(r"BT(\d+)", value, flags=re.IGNORECASE)
        if m:
            max_num = max(max_num, int(m.group(1)))

    for row in rows:
        if safe_str(row.get("ID")):
            continue
        max_num += 1
        row["ID"] = f"BT{max_num:04d}"

    return rows


def canonical_sheet_rules() -> dict[str, list[str]]:
    return {
        "EXECUTIVO DE CONTAS": ["executivo de contas", "account executive", "executivo comercial"],
        "ATENDIMENTO": ["atendimento", "account manager", "relacionamento", "cliente interno"],
        "COORDENADOR DE CONTEÚDO": ["coordenador de conteudo", "coordenador de conteúdo", "conteudo", "conteúdo", "social media", "editorial"],
        "DIRETOR DE ARTE BRANDING": ["diretor de arte branding", "branding", "identidade visual", "brand", "marca"],
        "DIRETOR DE ARTE DIGITAL": ["diretor de arte digital", "da digital", "performance digital", "marketplace", "ecommerce", "e-commerce"],
        "DIRETOR DE ARTE INSTITUCIONAL": ["diretor de arte institucional", "campanhas institucionais", "institucional", "diretor de arte", "designer grafico senior", "designer gráfico sênior"],
        "DIAGRAMADOR": ["diagramador", "diagramadora", "diagramação", "diagramacao", "tabloide", "tablóide", "ofertas", "encarte"],
        "PLAN. PERFORMANCE & GROWTH": ["performance", "growth", "grow", "google ads", "meta ads", "analytics", "tiktok ads"],
        "PLANEJAMENTO ESTRATÉGICO": ["estrategista", "planejamento estrategico", "planejamento estratégico", "brand strategy"],
        "MOTION DESIGNER": ["motion", "motion designer", "after effects", "premiere", "animação", "animacao"],
        "REDATOR": ["redator", "redatora", "copywriter", "copy", "redação", "redacao"],
    }


def choose_sheet_name(sheet_names: list[str], row: dict[str, Any]) -> str:
    normalized_sheet_names = [normalize_sheet_name(s) for s in sheet_names]

    direct_role_to_sheet = {
        "executivo_de_contas": "EXECUTIVO DE CONTAS",
        "atendimento_senior": "ATENDIMENTO",
        "coordenador_conteudo": "COORDENADOR DE CONTEÚDO",
        "diretor_arte_senior_branding_produto": "DIRETOR DE ARTE BRANDING",
        "diretor_arte_senior_digital": "DIRETOR DE ARTE DIGITAL",
        "diretor_arte_senior_campanhas": "DIRETOR DE ARTE INSTITUCIONAL",
        "diagramador_ofertas": "DIAGRAMADOR",
        "performance_growth_planejamento": "PLAN. PERFORMANCE & GROWTH",
        "estrategista_senior_planejamento": "PLANEJAMENTO ESTRATÉGICO",
        "motion_designer": "MOTION DESIGNER",
        "redator_digital": "REDATOR",
    }

    role_id = safe_str(row.get("Role ID sugerido"))
    if role_id in direct_role_to_sheet:
        target = direct_role_to_sheet[role_id]
        if target in normalized_sheet_names:
            return target

    cargo_sheet = normalize_role_to_sheet_name(safe_str(row.get("Cargo pretendido")))
    if cargo_sheet in normalized_sheet_names:
        return cargo_sheet

    title_sheet = normalize_role_to_sheet_name(safe_str(row.get("Título normalizado")))
    if title_sheet in normalized_sheet_names:
        return title_sheet

    haystack = norm(
        " | ".join(
            [
                safe_str(row.get("Role ID sugerido")),
                safe_str(row.get("Título normalizado")),
                safe_str(row.get("Cargo pretendido")),
                safe_str(row.get("Habilidades")),
                safe_str(row.get("Observações")),
            ]
        )
    )

    rules = canonical_sheet_rules()

    best_sheet = ""
    best_score = 0

    for sheet_name in normalized_sheet_names:
        aliases = rules.get(sheet_name, [])
        score = sum(1 for alias in aliases if norm(alias) in haystack)

        if score > best_score:
            best_score = score
            best_sheet = sheet_name

    if best_sheet:
        return best_sheet

    source_sheet = normalize_sheet_name(row.get("_source_sheet"))
    if source_sheet in normalized_sheet_names:
        return source_sheet

    return CANONICAL_SHEETS[0]


def redistribute_rows_by_sheet(
    rows: list[dict[str, Any]],
    sheet_names: list[str],
) -> dict[str, list[dict[str, Any]]]:
    buckets = {normalize_sheet_name(name): [] for name in sheet_names}

    for row in rows:
        row = normalize_row_data(row)
        sheet_name = choose_sheet_name(list(buckets.keys()), row)
        row["Cargo pretendido"] = sheet_display_title(sheet_name)
        buckets.setdefault(sheet_name, []).append(row)

    return buckets


def ensure_bank_workbook_structure(
    banco_path: Path = DEFAULT_BANCO_TALENTOS_XLSX,
    refined_path: Path = DEFAULT_REFINED_XLSX,
    create_backup: bool = True,
    redistribute_existing: bool = True,
) -> dict[str, Any]:
    banco_path = Path(banco_path)
    refined_path = Path(refined_path)

    banco_path.parent.mkdir(parents=True, exist_ok=True)

    backup_path = None
    if create_backup and banco_path.exists():
        backup_path = backup_workbook(banco_path)

    wb_bank = load_or_create_workbook(banco_path)
    existing_rows = merge_duplicate_rows(extract_rows_from_workbook(wb_bank))
    existing_rows = ensure_ids(existing_rows)

    wb_ref = get_refined_workbook(refined_path)

    target_sheet_names = CANONICAL_SHEETS.copy()

    # Remove abas fora do padrão canônico depois de preservar as linhas.
    for sheet_name in list(wb_bank.sheetnames):
        if normalize_sheet_name(sheet_name) not in target_sheet_names:
            del wb_bank[sheet_name]

    # Renomeia abas antigas equivalentes para o padrão canônico.
    for sheet_name in list(wb_bank.sheetnames):
        normalized = normalize_sheet_name(sheet_name)
        if normalized != sheet_name and normalized in target_sheet_names:
            if normalized not in wb_bank.sheetnames:
                wb_bank[sheet_name].title = normalized
            else:
                del wb_bank[sheet_name]

    for sheet_name in target_sheet_names:
        if sheet_name not in wb_bank.sheetnames:
            wb_bank.create_sheet(sheet_name)

    if "Sheet" in wb_bank.sheetnames and len(wb_bank.sheetnames) > 1:
        del wb_bank["Sheet"]

    buckets = (
        redistribute_rows_by_sheet(existing_rows, target_sheet_names)
        if redistribute_existing
        else {target_sheet_names[0]: existing_rows}
    )

    for sheet_name in target_sheet_names:
        ws = wb_bank[sheet_name]
        template_ws = (
            wb_ref[sheet_name]
            if wb_ref is not None and sheet_name in wb_ref.sheetnames
            else None
        )

        headers = build_bank_headers(None)

        clear_and_write_sheet(ws, headers, buckets.get(sheet_name, []), template_ws)

    wb_bank.save(banco_path)

    return {
        "ok": True,
        "banco_path": str(banco_path),
        "refined_path": str(refined_path),
        "backup_path": str(backup_path) if backup_path else "",
        "sheets": target_sheet_names,
        "rows_preserved": len(existing_rows),
    }


def normalize_duplicate_key(value: Any) -> str:
    text = norm(safe_str(value))
    text = re.sub(r"[^a-z0-9@./+-]", "", text)
    return text


def normalize_phone(value: Any) -> str:
    return re.sub(r"\D+", "", safe_str(value))


def is_duplicate_candidate(existing_rows: list[dict[str, Any]], values: dict[str, Any]) -> bool:
    new_values = normalize_row_data(values)

    new_email = normalize_duplicate_key(new_values.get("Email"))
    new_phone = normalize_phone(new_values.get("Telefone"))
    new_resume_name = normalize_duplicate_key(new_values.get("Nome do arquivo"))
    new_name = normalize_duplicate_key(new_values.get("Nome completo"))
    new_role = normalize_duplicate_key(new_values.get("Cargo pretendido"))

    for row in existing_rows:
        row = normalize_row_data(row)

        old_email = normalize_duplicate_key(row.get("Email"))
        old_phone = normalize_phone(row.get("Telefone"))
        old_resume_name = normalize_duplicate_key(row.get("Nome do arquivo"))
        old_name = normalize_duplicate_key(row.get("Nome completo"))
        old_role = normalize_duplicate_key(row.get("Cargo pretendido"))

        if new_email and old_email and new_email == old_email and new_role == old_role:
            return True

        if (
            new_phone
            and old_phone
            and len(new_phone) >= 8
            and new_phone == old_phone
            and new_role == old_role
        ):
            return True

        if new_resume_name and old_resume_name and new_resume_name == old_resume_name:
            return True

        if new_name and old_name and new_name == old_name and new_role == old_role:
            return True

    return False


def append_candidate_record(
    values: dict[str, Any],
    banco_path: Path = DEFAULT_BANCO_TALENTOS_XLSX,
    refined_path: Path = DEFAULT_REFINED_XLSX,
    target_sheet: str | None = None,
) -> str:
    banco_path = Path(banco_path)

    if not banco_path.exists():
        ensure_bank_workbook_structure(
            banco_path=banco_path,
            refined_path=refined_path,
            create_backup=False,
            redistribute_existing=True,
        )

    wb = load_workbook(banco_path)
    all_existing_rows = extract_rows_from_workbook(wb)

    values = normalize_row_data(dict(values))

    sheet_name = normalize_sheet_name(target_sheet) if target_sheet else choose_sheet_name(wb.sheetnames, values)

    if sheet_name not in wb.sheetnames:
        sheet_name = choose_sheet_name(wb.sheetnames, values)

    values["Cargo pretendido"] = sheet_display_title(sheet_name)

    if is_duplicate_candidate(all_existing_rows, values):
        print(
            "[TALENT BANK] Candidato duplicado detectado. Registro ignorado:",
            values.get("Nome completo") or values.get("Email") or values.get("Nome do arquivo"),
        )
        return "DUPLICADO"

    candidate_id = safe_str(values.get("ID")) or get_next_candidate_id_from_rows(all_existing_rows)

    values["ID"] = candidate_id

    ws = wb[sheet_name]
    header_map = build_header_map(ws)

    # Garante headers canônicos mesmo se a aba veio irregular.
    expected_headers = build_bank_headers(None)
    current_headers = get_headers(ws)
    if current_headers[: len(expected_headers)] != expected_headers:
        existing_sheet_rows = extract_rows_from_workbook(wb)
        wb.save(banco_path)
        ensure_bank_workbook_structure(
            banco_path=banco_path,
            refined_path=refined_path,
            create_backup=False,
            redistribute_existing=True,
        )
        wb = load_workbook(banco_path)
        ws = wb[sheet_name]
        header_map = build_header_map(ws)

    ws.insert_rows(2, amount=1)

    if ws.max_row >= 3:
        for col in range(1, ws.max_column + 1):
            copy_cell_style(ws.cell(row=3, column=col), ws.cell(row=2, column=col))

    for header, col in header_map.items():
        ws.cell(row=2, column=col, value=clean_excel_value(values.get(header, "")))

    wb.save(banco_path)
    return candidate_id


def find_rows_missing_core_fields(
    banco_path: Path = DEFAULT_BANCO_TALENTOS_XLSX,
) -> list[dict[str, Any]]:
    wb = load_workbook(banco_path, data_only=True)
    missing = []

    for ws in wb.worksheets:
        header_map = build_header_map(ws)
        headers = get_headers(ws)

        for row_idx in range(2, ws.max_row + 1):
            row_data = {"_sheet": ws.title, "_row": row_idx}
            has_any = False

            for col_idx, header in enumerate(headers, start=1):
                if not header:
                    continue
                value = ws.cell(row=row_idx, column=col_idx).value
                row_data[header] = value
                if value not in (None, ""):
                    has_any = True

            if not has_any:
                continue

            row_data = normalize_row_data(row_data)

            missing_name = not safe_str(row_data.get("Nome completo"))
            missing_location = not safe_str(row_data.get("Localização"))
            missing_phone = not safe_str(row_data.get("Telefone"))

            has_curriculum = bool(safe_str(row_data.get("Caminho do currículo")))

            if has_curriculum and (missing_name or missing_location or missing_phone):
                missing.append(row_data)

    return missing


def update_candidate_row(
    sheet_name: str,
    row_idx: int,
    updates: dict[str, Any],
    banco_path: Path = DEFAULT_BANCO_TALENTOS_XLSX,
    only_if_blank: bool = True,
) -> None:
    wb = load_workbook(banco_path)

    sheet_name = normalize_sheet_name(sheet_name)

    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Aba não encontrada: {sheet_name}")

    ws = wb[sheet_name]
    header_map = build_header_map(ws)
    updates = normalize_row_data(updates)

    for header, value in updates.items():
        header = normalize_header(header)
        if header not in header_map:
            continue

        col = header_map[header]
        current = ws.cell(row=row_idx, column=col).value

        if only_if_blank and safe_str(current):
            continue

        if safe_str(value):
            ws.cell(row=row_idx, column=col, value=value)

    wb.save(banco_path)
