from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook
from pypdf import PdfReader
from docx import Document

from rh.talent_bank_workbook import (
    DEFAULT_BANCO_TALENTOS_XLSX,
    safe_str,
    build_header_map,
)


STARIA_API_BASE = "http://127.0.0.1:8088"
STARIA_MODEL = "star-llama:latest"

CURRICULOS_FALLBACK_DIRS = [
    Path(r"G:\Drives compartilhados\STARMKT\StarIA\banco_talentos\curriculos"),
    Path(r"G:\Drives compartilhados\STARMKT\StarIA\curriculos"),
]

SUPPORTED_EXTS = {".pdf", ".docx", ".txt", ".rtf"}

FIELDS_TO_REPAIR = [
    "Nome completo",
    "Idade",
    "Localização",
    "Cargo pretendido",
    "Habilidades",
    "Formações",
    "Email",
    "Telefone",
    "Portfólio",
]


PORTFOLIO_PATTERNS = [
    r"https?://(?:www\.)?behance\.net/[^\s)>\]]+",
    r"https?://(?:www\.)?dribbble\.com/[^\s)>\]]+",
    r"https?://(?:www\.)?github\.com/[^\s)>\]]+",
    r"https?://(?:www\.)?linkedin\.com/[^\s)>\]]+",
    r"https?://[^\s)>\]]+",
]


def extract_text_from_pdf(path: Path) -> str:
    try:
        reader = PdfReader(str(path))
        return "\n".join([(p.extract_text() or "") for p in reader.pages]).strip()
    except Exception as e:
        print(f"[REPAIR] Erro PDF {path.name}: {e}")
        return ""


def extract_text_from_docx(path: Path) -> str:
    try:
        doc = Document(str(path))
        return "\n".join([p.text for p in doc.paragraphs if p.text.strip()]).strip()
    except Exception as e:
        print(f"[REPAIR] Erro DOCX {path.name}: {e}")
        return ""


def extract_text_from_file(path: Path) -> str:
    ext = path.suffix.lower()

    if ext == ".pdf":
        return extract_text_from_pdf(path)

    if ext == ".docx":
        return extract_text_from_docx(path)

    if ext in {".txt", ".rtf"}:
        try:
            return path.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            return ""

    return ""


def find_portfolio(text: str) -> str:
    for pattern in PORTFOLIO_PATTERNS:
        match = re.search(pattern, text or "", flags=re.IGNORECASE)
        if match:
            return match.group(0).strip().rstrip(".,;")
    return ""


def normalize_for_search(value: Any) -> str:
    text = safe_str(value).lower()
    text = re.sub(r"[^a-z0-9@._+\-\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def phone_digits(value: Any) -> str:
    return re.sub(r"\D+", "", safe_str(value))


def list_candidate_files() -> list[Path]:
    files = []
    for base in CURRICULOS_FALLBACK_DIRS:
        if not base.exists():
            continue
        for p in base.rglob("*"):
            if p.is_file() and p.suffix.lower() in SUPPORTED_EXTS:
                files.append(p)
    return files


def find_resume_path(row_data: dict[str, Any]) -> Path | None:
    direct = safe_str(row_data.get("Caminho do currículo"))
    filename = safe_str(row_data.get("Nome do arquivo"))
    name = safe_str(row_data.get("Nome completo"))
    email = safe_str(row_data.get("Email"))
    phone = phone_digits(row_data.get("Telefone"))

    if direct:
        p = Path(direct)
        if p.exists():
            return p

    candidate_files = list_candidate_files()

    if filename:
        for p in candidate_files:
            if p.name.lower() == filename.lower():
                return p

        filename_norm = normalize_for_search(filename)
        for p in candidate_files:
            if filename_norm and filename_norm in normalize_for_search(p.name):
                return p

    if name:
        terms = [t for t in normalize_for_search(name).split() if len(t) >= 3]
        for p in candidate_files:
            pn = normalize_for_search(p.name)
            if terms and all(t in pn for t in terms[:2]):
                return p

    # fallback mais caro: procura email/telefone dentro dos currículos textuais
    if email or phone:
        for p in candidate_files:
            text = extract_text_from_file(p)
            if not text:
                continue

            text_norm = normalize_for_search(text)
            text_phone = phone_digits(text)

            if email and normalize_for_search(email) in text_norm:
                return p

            if phone and len(phone) >= 8 and phone in text_phone:
                return p

    return None


def extract_json_block(text: str) -> str:
    text = (text or "").strip().replace("```json", "").replace("```", "").strip()
    start = text.find("{")
    end = text.rfind("}")

    if start >= 0 and end > start:
        return text[start:end + 1]

    return text


def ai_extract_candidate_data(curriculum_text: str, filename: str = "") -> dict[str, str]:
    prompt = f"""
Extraia do currículo abaixo as informações e devolva APENAS JSON válido.

Campos obrigatórios:
{{
  "nome_completo": "",
  "idade": "",
  "localizacao": "",
  "cargo_pretendido": "",
  "habilidades": "",
  "formacoes": "",
  "email": "",
  "telefone": "",
  "portfolio": ""
}}

Regras:
- Não invente informações.
- Se não encontrar, deixe "".
- Para "nome_completo", use apenas nome de pessoa explicitamente presente no currículo.
- Para "idade", só preencha se houver idade ou data de nascimento explícita.
- Para "localizacao", use cidade/região/endereço explicitamente presente.
- Para "cargo_pretendido", use headline profissional, cargo atual/mais provável ou título principal explicitamente presente no currículo.
- "habilidades" deve ser string curta separada por "; ".
- "formacoes" deve ser string curta separada por "; ".
- "portfolio" deve ser link de Behance, site pessoal, GitHub, Dribbble, LinkedIn ou outro link profissional encontrado.
- Responda somente JSON puro.

Arquivo: {filename}

Currículo:
{curriculum_text[:15000]}
""".strip()

    payload = {
        "question": prompt,
        "use_rag": False,
        "model": STARIA_MODEL,
    }

    resp = requests.post(f"{STARIA_API_BASE}/ask", json=payload, timeout=180)
    resp.raise_for_status()

    answer = (resp.json().get("answer") or "").strip()
    parsed = json.loads(extract_json_block(answer))

    return {
        "Nome completo": safe_str(parsed.get("nome_completo")),
        "Idade": safe_str(parsed.get("idade")),
        "Localização": safe_str(parsed.get("localizacao")),
        "Cargo pretendido": safe_str(parsed.get("cargo_pretendido")),
        "Habilidades": safe_str(parsed.get("habilidades")),
        "Formações": safe_str(parsed.get("formacoes")),
        "Email": safe_str(parsed.get("email")),
        "Telefone": safe_str(parsed.get("telefone")),
        "Portfólio": safe_str(parsed.get("portfolio")),
    }


def set_if_blank(ws, header_map: dict[str, int], row_idx: int, field: str, value: Any) -> bool:
    if field not in header_map:
        return False

    if safe_str(ws.cell(row=row_idx, column=header_map[field]).value):
        return False

    if not safe_str(value):
        return False

    ws.cell(row=row_idx, column=header_map[field], value=value)
    return True


def repair_missing_fields(
    banco_path: Path = DEFAULT_BANCO_TALENTOS_XLSX,
    save_every: int = 10,
) -> dict[str, Any]:
    wb = load_workbook(banco_path)

    processed = 0
    repaired_rows = 0
    missing_resume = 0
    unsupported_resume = 0
    errors = 0

    for ws in wb.worksheets:
        header_map = build_header_map(ws)

        if not header_map:
            continue

        for row_idx in range(2, ws.max_row + 1):
            row_data = {
                h: ws.cell(row=row_idx, column=c).value
                for h, c in header_map.items()
            }

            if not any(safe_str(v) for v in row_data.values()):
                continue

            missing_fields = [
                f for f in FIELDS_TO_REPAIR
                if f in header_map and not safe_str(row_data.get(f))
            ]

            missing_path = "Caminho do currículo" in header_map and not safe_str(row_data.get("Caminho do currículo"))
            missing_filename = "Nome do arquivo" in header_map and not safe_str(row_data.get("Nome do arquivo"))

            if not missing_fields and not missing_path and not missing_filename:
                continue

            resume_path = find_resume_path(row_data)

            if not resume_path:
                missing_resume += 1
                continue

            if resume_path.suffix.lower() not in SUPPORTED_EXTS:
                unsupported_resume += 1
                continue

            text = extract_text_from_file(resume_path)

            if not text.strip():
                missing_resume += 1
                continue

            try:
                extracted = ai_extract_candidate_data(text, resume_path.name)
            except Exception as e:
                errors += 1
                print(f"[REPAIR] Erro IA linha {row_idx} aba {ws.title}: {e}")
                continue

            changed = False

            for field in missing_fields:
                value = extracted.get(field, "")

                if field == "Portfólio":
                    detected = value or find_portfolio(text) or str(resume_path)
                    changed = set_if_blank(ws, header_map, row_idx, field, detected) or changed
                    continue

                changed = set_if_blank(ws, header_map, row_idx, field, value) or changed

            changed = set_if_blank(ws, header_map, row_idx, "Caminho do currículo", str(resume_path)) or changed
            changed = set_if_blank(ws, header_map, row_idx, "Nome do arquivo", resume_path.name) or changed

            if changed:
                repaired_rows += 1

            processed += 1

            if save_every and processed % save_every == 0:
                wb.save(banco_path)
                print(f"[REPAIR] Parcial salvo. Processados: {processed} | Corrigidos: {repaired_rows}")

    wb.save(banco_path)

    return {
        "ok": True,
        "processed": processed,
        "repaired_rows": repaired_rows,
        "missing_resume": missing_resume,
        "unsupported_resume": unsupported_resume,
        "errors": errors,
        "path": str(banco_path),
    }


if __name__ == "__main__":
    result = repair_missing_fields()
    print(json.dumps(result, ensure_ascii=False, indent=2))