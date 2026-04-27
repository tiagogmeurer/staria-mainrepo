from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader
from docx import Document

from datasets.professional_profiles.matching_engine import score_candidate_against_profiles
from rh.talent_bank_workbook import (
    DEFAULT_BANCO_TALENTOS_XLSX,
    safe_str,
    build_header_map,
)


CURRICULOS_FALLBACK_DIR = Path(
    r"G:\Drives compartilhados\STARMKT\StarIA\banco_talentos\curriculos"
)

PORTFOLIO_PATTERNS = [
    r"https?://(?:www\.)?behance\.net/[^\s)>\]]+",
    r"https?://(?:www\.)?dribbble\.com/[^\s)>\]]+",
    r"https?://(?:www\.)?github\.com/[^\s)>\]]+",
    r"https?://(?:www\.)?linkedin\.com/[^\s)>\]]+",
    r"https?://[^\s)>\]]+",
]


def extract_text_from_pdf(path: Path) -> str:
    try:
        reader = PdfReader(str(path))
        parts = []
        for page in reader.pages:
            text = page.extract_text() or ""
            if text.strip():
                parts.append(text)
        return "\n".join(parts).strip()
    except Exception as e:
        print(f"[BACKFILL] Falha lendo PDF {path.name}: {e}")
        return ""


def extract_text_from_docx(path: Path) -> str:
    try:
        doc = Document(str(path))
        parts = []
        for p in doc.paragraphs:
            if p.text and p.text.strip():
                parts.append(p.text)
        return "\n".join(parts).strip()
    except Exception as e:
        print(f"[BACKFILL] Falha lendo DOCX {path.name}: {e}")
        return ""


def extract_text_from_file(path: Path) -> str:
    ext = path.suffix.lower()

    if ext == ".pdf":
        return extract_text_from_pdf(path)

    if ext == ".docx":
        return extract_text_from_docx(path)

    if ext in {".txt", ".rtf"}:
        try:
            return path.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            return ""

    return ""


def find_portfolio(text: str) -> str:
    for pattern in PORTFOLIO_PATTERNS:
        match = re.search(pattern, text or "", flags=re.IGNORECASE)
        if match:
            return match.group(0).strip().rstrip(".,;")
    return ""


def row_to_candidate(ws, row_idx: int, header_map: dict[str, int]) -> dict[str, Any]:
    data = {}

    for header, col in header_map.items():
        data[header] = ws.cell(row=row_idx, column=col).value

    return {
        "nome_completo": safe_str(data.get("Nome completo")),
        "idade": safe_str(data.get("Idade")),
        "localizacao": safe_str(data.get("Localização")),
        "cargo_pretendido": safe_str(data.get("Cargo pretendido")),
        "nivel": safe_str(data.get("Nível")),
        "portfolio": safe_str(data.get("Portfólio")),
        "habilidades": safe_str(data.get("Habilidades")),
        "formacoes": safe_str(data.get("Formações")),
        "email": safe_str(data.get("Email")),
        "telefone": safe_str(data.get("Telefone")),
        "observacoes": safe_str(data.get("Observações")),
        "Role ID sugerido": safe_str(data.get("Role ID sugerido")),
        "Título normalizado": safe_str(data.get("Título normalizado")),
        "Top 3 roles aderentes": safe_str(data.get("Top 3 roles aderentes")),
        "Resumo de aderência": safe_str(data.get("Resumo de aderência")),
        "Flags de risco": safe_str(data.get("Flags de risco")),
    }


def find_resume_path(row_data: dict[str, Any]) -> Path | None:
    direct_path = safe_str(row_data.get("Caminho do currículo"))
    filename = safe_str(row_data.get("Nome do arquivo"))

    if direct_path:
        p = Path(direct_path)
        if p.exists():
            return p

    if filename:
        p = CURRICULOS_FALLBACK_DIR / filename
        if p.exists():
            return p

        matches = list(CURRICULOS_FALLBACK_DIR.rglob(filename))
        if matches:
            return matches[0]

    candidate_name = safe_str(row_data.get("Nome completo"))
    if candidate_name and CURRICULOS_FALLBACK_DIR.exists():
        name_terms = [
            t.lower()
            for t in re.split(r"\s+", candidate_name)
            if len(t) >= 3
        ]

        for p in CURRICULOS_FALLBACK_DIR.rglob("*"):
            if not p.is_file() or p.suffix.lower() not in {".pdf", ".docx", ".txt", ".rtf"}:
                continue

            pn = p.name.lower()
            if name_terms and all(t in pn for t in name_terms[:2]):
                return p

    return None


def set_cell_if_exists(ws, header_map: dict[str, int], row_idx: int, header: str, value: Any, only_if_blank: bool = False):
    if header not in header_map:
        return

    col = header_map[header]

    if only_if_blank and safe_str(ws.cell(row=row_idx, column=col).value):
        return

    ws.cell(row=row_idx, column=col, value=value)


def backfill_scores(
    banco_path: Path = DEFAULT_BANCO_TALENTOS_XLSX,
    save_every: int = 25,
) -> dict[str, Any]:
    wb = load_workbook(banco_path)

    processed = 0
    updated = 0
    missing_resume = 0

    for ws in wb.worksheets:
        header_map = build_header_map(ws)

        if not header_map:
            continue

        for row_idx in range(2, ws.max_row + 1):
            row_data = {h: ws.cell(row=row_idx, column=c).value for h, c in header_map.items()}

            if not any(safe_str(v) for v in row_data.values()):
                continue

            resume_path = find_resume_path(row_data)
            curriculum_text = ""

            if resume_path:
                curriculum_text = extract_text_from_file(resume_path)
                set_cell_if_exists(ws, header_map, row_idx, "Caminho do currículo", str(resume_path), only_if_blank=True)
                set_cell_if_exists(ws, header_map, row_idx, "Nome do arquivo", resume_path.name, only_if_blank=True)
            else:
                missing_resume += 1

            candidate = row_to_candidate(ws, row_idx, header_map)

            match = score_candidate_against_profiles(
                candidate=candidate,
                curriculum_text=curriculum_text,
            )

            best = match.get("best") or {}
            top_matches = match.get("top_matches") or []

            top_matches_text = "; ".join(
                f"{m.get('title', '')} ({m.get('nota', m.get('score_pct', 0))})"
                for m in top_matches
                if m.get("title")
            )

            set_cell_if_exists(ws, header_map, row_idx, "Nota", match.get("nota", 0))
            set_cell_if_exists(ws, header_map, row_idx, "Role ID sugerido", best.get("role_id", ""))
            set_cell_if_exists(ws, header_map, row_idx, "Título normalizado", best.get("title", ""))
            set_cell_if_exists(ws, header_map, row_idx, "Top 3 roles aderentes", top_matches_text)
            set_cell_if_exists(ws, header_map, row_idx, "Resumo de aderência", match.get("summary", ""))
            set_cell_if_exists(ws, header_map, row_idx, "Flags de risco", "; ".join(best.get("gaps", [])[:6]))

            current_portfolio = safe_str(row_data.get("Portfólio"))
            detected_portfolio = find_portfolio(curriculum_text)

            if detected_portfolio:
                set_cell_if_exists(ws, header_map, row_idx, "Portfólio", detected_portfolio, only_if_blank=True)
            elif resume_path:
                set_cell_if_exists(ws, header_map, row_idx, "Portfólio", str(resume_path), only_if_blank=True)

            processed += 1
            updated += 1

            if save_every and processed % save_every == 0:
                wb.save(banco_path)
                print(f"[BACKFILL] Parcial salvo. Processados: {processed}")

    wb.save(banco_path)

    return {
        "ok": True,
        "processed": processed,
        "updated": updated,
        "missing_resume": missing_resume,
        "path": str(banco_path),
    }


if __name__ == "__main__":
    result = backfill_scores()
    print(json.dumps(result, ensure_ascii=False, indent=2))