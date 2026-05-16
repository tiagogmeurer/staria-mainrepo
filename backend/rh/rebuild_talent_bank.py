from pathlib import Path
from openpyxl import load_workbook

from rh.talent_bank_workbook import append_candidate_record, normalize_role_to_sheet_name

BANCO = Path(r"G:\Drives compartilhados\STARMKT\StarIA\banco_talentos\banco_talentos.xlsx")

def rebuild():
    wb = load_workbook(BANCO)

    all_rows = []

    for ws in wb.worksheets:
        headers = [c.value for c in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            all_rows.append(data)

    # 🔥 limpa tudo
    for ws in wb.worksheets:
        ws.delete_rows(2, ws.max_row)

    wb.save(BANCO)

    seen = set()

    for row in all_rows:
        key = (row.get("Email"), row.get("Telefone"))

        if key in seen:
            continue
        seen.add(key)

        role = row.get("Cargo pretendido", "")
        sheet = normalize_role_to_sheet_name(role)

        row["Cargo pretendido"] = sheet.title()

        append_candidate_record(
            values=row,
            banco_path=BANCO,
            target_sheet=sheet,
        )

    print("✔ Banco reconstruído com sucesso")


if __name__ == "__main__":
    rebuild()    
