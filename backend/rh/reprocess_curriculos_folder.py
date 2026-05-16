from __future__ import annotations

import json
import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from datasets.professional_profiles.matching_engine import score_candidate_against_profiles
from rh.talent_bank_workbook import (
    DEFAULT_BANCO_TALENTOS_XLSX,
    DEFAULT_REFINED_XLSX,
    CANONICAL_SHEETS,
    append_candidate_record,
    backup_workbook,
    build_header_map,
    ensure_bank_workbook_structure,
    normalize_role_to_sheet_name,
    safe_str,
    sheet_display_title,
)

from workers.gmail_worker import (
    extract_candidate_data_with_ai,
    extract_email_regex,
    extract_phone_regex,
    extract_portfolio_regex,
    extract_text_from_file,
    guess_job_title_from_text,
)


CURRICULOS_DIR = Path(
    r"G:\Drives compartilhados\STARMKT\StarIA\banco_talentos\curriculos"
)

REJECTED_DIR = CURRICULOS_DIR / "_rejeitados_nota_menor_50"

SUPPORTED_EXTS = {".pdf", ".docx", ".txt", ".rtf"}
MIN_SCORE = 50


def clear_all_data_rows(path: Path) -> None:
    wb = load_workbook(path)

    for sheet_name in CANONICAL_SHEETS:
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)

        ws = wb[sheet_name]

        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

    wb.save(path)


def duplicate_key(values: dict[str, Any]) -> str:
    email = safe_str(values.get("Email")).lower()
    phone = "".join(ch for ch in safe_str(values.get("Telefone")) if ch.isdigit())
    name = safe_str(values.get("Nome completo")).lower()
    role = safe_str(values.get("Cargo pretendido")).lower()
    filename = safe_str(values.get("Nome do arquivo")).lower()

    if email and role:
        return f"email:{email}|role:{role}"

    if phone and len(phone) >= 8 and role:
        return f"phone:{phone}|role:{role}"

    if name and role:
        return f"name:{name}|role:{role}"

    return f"file:{filename}"


def move_rejected(file_path: Path) -> str:
    REJECTED_DIR.mkdir(parents=True, exist_ok=True)

    target = REJECTED_DIR / file_path.name

    if target.exists():
        target = REJECTED_DIR / f"{file_path.stem}_{file_path.stat().st_mtime_ns}{file_path.suffix}"

    shutil.move(str(file_path), str(target))
    return str(target)


def reprocess_curriculos_folder(
    banco_path: Path = DEFAULT_BANCO_TALENTOS_XLSX,
    refined_path: Path = DEFAULT_REFINED_XLSX,
    curriculos_dir: Path = CURRICULOS_DIR,
    min_score: int = MIN_SCORE,
) -> dict[str, Any]:
    banco_path = Path(banco_path)
    curriculos_dir = Path(curriculos_dir)

    if not curriculos_dir.exists():
        raise RuntimeError(f"Pasta de currículos não encontrada: {curriculos_dir}")

    backup = backup_workbook(banco_path)

    ensure_bank_workbook_structure(
        banco_path=banco_path,
        refined_path=refined_path,
        create_backup=False,
        redistribute_existing=True,
    )

    clear_all_data_rows(banco_path)

    files = [
        p for p in curriculos_dir.rglob("*")
        if p.is_file()
        and p.suffix.lower() in SUPPORTED_EXTS
        and "_rejeitados_nota_menor_50" not in str(p)
    ]

    files = files[:20]
    print(f"[REPROCESS] MODO TESTE: {len(files)} arquivos")

    processed = 0
    inserted = 0
    rejected = 0
    duplicates = 0
    errors = 0

    seen: set[str] = set()

    for file_path in files:
        processed += 1
        print(f"\n[REPROCESS] {processed}/{len(files)} - {file_path.name}")

        try:
            curriculum_text = extract_text_from_file(file_path)

            extracted = extract_candidate_data_with_ai(file_path)

            if not extracted.get("email"):
                extracted["email"] = extract_email_regex(curriculum_text)

            if not extracted.get("telefone"):
                extracted["telefone"] = extract_phone_regex(curriculum_text)

            if not extracted.get("portfolio"):
                extracted["portfolio"] = extract_portfolio_regex(curriculum_text)

            if not extracted.get("cargo_pretendido"):
                extracted["cargo_pretendido"] = guess_job_title_from_text(curriculum_text)

            filename_role = file_path.stem

            requested_role = (
                filename_role
                or extracted.get("cargo_pretendido")
                or guess_job_title_from_text(curriculum_text)
                or ""
            )

            target_sheet = normalize_role_to_sheet_name(requested_role)

            # Força o cargo canônico para registro e matching
            extracted["cargo_pretendido"] = sheet_display_title(target_sheet)

            match = score_candidate_against_profiles(
                candidate=extracted,
                curriculum_text=curriculum_text,
                requested_role=requested_role,
                extra_query=f"{requested_role} {sheet_display_title(target_sheet)}",
            )

            nota = int(match.get("nota", 0) or 0)
            best = match.get("best") or {}
            top_matches = match.get("top_matches") or []

            if nota < min_score:
                moved_to = move_rejected(file_path)
                print(f"[REPROCESS] Rejeitado nota {nota}. Movido para: {moved_to}")
                rejected += 1
                continue

            top_matches_text = "; ".join(
                f"{m.get('title', '')} ({m.get('nota', m.get('score_pct', 0))})"
                for m in top_matches
                if m.get("title")
            )

            values = {
                "Nota": nota,
                "Nome completo": extracted.get("nome_completo", ""),
                "Localização": extracted.get("localizacao", ""),
                "Cargo pretendido": sheet_display_title(target_sheet),
                "Nível": "",
                "Portfólio": extracted.get("portfolio") or str(file_path),
                "Habilidades": extracted.get("habilidades", ""),
                "Formação": extracted.get("formacoes", ""),
                "Email": extracted.get("email", ""),
                "Telefone": extracted.get("telefone", ""),
                "Caminho do currículo": str(file_path),
                "Nome do arquivo": file_path.name,
                "Data de entrada": "",
                "Origem": "reprocess_curriculos_folder",
                "Remetente do email": "",
                "Status": "Banco de talentos",
                "Observações": "",
                "Role ID sugerido": best.get("role_id", ""),
                "Título normalizado": best.get("title", ""),
                "Top 3 roles aderentes": top_matches_text,
                "Resumo de aderência": match.get("summary", ""),
                "Flags de risco": "; ".join(best.get("gaps", [])[:6]),
            }

            key = duplicate_key(values)
            if key in seen:
                print("[REPROCESS] Duplicado ignorado:", values.get("Nome completo") or file_path.name)
                duplicates += 1
                continue

            seen.add(key)

            candidate_id = append_candidate_record(
                values=values,
                banco_path=banco_path,
                refined_path=refined_path,
                target_sheet=target_sheet,
            )

            if candidate_id == "DUPLICADO":
                duplicates += 1
                continue

            print(
                f"[REPROCESS] Inserido {candidate_id} | "
                f"Aba: {target_sheet} | Nota: {nota} | "
                f"Nome: {values.get('Nome completo') or '(sem nome)'}"
            )

            inserted += 1

        except Exception as e:
            print(f"[REPROCESS] Erro em {file_path.name}: {e}")
            errors += 1

    return {
        "ok": True,
        "backup": str(backup) if backup else "",
        "processed": processed,
        "inserted": inserted,
        "rejected_under_50": rejected,
        "duplicates": duplicates,
        "errors": errors,
        "banco_path": str(banco_path),
        "curriculos_dir": str(curriculos_dir),
    }


if __name__ == "__main__":
    result = reprocess_curriculos_folder()
    print(json.dumps(result, ensure_ascii=False, indent=2))
