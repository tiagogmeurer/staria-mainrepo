from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from rh.talent_bank_workbook import (
    DEFAULT_BANCO_TALENTOS_XLSX,
    safe_str,
    build_header_map,
    normalize_header,
    choose_sheet_name,
)

MIN_SCORE = 50


def norm(value: Any) -> str:
    text = safe_str(value).lower()
    text = re.sub(r"\s+", " ", text).strip()
    return text


def phone_digits(value: Any) -> str:
    return re.sub(r"\D+", "", safe_str(value))


def detect_seniority(row: dict[str, Any]) -> str:
    text = norm(
        " | ".join(
            [
                safe_str(row.get("Cargo pretendido")),
                safe_str(row.get("Título normalizado")),
                safe_str(row.get("Habilidades")),
                safe_str(row.get("Resumo de aderência")),
                safe_str(row.get("Observações")),
                safe_str(row.get("Top 3 roles aderentes")),
            ]
        )
    )

    if any(x in text for x in ["sênior", "senior", " sr", " sr."]):
        return "Sênior"
    if "pleno" in text:
        return "Pleno"
    if any(x in text for x in ["júnior", "junior", " jr", " jr."]):
        return "Júnior"

    role_id = norm(row.get("Role ID sugerido"))
    title = norm(row.get("Título normalizado"))

    if "senior" in role_id or "sênior" in title or "senior" in title:
        return "Sênior"

    return ""


def duplicate_key(row: dict[str, Any]) -> str:
    email = norm(row.get("Email"))
    phone = phone_digits(row.get("Telefone"))
    name = norm(row.get("Nome completo"))
    role = norm(row.get("Cargo pretendido") or row.get("Título normalizado"))
    filename = norm(row.get("Nome do arquivo"))

    if email and role:
        return f"email:{email}|role:{role}"
    if phone and len(phone) >= 8 and role:
        return f"phone:{phone}|role:{role}"
    if name and role:
        return f"name:{name}|role:{role}"
    if filename:
        return f"file:{filename}"

    return f"row:{name}|{phone}|{email}|{role}|{filename}"


def get_score(row: dict[str, Any]) -> float:
    raw = safe_str(row.get("Nota"))
    if not raw:
        return 0.0

    raw = raw.replace(",", ".")
    try:
        return float(raw)
    except ValueError:
        return 0.0


def extract_all_rows(wb) -> list[dict[str, Any]]:
    rows = []

    for ws in wb.worksheets:
        headers = [normalize_header(c.value) for c in ws[1]]

        for row_idx in range(2, ws.max_row + 1):
            row_data = {"_source_sheet": ws.title}
            has_any = False

            for col_idx, header in enumerate(headers, start=1):
                if not header:
                    continue
                value = ws.cell(row=row_idx, column=col_idx).value
                row_data[header] = value
                if value not in (None, ""):
                    has_any = True

            if has_any:
                rows.append(row_data)

    return rows


def clear_data_rows(ws):
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def write_row(ws, row_data: dict[str, Any]):
    header_map = build_header_map(ws)
    ws.append([row_data.get(header, "") for header in header_map.keys()])


def cleanup_talent_bank(
    banco_path: Path = DEFAULT_BANCO_TALENTOS_XLSX,
    min_score: int = MIN_SCORE,
) -> dict[str, Any]:
    wb = load_workbook(banco_path)
    all_rows = extract_all_rows(wb)

    before = len(all_rows)

    # 1) remove abaixo da nota de corte
    rows_score_ok = [r for r in all_rows if get_score(r) >= min_score]
    removed_low_score = before - len(rows_score_ok)

    # 2) preenche senioridade quando possível
    seniority_filled = 0
    for row in rows_score_ok:
        if not safe_str(row.get("Nível")):
            seniority = detect_seniority(row)
            if seniority:
                row["Nível"] = seniority
                seniority_filled += 1

    # 3) remove duplicatas, mantendo a melhor nota
    best_by_key: dict[str, dict[str, Any]] = {}

    for row in rows_score_ok:
        key = duplicate_key(row)
        current = best_by_key.get(key)

        if current is None:
            best_by_key[key] = row
            continue

        if get_score(row) > get_score(current):
            best_by_key[key] = row

    deduped_rows = list(best_by_key.values())
    removed_duplicates = len(rows_score_ok) - len(deduped_rows)

    # 4) limpa dados das abas e redistribui corretamente
    for ws in wb.worksheets:
        clear_data_rows(ws)

    sheet_names = wb.sheetnames
    written = 0

    for row in deduped_rows:
        sheet_name = choose_sheet_name(sheet_names, row)
        ws = wb[sheet_name]
        write_row(ws, row)
        written += 1

    wb.save(banco_path)

    return {
        "ok": True,
        "path": str(banco_path),
        "before": before,
        "removed_low_score": removed_low_score,
        "removed_duplicates": removed_duplicates,
        "seniority_filled": seniority_filled,
        "after": written,
        "min_score": min_score,
    }


if __name__ == "__main__":
    result = cleanup_talent_bank()
    print(json.dumps(result, ensure_ascii=False, indent=2))