from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook
from pypdf import PdfReader
from docx import Document

from rh.talent_bank_workbook import (
    DEFAULT_BANCO_TALENTOS_XLSX,
    safe_str,
    build_header_map,
)

STARIA_API_BASE = "http://127.0.0.1:8088"
STARIA_MODEL = "star-llama:latest"

CURRICULOS_DIRS = [
    Path(r"G:\Drives compartilhados\STARMKT\StarIA\curriculos"),
    Path(r"G:\Drives compartilhados\STARMKT\StarIA\banco_talentos\curriculos"),
]


def extract_text_from_pdf(path: Path) -> str:
    try:
        reader = PdfReader(str(path))
        return "\n".join([(p.extract_text() or "") for p in reader.pages]).strip()
    except Exception:
        return ""


def extract_text_from_docx(path: Path) -> str:
    try:
        doc = Document(str(path))
        return "\n".join([p.text for p in doc.paragraphs if p.text.strip()]).strip()
    except Exception:
        return ""


def extract_text_from_file(path: Path) -> str:
    ext = path.suffix.lower()
    if ext == ".pdf":
        return extract_text_from_pdf(path)
    if ext == ".docx":
        return extract_text_from_docx(path)
    if ext in {".txt", ".rtf"}:
        return path.read_text(encoding="utf-8", errors="ignore")
    return ""


def find_resume_path(row: dict[str, Any]) -> Path | None:
    direct = safe_str(row.get("Caminho do currículo"))
    filename = safe_str(row.get("Nome do arquivo"))

    if direct:
        p = Path(direct)
        if p.exists():
            return p

    for base in CURRICULOS_DIRS:
        if not base.exists():
            continue

        if filename:
            p = base / filename
            if p.exists():
                return p

            matches = list(base.rglob(filename))
            if matches:
                return matches[0]

    return None


def extract_json_block(text: str) -> str:
    text = (text or "").strip().replace("```json", "").replace("```", "").strip()
    start = text.find("{")
    end = text.rfind("}")
    if start >= 0 and end > start:
        return text[start:end + 1]
    return text


def level_from_years(years: float) -> str:
    if years >= 5:
        return "Sênior"
    if years >= 2:
        return "Pleno"
    return "Júnior"


def estimate_years_by_regex(text: str) -> tuple[float | None, str]:
    """
    Heurística objetiva:
    - captura frases tipo "5 anos de experiência"
    - captura anos em períodos profissionais tipo 2019 - 2024
    """
    t = text or ""
    current_year = datetime.now().year

    explicit_years = []
    for m in re.finditer(r"(\d{1,2})\s*(?:\+)?\s+anos?\s+(?:de\s+)?experi", t, flags=re.I):
        try:
            explicit_years.append(float(m.group(1)))
        except Exception:
            pass

    if explicit_years:
        y = max(explicit_years)
        return y, f"Menção explícita no currículo: {int(y)} anos de experiência"

    ranges = []
    pattern = r"\b(20\d{2}|19\d{2})\b\s*(?:-|–|—|a|até|ate)\s*\b(20\d{2}|19\d{2}|atual|presente|hoje)\b"
    for m in re.finditer(pattern, t, flags=re.I):
        start_raw = m.group(1)
        end_raw = m.group(2).lower()

        start = int(start_raw)
        end = current_year if end_raw in {"atual", "presente", "hoje"} else int(end_raw)

        if 1980 <= start <= current_year and start <= end <= current_year + 1:
            ranges.append(max(0, end - start))

    if ranges:
        y = max(ranges)
        return float(y), f"Período profissional identificado no currículo: aproximadamente {int(y)} anos"

    years = [int(y) for y in re.findall(r"\b(20\d{2}|19\d{2})\b", t)]
    years = [y for y in years if 1980 <= y <= current_year]

    if years:
        y = max(0, current_year - min(years))
        if y > 0:
            return float(y), f"Anos no currículo sugerem trajetória desde {min(years)}"

    return None, ""


def ai_estimate_years(curriculum_text: str, filename: str) -> dict[str, Any]:
    prompt = f"""
Analise o currículo abaixo e estime os anos de experiência profissional do candidato.

Responda APENAS JSON válido:
{{
  "anos_experiencia": null,
  "evidencia": ""
}}

Regras:
- Use somente o texto do currículo.
- NÃO use o título da vaga.
- NÃO invente.
- Se não houver datas ou menção de experiência, retorne null.
- "evidencia" deve conter o trecho/dado do currículo usado para estimar.

Arquivo: {filename}

Currículo:
{curriculum_text[:15000]}
""".strip()

    payload = {
        "question": prompt,
        "use_rag": False,
        "model": STARIA_MODEL,
    }

    resp = requests.post(f"{STARIA_API_BASE}/ask", json=payload, timeout=180)
    resp.raise_for_status()

    parsed = json.loads(extract_json_block(resp.json().get("answer", "")))

    years = parsed.get("anos_experiencia")
    evidence = safe_str(parsed.get("evidencia"))

    try:
        if years is None or years == "":
            return {"years": None, "evidence": ""}
        return {"years": float(years), "evidence": evidence}
    except Exception:
        return {"years": None, "evidence": ""}


def repair_candidate_seniority(
    banco_path: Path = DEFAULT_BANCO_TALENTOS_XLSX,
    clear_existing: bool = True,
    save_every: int = 20,
) -> dict[str, Any]:
    wb = load_workbook(banco_path)

    processed = 0
    updated = 0
    defaulted_junior = 0
    no_resume_or_text = 0
    errors = 0

    for ws in wb.worksheets:
        header_map = build_header_map(ws)

        if "Nível" not in header_map:
            continue

        nivel_col = header_map["Nível"]
        obs_col = header_map.get("Observações")

        for row_idx in range(2, ws.max_row + 1):
            row = {h: ws.cell(row=row_idx, column=c).value for h, c in header_map.items()}

            if not any(safe_str(v) for v in row.values()):
                continue

            if clear_existing:
                ws.cell(row=row_idx, column=nivel_col, value="")

            resume_path = find_resume_path(row)
            if not resume_path:
                no_resume_or_text += 1
                nivel = "Júnior"
                evidence = "Sem currículo localizado para estimativa; classificado conservadoramente como Júnior."
            else:
                text = extract_text_from_file(resume_path)

                if not text.strip():
                    no_resume_or_text += 1
                    nivel = "Júnior"
                    evidence = "Currículo sem texto extraível; classificado conservadoramente como Júnior."
                else:
                    years, evidence = estimate_years_by_regex(text)

                    if years is None:
                        try:
                            ai_result = ai_estimate_years(text, resume_path.name)
                            years = ai_result["years"]
                            evidence = ai_result["evidence"]
                        except Exception as e:
                            errors += 1
                            print(f"[SENIORITY] Erro IA linha {row_idx} aba {ws.title}: {e}")
                            years = None
                            evidence = ""

                    if years is None:
                        nivel = "Júnior"
                        evidence = "Não foi encontrada evidência suficiente de tempo de experiência; classificado conservadoramente como Júnior."
                        defaulted_junior += 1
                    else:
                        nivel = level_from_years(float(years))

            ws.cell(row=row_idx, column=nivel_col, value=nivel)

            if obs_col:
                current_obs = safe_str(ws.cell(row=row_idx, column=obs_col).value)
                note = f"Senioridade: {nivel}. Critério: {evidence}"
                if note not in current_obs:
                    ws.cell(
                        row=row_idx,
                        column=obs_col,
                        value=(current_obs + " | " + note).strip(" |"),
                    )

            updated += 1
            processed += 1

            if save_every and processed % save_every == 0:
                wb.save(banco_path)
                print(f"[SENIORITY] Parcial salvo. Processados: {processed} | Atualizados: {updated}")

    wb.save(banco_path)

    return {
        "ok": True,
        "path": str(banco_path),
        "processed": processed,
        "updated": updated,
        "defaulted_junior": defaulted_junior,
        "no_resume_or_text": no_resume_or_text,
        "errors": errors,
    }


if __name__ == "__main__":
    result = repair_candidate_seniority()
    print(json.dumps(result, ensure_ascii=False, indent=2))