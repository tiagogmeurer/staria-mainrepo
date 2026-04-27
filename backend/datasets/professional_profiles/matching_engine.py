from __future__ import annotations

import json
import os
import re
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from datasets.professional_profiles.loader import load_profiles
from datasets.professional_profiles.normalizer import normalize_role_query, resolve_role_id
from datasets.professional_profiles.schema import ProfessionalProfile


DEFAULT_DRIVE_ROOT = os.getenv(
    "STARIA_DRIVE_ROOT",
    r"G:\Drives compartilhados\STARMKT\StarIA",
)

BANCO_TALENTOS_XLSX = Path(
    os.getenv(
        "STARIA_TALENTS_XLSX",
        str(Path(DEFAULT_DRIVE_ROOT) / "banco_talentos" / "banco_talentos.xlsx"),
    )
)

BASE_DIR = Path(__file__).resolve().parent
MATCHING_RULES_JSON = BASE_DIR / "matching_rules.json"


DEFAULT_MATCHING_RULES = {
    "default_weights": {
        "role_title": 20,
        "profile_required_skills": 20,
        "priority_specs": 40,
        "seniority": 10,
        "location": 10,
    },
    "global_terms": {
        "varejo": [
            "varejo", "retail", "atacarejo", "supermercado", "atacado",
            "assaí", "assai", "carrefour", "extra", "pão de açúcar",
            "pao de acucar", "mercado", "loja alimentar", "alimentar"
        ],
        "farma": [
            "farma", "farmácia", "farmacia", "drogaria", "drogarias",
            "pharma", "saúde", "saude", "bem estar", "bem-estar",
            "categoria regulada", "categorias reguladas"
        ],
        "ia_comunicacao": [
            "ia", "inteligência artificial", "inteligencia artificial",
            "chatgpt", "midjourney", "firefly", "adobe firefly",
            "runway", "sora", "kling", "kling 3.0", "higgsfield",
            "heygen", "stable diffusion", "comfyui", "comfy ui",
            "leonardo ai", "flux", "generative ai", "genai"
        ],
        "adobe": [
            "photoshop", "illustrator", "indesign", "after effects",
            "premiere", "adobe", "pacote adobe", "creative suite"
        ],
        "performance_ads": [
            "analytics", "google analytics", "ga4", "google ads",
            "meta ads", "facebook ads", "instagram ads", "tiktok ads",
            "dashboard", "looker", "data studio", "power bi"
        ],
        "marketplace": [
            "marketplace", "market place", "ecommerce", "e-commerce",
            "digital commerce", "performance digital", "programática",
            "programatica", "varejo digital"
        ],
        "alto_volume": [
            "alto volume", "alta cadência", "alta cadencia", "volume alto",
            "escala", "múltiplas entregas", "multiplas entregas",
            "velocidade", "alta demanda"
        ],
        "stakeholder": [
            "stakeholder", "stakeholder engagement", "comunicação corporativa",
            "comunicacao corporativa", "cliente interno", "clientes internos",
            "interface", "relacionamento", "briefing"
        ],
        "encarte": [
            "encarte", "tablóide", "tabloide", "ofertas",
            "material promocional", "varejo promocional", "diagramação",
            "diagramacao", "indesign"
        ],
    },
    "role_rules": {
        "estrategista_senior_planejamento": [
            {"label": "Experiência em farma", "terms_key": "farma", "points": 10},
            {"label": "Experiência em varejo", "terms_key": "varejo", "points": 10},
            {"label": "IA aplicada a marca/comunicação", "terms_key": "ia_comunicacao", "points": 20},
        ],
        "coordenador_conteudo": [
            {"label": "Experiência em farma", "terms_key": "farma", "points": 10},
            {"label": "Experiência em varejo", "terms_key": "varejo", "points": 10},
            {"label": "IA aplicada a marca/comunicação", "terms_key": "ia_comunicacao", "points": 20},
        ],
        "performance_growth_planejamento": [
            {"label": "Analytics, Google Ads, Meta Ads ou TikTok Ads", "terms_key": "performance_ads", "points": 20},
            {"label": "Experiência em farma", "terms_key": "farma", "points": 8},
            {"label": "Experiência em varejo", "terms_key": "varejo", "points": 8},
            {"label": "IA aplicada a marca/comunicação", "terms_key": "ia_comunicacao", "points": 14},
        ],
        "diretor_arte_senior_campanhas": [
            {"label": "Experiência em farma", "terms_key": "farma", "points": 8},
            {"label": "Experiência em varejo", "terms_key": "varejo", "points": 8},
            {"label": "Domínio em Adobe/Photoshop/Illustrator", "terms_key": "adobe", "points": 12},
            {"label": "IA aplicada a criação e comunicação", "terms_key": "ia_comunicacao", "points": 12},
        ],
        "diretor_arte_senior_branding_produto": [
            {"label": "Experiência em farma", "terms_key": "farma", "points": 10},
            {"label": "Experiência em varejo", "terms_key": "varejo", "points": 10},
            {"label": "IA aplicada a marca/comunicação", "terms_key": "ia_comunicacao", "points": 20},
        ],
        "diretor_arte_senior_digital": [
            {"label": "Performance/marketplace digital no varejo", "terms_key": "marketplace", "points": 16},
            {"label": "Materiais com inteligência artificial", "terms_key": "ia_comunicacao", "points": 14},
            {"label": "Ambiente de alto volume", "terms_key": "alto_volume", "points": 10},
            {"label": "Experiência em varejo", "terms_key": "varejo", "points": 8},
        ],
        "motion_designer": [
            {
                "label": "IA para motion, Kling ou Higgsfield",
                "terms": ["kling", "kling 3.0", "higgsfield", "ia", "inteligência artificial", "inteligencia artificial"],
                "points": 18,
            },
            {
                "label": "Animações, logos animados ou interfaces dinâmicas",
                "terms": ["animação", "animacao", "logo animado", "logos animados", "ui", "interface", "interfaces dinâmicas", "interfaces dinamicas", "motion", "reels"],
                "points": 14,
            },
            {"label": "After Effects/Premiere", "terms": ["after effects", "premiere"], "points": 8},
        ],
        "redator_digital": [
            {
                "label": "Categorias reguladas",
                "terms": ["categoria regulada", "categorias reguladas", "saúde", "saude", "bem-estar", "bem estar"],
                "points": 20,
            },
            {"label": "Varejo farma", "terms_key": "farma", "points": 12},
            {"label": "Varejo", "terms_key": "varejo", "points": 8},
        ],
        "atendimento_senior": [
            {"label": "Comunicação corporativa / stakeholder engagement", "terms_key": "stakeholder", "points": 18},
            {"label": "Atendimento a varejo alimentar", "terms_key": "varejo", "points": 12},
            {"label": "Atendimento a farma", "terms_key": "farma", "points": 12},
        ],
        "diagramador_ofertas": [
            {"label": "Encarte, tabloide ou ofertas", "terms_key": "encarte", "points": 18},
            {"label": "Experiência em varejo", "terms_key": "varejo", "points": 12},
            {"label": "Experiência em farma", "terms_key": "farma", "points": 12},
            {"label": "InDesign/Illustrator", "terms": ["indesign", "illustrator"], "points": 8},
        ],
    },
}


def _strip_accents(text: str) -> str:
    text = text or ""
    return "".join(
        c for c in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(c)
    )


def normalize_text(text: Any) -> str:
    text = _strip_accents(str(text or "")).lower().strip()
    text = re.sub(r"[^a-z0-9\s@+./_-]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def singularize_pt(text: Any) -> str:
    words = []
    for w in normalize_text(text).split():
        if len(w) > 3 and w.endswith("s"):
            words.append(w[:-1])
        else:
            words.append(w)
    return " ".join(words).strip()


def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def load_matching_rules() -> dict:
    if MATCHING_RULES_JSON.exists():
        try:
            data = json.loads(MATCHING_RULES_JSON.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                merged = DEFAULT_MATCHING_RULES.copy()
                merged.update(data)
                return merged
        except Exception as e:
            print("[MATCHING] Falha ao ler matching_rules.json:", repr(e))

    return DEFAULT_MATCHING_RULES


def _contains_any(text: Any, values: list[str]) -> bool:
    normalized = singularize_pt(text)
    for value in values:
        needle = singularize_pt(value)
        if needle and needle in normalized:
            return True
    return False


def _count_hits(text: Any, values: list[str]) -> int:
    normalized = singularize_pt(text)
    hits = 0
    seen = set()

    for value in values:
        needle = singularize_pt(value)
        if needle and needle in normalized and needle not in seen:
            hits += 1
            seen.add(needle)

    return hits


def _matched_terms(text: Any, values: list[str]) -> list[str]:
    normalized = singularize_pt(text)
    out = []

    for value in values:
        needle = singularize_pt(value)
        if needle and needle in normalized:
            out.append(value)

    return sorted(set(out))


def _token_score(text: str, targets: list[str]) -> tuple[float, list[str]]:
    haystack = singularize_pt(text)
    matched: list[str] = []

    valid_targets = []
    for item in targets:
        needle = singularize_pt(item)
        if not needle:
            continue
        valid_targets.append(item)
        if needle in haystack:
            matched.append(item)

    if not valid_targets:
        return 0.0, []

    score = len(set(matched)) / len(valid_targets)
    return min(score, 1.0), sorted(set(matched))


def load_banco_talentos_rows(
    path: Path = BANCO_TALENTOS_XLSX,
    limit: int = 5000,
) -> list[dict[str, Any]]:
    if not path.exists():
        return []

    wb = load_workbook(path, data_only=True)

    rows: list[dict[str, Any]] = []

    for ws in wb.worksheets:
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            item: dict[str, Any] = {"_sheet": ws.title}
            has_any = False

            for idx, val in enumerate(row):
                key = headers[idx] if idx < len(headers) else f"col_{idx + 1}"
                item[key] = val
                if val not in (None, ""):
                    has_any = True

            if has_any:
                rows.append(item)

            if len(rows) >= limit:
                return rows

    return rows


def candidate_name(row: dict[str, Any]) -> str:
    return (
        _safe_str(row.get("Nome completo"))
        or _safe_str(row.get("Nome"))
        or _safe_str(row.get("Candidato"))
        or "Sem nome"
    )


def candidate_role_text(row: dict[str, Any]) -> str:
    return (
        _safe_str(row.get("Cargo pretendido"))
        or _safe_str(row.get("Cargo"))
        or _safe_str(row.get("Função"))
        or _safe_str(row.get("Funcao"))
        or _safe_str(row.get("Título normalizado"))
    )


def candidate_seniority_text(row: dict[str, Any]) -> str:
    return (
        _safe_str(row.get("Nível"))
        or _safe_str(row.get("Senioridade"))
        or _safe_str(row.get("Nivel"))
    )


def candidate_location_text(row: dict[str, Any]) -> str:
    return (
        _safe_str(row.get("Localização"))
        or _safe_str(row.get("Localizacao"))
        or _safe_str(row.get("Cidade"))
        or _safe_str(row.get("Região"))
        or _safe_str(row.get("Regiao"))
    )


def candidate_contact_text(row: dict[str, Any]) -> dict[str, str]:
    return {
        "email": _safe_str(row.get("Email") or row.get("E-mail")),
        "telefone": _safe_str(row.get("Telefone") or row.get("Celular") or row.get("WhatsApp")),
        "curriculo": _safe_str(row.get("Caminho do currículo") or row.get("Currículo") or row.get("Curriculo")),
    }


def candidate_resume_text(row: dict[str, Any], curriculum_text: str = "") -> str:
    preferred_keys = [
        "Resumo",
        "Observações",
        "Observacoes",
        "Skills",
        "Habilidades",
        "Competências",
        "Competencias",
        "Ferramentas",
        "Experiência",
        "Experiencia",
        "Formações",
        "Formacao",
        "Formação",
        "Portfólio",
        "Portfolio",
        "Top 3 roles aderentes",
        "Resumo de aderência",
        "Flags de risco",
    ]

    parts = [
        candidate_name(row),
        candidate_role_text(row),
        candidate_seniority_text(row),
        candidate_location_text(row),
        curriculum_text,
    ]

    for key in preferred_keys:
        value = _safe_str(row.get(key))
        if value:
            parts.append(value)

    for key, value in row.items():
        if key in preferred_keys or str(key).startswith("_"):
            continue
        value_str = _safe_str(value)
        if value_str and value_str not in parts:
            parts.append(value_str)

    return " | ".join([p for p in parts if p])


def _parse_query_constraints(query: str) -> dict[str, str]:
    q = normalize_text(query)

    seniority = ""
    if "senior" in q:
        seniority = "senior"
    elif "pleno" in q:
        seniority = "pleno"
    elif "junior" in q:
        seniority = "junior"

    location = ""
    patterns = [
        r"proxim[oa]\s+a\s+([a-z0-9\s-]+)",
        r"perto\s+de\s+([a-z0-9\s-]+)",
        r"pr[oó]ximo\s+ao\s+([a-z0-9\s-]+)",
        r"proximo\s+ao\s+([a-z0-9\s-]+)",
        r"na\s+regiao\s+de\s+([a-z0-9\s-]+)",
        r"na\s+região\s+de\s+([a-z0-9\s-]+)",
        r"em\s+([a-z0-9\s-]+)$",
    ]
    for pattern in patterns:
        match = re.search(pattern, q)
        if match:
            location = match.group(1).strip(" .,:;!?")
            break

    return {"seniority": seniority, "location": location}


def _resolve_target_profile(query: str) -> ProfessionalProfile | None:
    role_id = resolve_role_id(query)
    profiles = load_profiles()

    if role_id:
        for profile in profiles:
            if profile.role_id == role_id:
                return profile

    normalized_query = normalize_role_query(query)
    query_tokens = normalized_query.split()

    best_profile = None
    best_score = -1.0

    for profile in profiles:
        text = " | ".join([profile.title, *profile.aliases, *profile.keywords, profile.summary])
        score, _ = _token_score(text, query_tokens)
        if score > best_score:
            best_score = score
            best_profile = profile

    return best_profile


def _role_score(profile: ProfessionalProfile, candidate_text: str) -> tuple[int, list[str]]:
    role_targets = [profile.title, profile.role_id, *profile.aliases, *profile.keywords]
    hits = _matched_terms(candidate_text, role_targets)

    if len(hits) >= 3:
        return 20, hits[:6]
    if len(hits) == 2:
        return 16, hits[:6]
    if len(hits) == 1:
        return 10, hits[:6]

    return 0, []


def _required_skills_score(profile: ProfessionalProfile, candidate_text: str) -> tuple[int, list[str], list[str]]:
    required = list(profile.required_skills or [])
    if not required:
        return 0, [], []

    matched = _matched_terms(candidate_text, required)
    missing = [item for item in required if item not in matched]

    ratio = len(matched) / max(len(required), 1)
    score = round(min(20, ratio * 20))

    return score, matched[:10], missing[:10]


def _priority_specs_score(profile: ProfessionalProfile, candidate_text: str) -> tuple[int, list[str], list[str]]:
    rules = load_matching_rules()
    global_terms = rules.get("global_terms", {}) or {}
    role_rules = (rules.get("role_rules", {}) or {}).get(profile.role_id, [])

    total = 0
    matched = []
    gaps = []

    for rule in role_rules:
        label = _safe_str(rule.get("label")) or "Critério"
        points = int(rule.get("points", 0) or 0)

        terms = rule.get("terms")
        terms_key = rule.get("terms_key")

        if terms_key:
            terms = global_terms.get(terms_key, [])

        terms = terms or []

        if _contains_any(candidate_text, terms):
            total += points
            matched.append(f"{label} (+{points})")
        else:
            gaps.append(label)

    return min(total, 40), matched[:10], gaps[:10]


def _seniority_score(profile: ProfessionalProfile, candidate_text: str, query: str = "") -> tuple[int, list[str], list[str]]:
    text = singularize_pt(candidate_text + " " + query)
    expected = singularize_pt(profile.seniority)

    if "senior" in expected and "senior" in text:
        return 10, ["Senioridade sênior compatível"], []
    if "pleno" in expected and ("pleno" in text or "senior" in text):
        return 8, ["Senioridade compatível"], []
    if "junior" in expected and ("junior" in text or "júnior" in text):
        return 8, ["Senioridade júnior compatível"], []

    constraints = _parse_query_constraints(query)
    if constraints["seniority"] and constraints["seniority"] in text:
        return 10, [f"Senioridade pedida compatível: {constraints['seniority']}"], []

    if any(x in text for x in ["senior", "pleno", "junior"]):
        return 3, ["Senioridade detectada, mas pode não bater perfeitamente"], ["Senioridade pode não bater totalmente"]

    return 0, [], ["Senioridade não identificada"]


def _location_score(candidate_text: str, query: str = "") -> tuple[int, list[str], list[str]]:
    text = singularize_pt(candidate_text)
    constraints = _parse_query_constraints(query)
    requested_location = singularize_pt(constraints.get("location", ""))

    if requested_location and requested_location in text:
        return 10, [f"Localização compatível com {constraints['location']}"], []

    sao_paulo_terms = [
        "sao paulo", "sp", "zona leste", "zl", "aricanduva",
        "tatuape", "itaquera", "mooca", "analia franco", "penha",
        "vila formosa", "carrão", "carrao"
    ]

    if _contains_any(text, sao_paulo_terms):
        return 7, ["Localização indica São Paulo/Zona Leste ou proximidade"], []

    if requested_location:
        return 0, [], [f"Não encontrei localização compatível com {constraints['location']}"]

    return 0, [], ["Localização/proximidade não identificada"]


def score_candidate_against_profile(
    row: dict[str, Any],
    profile: ProfessionalProfile,
    query: str = "",
    curriculum_text: str = "",
) -> dict[str, Any]:
    candidate_text = candidate_resume_text(row, curriculum_text=curriculum_text)
    contacts = candidate_contact_text(row)

    role_points, matched_role = _role_score(profile, candidate_text)
    required_points, matched_required, missing_required = _required_skills_score(profile, candidate_text)
    priority_points, matched_priority, missing_priority = _priority_specs_score(profile, candidate_text)
    seniority_points, matched_seniority, missing_seniority = _seniority_score(profile, candidate_text, query=query)
    location_points, matched_location, missing_location = _location_score(candidate_text, query=query)

    total = role_points + required_points + priority_points + seniority_points + location_points
    total = max(0, min(100, int(round(total))))

    reasons = []
    if matched_role:
        reasons.append("Cargo/função: " + ", ".join(matched_role[:4]))
    if matched_required:
        reasons.append("Skills: " + ", ".join(matched_required[:6]))
    if matched_priority:
        reasons.extend(matched_priority[:6])
    if matched_seniority:
        reasons.extend(matched_seniority)
    if matched_location:
        reasons.extend(matched_location)

    gaps = []
    gaps.extend(missing_priority)
    gaps.extend(missing_required[:5])
    gaps.extend(missing_seniority)
    gaps.extend(missing_location)

    return {
        "candidate_name": candidate_name(row),
        "candidate_role": candidate_role_text(row),
        "candidate_seniority": candidate_seniority_text(row),
        "candidate_location": candidate_location_text(row),
        "candidate_email": contacts["email"],
        "candidate_phone": contacts["telefone"],
        "candidate_resume_path": contacts["curriculo"],
        "profile_role_id": profile.role_id,
        "profile_title": profile.title,
        "role_id": profile.role_id,
        "title": profile.title,
        "score": round(total / 100, 4),
        "score_pct": round(float(total), 1),
        "nota": total,
        "breakdown": {
            "role_title": role_points,
            "profile_required_skills": required_points,
            "priority_specs": priority_points,
            "seniority": seniority_points,
            "location": location_points,
        },
        "matched_role_terms": matched_role,
        "matched_required_skills": matched_required,
        "matched_priority_specs": matched_priority,
        "missing_required_skills": missing_required,
        "missing_priority_specs": missing_priority,
        "reasons": reasons[:12],
        "gaps": gaps[:12],
        "query_constraints": _parse_query_constraints(query),
        "raw_row": row,
    }


def score_candidate_against_profiles(
    candidate: dict[str, Any],
    curriculum_text: str = "",
    requested_role: str | None = None,
    requested_location: str = "",
    extra_query: str = "",
    top_k: int = 3,
) -> dict[str, Any]:
    profiles = [p for p in load_profiles() if p.status == "active"]

    query = " ".join(
        [
            requested_role or "",
            requested_location or "",
            extra_query or "",
            candidate_role_text(candidate),
        ]
    ).strip()

    if requested_role:
        role_id = resolve_role_id(requested_role)
        if role_id:
            profiles = [p for p in profiles if p.role_id == role_id]

    scored = [
        score_candidate_against_profile(
            row=candidate,
            profile=profile,
            query=query,
            curriculum_text=curriculum_text,
        )
        for profile in profiles
    ]

    scored.sort(key=lambda item: (-item["nota"], item["profile_title"]))

    best = scored[0] if scored else {
        "role_id": "",
        "title": "",
        "profile_role_id": "",
        "profile_title": "",
        "nota": 0,
        "score_pct": 0.0,
        "breakdown": {},
        "reasons": [],
        "gaps": [],
    }

    return {
        "best": best,
        "top_matches": scored[:top_k],
        "nota": best.get("nota", 0),
        "role_id": best.get("role_id", ""),
        "title": best.get("title", ""),
        "summary": build_match_summary(best),
    }


def search_candidates_by_profile_query(
    query: str,
    limit: int = 10,
    min_score: float = 0.15,
) -> dict[str, Any]:
    profile = _resolve_target_profile(query)

    if not profile:
        return {
            "query": query,
            "profile": None,
            "matches": [],
            "total_candidates": 0,
            "message": "Não consegui resolver um perfil profissional para essa busca.",
        }

    rows = load_banco_talentos_rows()

    scored = [
        score_candidate_against_profile(row, profile=profile, query=query)
        for row in rows
    ]

    scored = [item for item in scored if item["score"] >= min_score]
    scored.sort(key=lambda item: (-item["nota"], item["candidate_name"].lower()))

    return {
        "query": query,
        "profile": {
            "role_id": profile.role_id,
            "title": profile.title,
            "family": profile.family,
            "hub": profile.hub,
            "seniority": profile.seniority,
            "location": profile.location,
        },
        "matches": scored[:limit],
        "total_candidates": len(rows),
        "message": f"Encontrei {len(scored[:limit])} candidato(s) aderentes para '{profile.title}'.",
    }


def build_match_summary(match: dict[str, Any]) -> str:
    score = match.get("nota", match.get("score_pct", 0))
    title = match.get("title") or match.get("profile_title") or ""
    reasons = match.get("reasons") or []
    gaps = match.get("gaps") or []

    parts = [f"Nota {score}/100 para {title}".strip()]

    if reasons:
        parts.append("Pontos fortes: " + "; ".join(reasons[:4]))

    if gaps:
        parts.append("Gaps: " + "; ".join(gaps[:4]))

    return " | ".join(parts)


def format_match_summary(result: dict[str, Any]) -> str:
    profile = result.get("profile")
    matches = result.get("matches") or []

    if not profile:
        return result.get("message") or "Não consegui resolver um perfil para essa busca."

    if not matches:
        return f"Não encontrei candidatos aderentes para '{profile['title']}' no banco de talentos."

    lines = [f"Perfil alvo: {profile['title']}", "", "Melhores aderências:"]

    for idx, item in enumerate(matches, start=1):
        line = f"{idx}. {item['candidate_name']} | Nota: {item['nota']}/100"

        if item.get("candidate_role"):
            line += f" | Cargo: {item['candidate_role']}"
        if item.get("candidate_seniority"):
            line += f" | Nível: {item['candidate_seniority']}"
        if item.get("candidate_location"):
            line += f" | Local: {item['candidate_location']}"

        lines.append(line)

        details = []
        if item.get("candidate_resume_path"):
            details.append(f"Currículo: {item['candidate_resume_path']}")
        if item.get("candidate_phone"):
            details.append(f"Telefone: {item['candidate_phone']}")
        if item.get("candidate_email"):
            details.append(f"Email: {item['candidate_email']}")

        if details:
            lines.append("   " + " | ".join(details))

        evidence_parts = []
        if item.get("reasons"):
            evidence_parts.append("pontos fortes: " + "; ".join(item["reasons"][:3]))
        if item.get("gaps"):
            evidence_parts.append("gaps: " + "; ".join(item["gaps"][:3]))

        if evidence_parts:
            lines.append("   " + " | ".join(evidence_parts))

    return "\n".join(lines)