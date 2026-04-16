from __future__ import annotations

import json
import os
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from datasets.professional_profiles.schema import (
    PersonaModel,
    ProfessionalProfile,
    ProfessionalProfilesCatalog,
)


BASE_DIR = Path(__file__).resolve().parent
LOCAL_CATALOG_JSON = BASE_DIR / "profiles_catalog.json"

DEFAULT_DRIVE_ROOT = os.getenv(
    "STARIA_DRIVE_ROOT",
    r"G:\Drives compartilhados\STARMKT\StarIA",
)
PROFILES_DIR = Path(
    os.getenv(
        "STARIA_PROFILES_DIR",
        str(Path(DEFAULT_DRIVE_ROOT) / "banco_talentos" / "perfis"),
    )
)
SHARED_CATALOG_XLSX = Path(
    os.getenv(
        "STARIA_PROFILES_XLSX",
        str(PROFILES_DIR / "profiles_catalog.xlsx"),
    )
)


XLSX_HEADERS = [
    "role_id",
    "title",
    "family",
    "hub",
    "seniority",
    "work_model",
    "location",
    "summary",
    "responsibilities",
    "required_skills",
    "preferred_skills",
    "tools",
    "experience_requirements",
    "education_requirements",
    "keywords",
    "aliases",
    "persona_archetype",
    "persona_summary",
    "persona_work_style",
    "persona_behavioral_traits",
    "persona_decision_style",
    "persona_ideal_signals",
    "persona_risk_signals",
    "source_docs",
    "version",
    "status",
    "updated_at",
]


def _safe_list(value: Any) -> list[str]:
    if value is None:
        return []

    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]

    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return []

        try:
            parsed = json.loads(raw)
            if isinstance(parsed, list):
                return [str(item).strip() for item in parsed if str(item).strip()]
        except Exception:
            pass

        return [item.strip() for item in raw.split("|") if item.strip()]

    return [str(value).strip()] if str(value).strip() else []


def _list_to_cell(value: list[str]) -> str:
    return " | ".join(_safe_list(value))


def _read_json_file(path: Path) -> dict:
    if not path.exists():
        return {"profiles": []}

    raw = path.read_text(encoding="utf-8").strip()
    if not raw:
        return {"profiles": []}

    data = json.loads(raw)

    if isinstance(data, list):
        return {"profiles": data}

    if isinstance(data, dict) and isinstance(data.get("profiles"), list):
        return data

    return {"profiles": []}


def ensure_profiles_dir(path: Path = PROFILES_DIR) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def load_catalog_from_json(path: Path = LOCAL_CATALOG_JSON) -> ProfessionalProfilesCatalog:
    data = _read_json_file(path)
    return ProfessionalProfilesCatalog(**data)


def save_catalog_to_json(
    catalog: ProfessionalProfilesCatalog,
    path: Path = LOCAL_CATALOG_JSON,
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = catalog.model_dump(mode="json")
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return path


def profile_to_row(profile: ProfessionalProfile) -> list[Any]:
    persona = profile.persona or PersonaModel()

    return [
        profile.role_id,
        profile.title,
        profile.family,
        profile.hub,
        profile.seniority,
        profile.work_model,
        profile.location,
        profile.summary,
        _list_to_cell(profile.responsibilities),
        _list_to_cell(profile.required_skills),
        _list_to_cell(profile.preferred_skills),
        _list_to_cell(profile.tools),
        _list_to_cell(profile.experience_requirements),
        _list_to_cell(profile.education_requirements),
        _list_to_cell(profile.keywords),
        _list_to_cell(profile.aliases),
        persona.archetype,
        persona.summary,
        _list_to_cell(persona.work_style),
        _list_to_cell(persona.behavioral_traits),
        _list_to_cell(persona.decision_style),
        _list_to_cell(persona.ideal_signals),
        _list_to_cell(persona.risk_signals),
        _list_to_cell(profile.source_docs),
        profile.version,
        profile.status,
        profile.updated_at,
    ]


def export_json_to_xlsx(
    json_path: Path = LOCAL_CATALOG_JSON,
    xlsx_path: Path = SHARED_CATALOG_XLSX,
) -> Path:
    ensure_profiles_dir(xlsx_path.parent)
    catalog = load_catalog_from_json(json_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "profiles"

    ws.append(XLSX_HEADERS)

    for profile in catalog.profiles:
        ws.append(profile_to_row(profile))

    wb.save(xlsx_path)
    return xlsx_path


def _row_to_profile_dict(headers: list[str], row: list[Any]) -> dict[str, Any]:
    values = dict(zip(headers, row))

    persona = {
        "archetype": (values.get("persona_archetype") or "").strip(),
        "summary": (values.get("persona_summary") or "").strip(),
        "work_style": _safe_list(values.get("persona_work_style")),
        "behavioral_traits": _safe_list(values.get("persona_behavioral_traits")),
        "decision_style": _safe_list(values.get("persona_decision_style")),
        "ideal_signals": _safe_list(values.get("persona_ideal_signals")),
        "risk_signals": _safe_list(values.get("persona_risk_signals")),
    }

    return {
        "role_id": (values.get("role_id") or "").strip(),
        "title": (values.get("title") or "").strip(),
        "family": (values.get("family") or "").strip(),
        "hub": (values.get("hub") or "").strip(),
        "seniority": (values.get("seniority") or "").strip(),
        "work_model": (values.get("work_model") or "").strip(),
        "location": (values.get("location") or "").strip(),
        "summary": (values.get("summary") or "").strip(),
        "responsibilities": _safe_list(values.get("responsibilities")),
        "required_skills": _safe_list(values.get("required_skills")),
        "preferred_skills": _safe_list(values.get("preferred_skills")),
        "tools": _safe_list(values.get("tools")),
        "experience_requirements": _safe_list(values.get("experience_requirements")),
        "education_requirements": _safe_list(values.get("education_requirements")),
        "keywords": _safe_list(values.get("keywords")),
        "aliases": _safe_list(values.get("aliases")),
        "persona": persona,
        "source_docs": _safe_list(values.get("source_docs")),
        "version": int(values.get("version") or 1),
        "status": (values.get("status") or "active").strip(),
        "updated_at": (values.get("updated_at") or "").strip(),
    }


def load_catalog_from_xlsx(path: Path = SHARED_CATALOG_XLSX) -> ProfessionalProfilesCatalog:
    if not path.exists():
        return ProfessionalProfilesCatalog(profiles=[])

    wb = load_workbook(path)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ProfessionalProfilesCatalog(profiles=[])

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    profiles: list[ProfessionalProfile] = []

    for row in rows[1:]:
        if not row or not any(cell not in (None, "") for cell in row):
            continue

        data = _row_to_profile_dict(headers, list(row))
        if not data["role_id"] or not data["title"]:
            continue

        profiles.append(ProfessionalProfile(**data))

    return ProfessionalProfilesCatalog(profiles=profiles)


def export_xlsx_to_json(
    xlsx_path: Path = SHARED_CATALOG_XLSX,
    json_path: Path = LOCAL_CATALOG_JSON,
) -> Path:
    catalog = load_catalog_from_xlsx(xlsx_path)
    return save_catalog_to_json(catalog, json_path)


def sync_profiles(
    source: str = "xlsx",
    json_path: Path = LOCAL_CATALOG_JSON,
    xlsx_path: Path = SHARED_CATALOG_XLSX,
) -> Path:
    source_norm = (source or "xlsx").strip().lower()

    if source_norm == "json":
        return export_json_to_xlsx(json_path=json_path, xlsx_path=xlsx_path)

    if source_norm == "xlsx":
        return export_xlsx_to_json(xlsx_path=xlsx_path, json_path=json_path)

    raise ValueError("source must be 'json' or 'xlsx'")


if __name__ == "__main__":
    ensure_profiles_dir()
    output = sync_profiles("xlsx")
    print(f"Sync concluído: {output}")