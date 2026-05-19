from __future__ import annotations

import datetime
import email
import hashlib
import imaplib
import json
import os
import re
import shutil
import subprocess
import sys
import unicodedata
from email.header import decode_header
from email.utils import parseaddr
from pathlib import Path
from typing import Any

import requests
from docx import Document
from dotenv import load_dotenv
from openpyxl import load_workbook
from pypdf import PdfReader

BASE_DIR = Path(__file__).resolve().parents[1]
sys.path.append(str(BASE_DIR))

from datasets.professional_profiles.matching_engine import (
    score_candidate_against_profiles,
)
from rh.talent_bank_workbook import (
    CANONICAL_SHEETS,
    append_candidate_record,
    backup_workbook,
    build_bank_headers,
    ensure_bank_workbook_structure,
    normalize_role_to_sheet_name,
    sheet_display_title,
)

# =========================
# LOAD .ENV DO BACKEND
# =========================

load_dotenv(BASE_DIR / ".env", override=True)


# =========================
# CONFIG
# =========================

IMAP_SERVER = "imap.gmail.com"

EMAIL_ACCOUNT = os.getenv("STARIA_EMAIL_ACCOUNT", "staria@starmkt.com.br").strip()
EMAIL_PASSWORD = (
    os.getenv("STARIA_EMAIL_PASSWORD", "")
    .strip()
    .replace(" ", "")
    .replace('"', "")
    .replace("'", "")
)

STARIA_ROOT = Path(
    os.getenv("STARIA_DRIVE_ROOT", r"G:\Drives compartilhados\STARMKT\StarIA")
)

CURRICULOS_DIR = Path(
    os.getenv(
        "STARIA_CURRICULOS_DIR",
        str(STARIA_ROOT / "banco_talentos" / "curriculos"),
    )
)

BANCO_TALENTOS_XLSX = Path(
    os.getenv(
        "STARIA_TALENTS_XLSX",
        str(STARIA_ROOT / "banco_talentos" / "banco_talentos.xlsx"),
    )
)

CANDIDATOS_REFINADOS_XLSX = Path(
    os.getenv(
        "CANDIDATOS_REFINADOS_XLSX",
        str(STARIA_ROOT / "banco_talentos" / "candidatos_refinados.xlsx"),
    )
)

REJECTED_DIR = CURRICULOS_DIR / "_rejeitados_nota_menor_40"
INDEX_PATH = CURRICULOS_DIR / "_curriculos_index.json"

ALLOWED_EXTS = {".pdf", ".doc", ".docx", ".txt", ".rtf"}

STARIA_API_BASE = os.getenv("STARIA_API_BASE", "http://127.0.0.1:8088").strip()
STARIA_OLLAMA_MODEL = os.getenv("STAR_OLLAMA_MODEL", "star-llama").strip()

DEFAULT_STATUS = "Banco de talentos"
DEFAULT_ORIGEM = "email"

MIN_SCORE = int(os.getenv("STARIA_MIN_SCORE", "40"))

REPROCESS_ALL_EMAILS = os.getenv("STARIA_REPROCESS_ALL_EMAILS", "false").lower() in {
    "1",
    "true",
    "yes",
    "sim",
}

RESET_TALENT_BANK_BEFORE_REPROCESS = os.getenv(
    "STARIA_RESET_TALENT_BANK_BEFORE_REPROCESS", "false"
).lower() in {"1", "true", "yes", "sim"}

EMAIL_LIMIT = int(os.getenv("STARIA_EMAIL_LIMIT", "0") or 0)

MOVE_REJECTED_UNDER_MIN_SCORE = os.getenv(
    "STARIA_MOVE_REJECTED_UNDER_MIN_SCORE", "true"
).lower() in {"1", "true", "yes", "sim"}


# =========================
# UTILS
# =========================


def decode_mime_words(s: str) -> str:
    if not s:
        return ""

    decoded = decode_header(s)
    result = ""

    for part, encoding in decoded:
        if isinstance(part, bytes):
            result += part.decode(encoding or "utf-8", errors="ignore")
        else:
            result += part

    return result


def normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "")).strip()


def safe_str(value: Any) -> str:
    if value is None:
        return ""
    return normalize_spaces(str(value))


def normalize_text(text: str) -> str:
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = text.lower().strip()
    text = re.sub(r"\s+", " ", text)
    return text


def sanitize_filename(value: str, max_len: int = 140) -> str:
    value = decode_mime_words(value or "")
    value = re.sub(r'[<>:"/\\|?*\n\r\t]+', " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value[:max_len].strip() or "curriculo"


def decode_sender(sender_raw: str) -> str:
    name, addr = parseaddr(sender_raw or "")
    name = decode_mime_words(name)
    if name and addr:
        return f"{name} <{addr}>"
    return addr or name or ""


def ensure_paths():
    CURRICULOS_DIR.mkdir(parents=True, exist_ok=True)
    REJECTED_DIR.mkdir(parents=True, exist_ok=True)

    if not BANCO_TALENTOS_XLSX.exists():
        raise RuntimeError(f"Planilha não encontrada: {BANCO_TALENTOS_XLSX}")


def validate_config():
    if not EMAIL_ACCOUNT:
        raise RuntimeError("STARIA_EMAIL_ACCOUNT não definido no .env")

    if not EMAIL_PASSWORD:
        raise RuntimeError(
            "STARIA_EMAIL_PASSWORD não definido no .env " f"({BASE_DIR / '.env'})"
        )


# =========================
# REGEX EXTRACTION
# =========================

EMAIL_REGEX = r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+"
PHONE_REGEX = r"(\+55\s?)?(\(?\d{2}\)?\s?)?\d{4,5}[-.\s]?\d{4}"

PORTFOLIO_PATTERNS = [
    r"https?://(?:www\.)?behance\.net/\S+",
    r"https?://(?:www\.)?dribbble\.com/\S+",
    r"https?://(?:www\.)?github\.com/\S+",
    r"https?://(?:www\.)?linkedin\.com/\S+",
    r"https?://[^\s)>\]]+",
]


def extract_email_regex(text: str) -> str:
    match = re.search(EMAIL_REGEX, text or "")
    return match.group(0).strip() if match else ""


def extract_phone_regex(text: str) -> str:
    match = re.search(PHONE_REGEX, text or "")
    return match.group(0).strip() if match else ""


def extract_portfolio_regex(text: str) -> str:
    for pattern in PORTFOLIO_PATTERNS:
        match = re.search(pattern, text or "", flags=re.IGNORECASE)
        if match:
            return match.group(0).strip().rstrip(".,;")
    return ""


# =========================
# TALENT BANK RESET
# =========================


def clear_talent_bank_data_rows():
    print("[EMAIL] Limpando dados da planilha mantendo abas, headers e estrutura...")

    backup = backup_workbook(BANCO_TALENTOS_XLSX)

    ensure_bank_workbook_structure(
        banco_path=BANCO_TALENTOS_XLSX,
        refined_path=CANDIDATOS_REFINADOS_XLSX,
        create_backup=False,
        redistribute_existing=False,
    )

    wb = load_workbook(BANCO_TALENTOS_XLSX)
    headers = build_bank_headers(None)

    for sheet_name in CANONICAL_SHEETS:
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)

        ws = wb[sheet_name]

        if ws.max_row:
            ws.delete_rows(1, ws.max_row)

        ws.append(headers)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=idx, value=header)
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = 22

    for sheet_name in list(wb.sheetnames):
        if sheet_name not in CANONICAL_SHEETS:
            del wb[sheet_name]

    wb.save(BANCO_TALENTOS_XLSX)

    print("[EMAIL] Planilha zerada com sucesso.")
    print("[EMAIL] Backup criado:", backup if backup else "(sem backup)")


# =========================
# EMAIL BODY
# =========================


def extract_email_body(msg) -> str:
    body = ""

    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            disposition = str(part.get("Content-Disposition") or "")

            if "attachment" in disposition.lower():
                continue

            payload = part.get_payload(decode=True)
            charset = part.get_content_charset() or "utf-8"

            if not payload:
                continue

            if content_type == "text/plain":
                body += payload.decode(charset, errors="ignore")

            elif content_type == "text/html" and not body:
                html = payload.decode(charset, errors="ignore")
                html = re.sub(r"<br\s*/?>", "\n", html, flags=re.I)
                html = re.sub(r"<[^>]+>", " ", html)
                body += normalize_spaces(html)
    else:
        payload = msg.get_payload(decode=True)
        charset = msg.get_content_charset() or "utf-8"

        if payload:
            body += payload.decode(charset, errors="ignore")

    return body


# =========================
# EXTRACT NÍVEL / PORTFOLIO
# =========================


def extract_level(email_body: str) -> str:
    if not email_body:
        return ""

    text = email_body.lower()

    patterns = [
        r"n[íi]vel\s*:\s*(j[uú]nior|pleno|s[êe]nior)",
        r"senioridade\s*:\s*(j[uú]nior|pleno|s[êe]nior)",
    ]

    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            value = match.group(1)

            if value in ["junior", "júnior"]:
                return "Júnior"
            if value == "pleno":
                return "Pleno"
            if value in ["senior", "sênior"]:
                return "Sênior"

    return ""


def infer_candidate_level(
    subject: str = "",
    body: str = "",
    curriculum_text: str = "",
    explicit_sheet: str = "",
) -> str:
    text = normalize_text(" ".join([subject or "", body or "", curriculum_text or ""]))

    # 1) Indício explícito no assunto/corpo/currículo
    senior_patterns = [
        r"\bsenior\b",
        r"\bs[eê]nior\b",
        r"\bsr\b",
        r"\bsr\.",
        r"\bespecialista\b",
        r"\bcoordena[cç][aã]o\b",
        r"\bcoordenador\b",
        r"\bcoordenadora\b",
        r"\bdiretor\b",
        r"\bdiretora\b",
        r"\bgerente\b",
    ]

    pleno_patterns = [
        r"\bpleno\b",
        r"\bpl\b",
        r"\bpl\.",
    ]

    junior_patterns = [
        r"\bjunior\b",
        r"\bj[uú]nior\b",
        r"\bjr\b",
        r"\bjr\.",
        r"\bestagio\b",
        r"\best[aá]gio\b",
        r"\bestagiario\b",
        r"\bestagi[aá]rio\b",
        r"\bassistente\b",
        r"\bauxiliar\b",
    ]

    for pattern in senior_patterns:
        if re.search(pattern, text):
            return "Sênior"

    for pattern in pleno_patterns:
        if re.search(pattern, text):
            return "Pleno"

    for pattern in junior_patterns:
        if re.search(pattern, text):
            return "Júnior"

    # 2) Experiência por anos explícitos
    years = []

    for m in re.finditer(r"(\d{1,2})\s*(?:anos|ano)\s+de\s+experi[eê]ncia", text):
        try:
            years.append(int(m.group(1)))
        except Exception:
            pass

    for m in re.finditer(r"experi[eê]ncia\s+de\s+(\d{1,2})\s*(?:anos|ano)", text):
        try:
            years.append(int(m.group(1)))
        except Exception:
            pass

    if years:
        max_years = max(years)

        if max_years >= 5:
            return "Sênior"

        if max_years >= 2:
            return "Pleno"

        return "Júnior"

    # 3) Alguns cargos do dataset já são naturalmente sênior
    if explicit_sheet in {
        "DIRETOR DE ARTE BRANDING",
        "DIRETOR DE ARTE DIGITAL",
        "DIRETOR DE ARTE INSTITUCIONAL",
        "ATENDIMENTO",
    }:
        if "senior" in text or "sênior" in text or "sr" in text:
            return "Sênior"

    return ""


def extract_portfolio(email_body: str) -> str:
    if not email_body:
        return ""

    found = extract_portfolio_regex(email_body)
    if found:
        return found

    pattern = r"(portf[oó]lio|portfolio)\s*:\s*(https?://\S+|\S+\.\S+)"
    match = re.search(pattern, email_body, flags=re.IGNORECASE)

    if match:
        return match.group(2).strip().rstrip(".,;")

    return ""


# =========================
# CONNECT MAILBOX
# =========================


def connect_mailbox():
    validate_config()
    ensure_paths()

    print("[EMAIL] BASE_DIR =", BASE_DIR)
    print("[EMAIL] EMAIL_ACCOUNT =", EMAIL_ACCOUNT)
    print("[EMAIL] PASSWORD_LOADED =", bool(EMAIL_PASSWORD))
    print("[EMAIL] CURRICULOS_DIR =", CURRICULOS_DIR)
    print("[EMAIL] BANCO_TALENTOS_XLSX =", BANCO_TALENTOS_XLSX)
    print("[EMAIL] STARIA_API_BASE =", STARIA_API_BASE)
    print("[EMAIL] REPROCESS_ALL_EMAILS =", REPROCESS_ALL_EMAILS)
    print(
        "[EMAIL] RESET_TALENT_BANK_BEFORE_REPROCESS =",
        RESET_TALENT_BANK_BEFORE_REPROCESS,
    )
    print("[EMAIL] EMAIL_LIMIT =", EMAIL_LIMIT if EMAIL_LIMIT else "sem limite")
    print("[EMAIL] MIN_SCORE =", MIN_SCORE)

    mail = imaplib.IMAP4_SSL(IMAP_SERVER)
    mail.login(EMAIL_ACCOUNT, EMAIL_PASSWORD)

    print("[EMAIL] Login IMAP realizado com sucesso.")
    return mail


# =========================
# FILE HASH / INDEX
# =========================


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def load_curriculos_index() -> dict[str, str]:
    if INDEX_PATH.exists():
        try:
            data = json.loads(INDEX_PATH.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                return {str(k): str(v) for k, v in data.items()}
        except Exception:
            pass

    index = {}
    for p in CURRICULOS_DIR.glob("*"):
        if p.is_file() and p.suffix.lower() in ALLOWED_EXTS.union({".pdf"}):
            try:
                index[sha256_file(p)] = str(p)
            except Exception:
                continue

    save_curriculos_index(index)
    return index


def save_curriculos_index(index: dict[str, str]):
    INDEX_PATH.write_text(
        json.dumps(index, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def existing_file_by_hash(file_hash: str, index: dict[str, str]) -> Path | None:
    existing = index.get(file_hash)
    if existing:
        p = Path(existing)
        if p.exists():
            return p
    return None


# =========================
# PDF CONVERSION / SAVE ATTACHMENTS
# =========================


def find_soffice() -> str:
    candidates = [
        "soffice",
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ]

    for c in candidates:
        try:
            result = subprocess.run(
                [c, "--version"],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                timeout=10,
            )
            if result.returncode == 0:
                return c
        except Exception:
            continue

    return "soffice"


def convert_to_pdf_if_possible(file_path: Path) -> Path:
    ext = file_path.suffix.lower()

    if ext == ".pdf":
        return file_path

    if ext not in {".doc", ".docx"}:
        return file_path

    try:
        output_dir = file_path.parent
        soffice = find_soffice()

        subprocess.run(
            [
                soffice,
                "--headless",
                "--convert-to",
                "pdf",
                str(file_path),
                "--outdir",
                str(output_dir),
            ],
            check=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            timeout=120,
        )

        pdf_path = file_path.with_suffix(".pdf")

        if pdf_path.exists():
            print(f"[EMAIL] Convertido para PDF: {pdf_path.name}")
            return pdf_path

    except Exception as e:
        print(f"[EMAIL] Falha ao converter PDF: {file_path.name} | {e}")

    return file_path


def build_attachment_safe_name(
    filename: str,
    subject: str = "",
    explicit_role: str = "",
) -> str:
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    original = sanitize_filename(filename)
    vacancy = sanitize_filename(
        explicit_role or extract_job_title_from_subject(subject)
    )

    if vacancy:
        return f"{timestamp}_{vacancy}_{original}"

    return f"{timestamp}_{original}"


def save_attachment(msg, subject: str = "", explicit_role: str = "") -> list[Path]:
    saved_files: list[Path] = []
    CURRICULOS_DIR.mkdir(parents=True, exist_ok=True)

    index = load_curriculos_index()
    index_changed = False

    for part in msg.walk():
        content_disposition = str(part.get("Content-Disposition") or "")

        if "attachment" not in content_disposition.lower():
            continue

        filename = part.get_filename()
        if not filename:
            continue

        filename = decode_mime_words(filename)
        ext = Path(filename).suffix.lower()

        if ext not in ALLOWED_EXTS:
            print(f"[EMAIL] Anexo ignorado por extensão: {filename}")
            continue

        payload = part.get_payload(decode=True)
        if not payload:
            print(f"[EMAIL] Anexo sem payload ignorado: {filename}")
            continue

        file_hash = sha256_bytes(payload)
        existing = existing_file_by_hash(file_hash, index)

        if existing:
            print(f"[EMAIL] Currículo já existe por hash. Reutilizando: {existing}")
            saved_files.append(existing)
            continue

        safe_name = build_attachment_safe_name(filename, subject, explicit_role)
        filepath = CURRICULOS_DIR / safe_name

        if filepath.exists():
            filepath = (
                CURRICULOS_DIR / f"{filepath.stem}_{file_hash[:10]}{filepath.suffix}"
            )

        with open(filepath, "wb") as f:
            f.write(payload)

        original_filepath = filepath
        filepath = convert_to_pdf_if_possible(filepath)

        try:
            final_hash = sha256_file(filepath)
        except Exception:
            final_hash = file_hash

        index[final_hash] = str(filepath)
        index[file_hash] = str(filepath)
        index_changed = True

        if filepath != original_filepath:
            print("[EMAIL] Currículo convertido para PDF:", filepath)
        else:
            print("[EMAIL] Currículo salvo:", filepath)

        saved_files.append(filepath)

    if index_changed:
        save_curriculos_index(index)

    return saved_files


def move_rejected_file(file_path: Path) -> Path:
    REJECTED_DIR.mkdir(parents=True, exist_ok=True)

    target = REJECTED_DIR / file_path.name

    if target.exists():
        target = (
            REJECTED_DIR
            / f"{file_path.stem}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}{file_path.suffix}"
        )

    try:
        shutil.move(str(file_path), str(target))
        return target
    except Exception:
        return file_path


# =========================
# TEXT EXTRACTION
# =========================


def extract_text_from_pdf(file_path: Path) -> str:
    try:
        reader = PdfReader(str(file_path))
        parts = []
        for page in reader.pages:
            txt = page.extract_text() or ""
            if txt.strip():
                parts.append(txt)
        return "\n".join(parts).strip()
    except Exception as e:
        print(f"[EMAIL] Falha ao ler PDF {file_path.name}: {e}")
        return ""


def extract_text_from_docx(file_path: Path) -> str:
    try:
        doc = Document(str(file_path))
        parts = []
        for p in doc.paragraphs:
            txt = p.text or ""
            if txt.strip():
                parts.append(txt)
        return "\n".join(parts).strip()
    except Exception as e:
        print(f"[EMAIL] Falha ao ler DOCX {file_path.name}: {e}")
        return ""


def extract_text_from_file(file_path: Path) -> str:
    ext = file_path.suffix.lower()

    if ext in {".txt", ".rtf"}:
        try:
            return file_path.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            try:
                return file_path.read_text(errors="ignore")
            except Exception as e:
                print(f"[EMAIL] Falha ao ler TXT/RTF {file_path.name}: {e}")
                return ""

    if ext == ".pdf":
        return extract_text_from_pdf(file_path)

    if ext == ".docx":
        return extract_text_from_docx(file_path)

    return ""


# =========================
# ROLE DETECTION - FONTE DA VERDADE
# =========================

ROLE_PATTERNS = [
    (
        "DIAGRAMADOR",
        [
            r"\bvaga\s+diagramador(?:a)?\b",
            r"\bdiagramador(?:a)?\b",
            r"\bdiagrama[cç][aã]o\b",
            r"\btabloide\b",
            r"\btabl[oó]ide\b",
            r"\bencarte\b",
            r"\bofertas\b",
        ],
    ),
    (
        "ATENDIMENTO",
        [
            r"\bvaga\s+atendimento\b",
            r"\batendimento\b",
            r"\batendimento\s+s[eê]nior\b",
            r"\baccount\s+manager\b",
            r"\brelacionamento\b",
        ],
    ),
    (
        "COORDENADOR DE CONTEÚDO",
        [
            r"\bvaga\s+designer\b",
            r"\bvaga\s+marketing\b",
            r"\bvaga\s+social\s+media\b",
            r"\bcoordenador(?:a)?\s+de\s+conte[uú]do\b",
            r"\bconte[uú]do\b",
            r"\bsocial\s+media\b",
            r"\bmarketing\b",
            r"\bdesigner\s+gr[aá]fico\b",
        ],
    ),
    (
        "DIRETOR DE ARTE BRANDING",
        [
            r"\bvaga\s+diretor(?:a)?\s+de\s+arte.*branding\b",
            r"\bdiretor(?:a)?\s+de\s+arte\b.*\bbranding\b",
            r"\bdiretor(?:a)?\s+de\s+arte\b.*\bidentidade\b",
            r"\bdiretor(?:a)?\s+de\s+arte\b.*\bproduto\b",
        ],
    ),
    (
        "DIRETOR DE ARTE DIGITAL",
        [
            r"\bvaga\s+diretor(?:a)?\s+de\s+arte.*digital\b",
            r"\bdiretor(?:a)?\s+de\s+arte\b.*\bdigital\b",
            r"\bdiretor(?:a)?\s+de\s+arte\b.*\bperformance\b",
            r"\bdiretor(?:a)?\s+de\s+arte\b.*\bmarketplace\b",
        ],
    ),
    (
        "DIRETOR DE ARTE INSTITUCIONAL",
        [
            r"\bvaga\s+diretor(?:a)?\s+de\s+arte\b",
            r"\bdiretor(?:a)?\s+de\s+arte\b.*\binstitucional\b",
            r"\bdiretor(?:a)?\s+de\s+arte\b.*\bcampanhas\b",
            r"\bdiretor(?:a)?\s+de\s+arte\b",
        ],
    ),
    (
        "PLAN. PERFORMANCE & GROWTH",
        [
            r"\bvaga\s+planejamento\s+performance\b",
            r"\bvaga\s+performance\b",
            r"\bperformance\b",
            r"\bgrowth\b",
            r"\bgoogle\s+ads\b",
            r"\bmeta\s+ads\b",
            r"\btiktok\s+ads\b",
            r"\banalytics\b",
        ],
    ),
    (
        "PLANEJAMENTO ESTRATÉGICO",
        [
            r"\bvaga\s+planejamento\b",
            r"\bvaga\s+planejamento\s+estrat[eé]gico\b",
            r"\bplanejamento\s+estrat[eé]gico\b",
            r"\bestrategista\b",
            r"\bplanejamento\b",
        ],
    ),
    (
        "MOTION DESIGNER",
        [
            r"\bvaga\s+designer\s+motion\b",
            r"\bvaga\s+motion\b",
            r"\bmotion\b",
            r"\bmotion\s+designer\b",
            r"\banima[cç][aã]o\b",
        ],
    ),
    (
        "REDATOR",
        [
            r"\bvaga\s+redator(?:a)?\b",
            r"\bredator(?:a)?\b",
            r"\bcopywriter\b",
            r"\breda[cç][aã]o\b",
        ],
    ),
    (
        "EXECUTIVO DE CONTAS",
        [
            r"\bvaga\s+executivo(?:a)?\s+de\s+contas\b",
            r"\bexecutivo(?:a)?\s+de\s+contas\b",
            r"\bexecutivo(?:a)?\s+comercial\b",
            r"\baccount\s+executive\b",
        ],
    ),
]


def detect_explicit_role_from_text(text: str) -> str:
    n = normalize_text(text)

    if not n:
        return ""

    for sheet_name, patterns in ROLE_PATTERNS:
        for pattern in patterns:
            if re.search(pattern, n, flags=re.IGNORECASE):
                return sheet_name

    return ""


def extract_level_from_text(text: str) -> str:
    n = normalize_text(text)

    if re.search(r"\b(senior|s[eê]nior|sr\.?|especialista)\b", n):
        return "Sênior"

    if re.search(r"\b(pleno|pl\.?)\b", n):
        return "Pleno"

    if re.search(
        r"\b(junior|j[uú]nior|jr\.?|est[aá]gio|estagi[aá]rio|assistente|auxiliar)\b", n
    ):
        return "Júnior"

    return ""


def infer_candidate_level(
    subject: str = "",
    body: str = "",
    curriculum_text: str = "",
    explicit_sheet: str = "",
) -> str:
    level = extract_level_from_text(subject)
    if level:
        return level

    level = extract_level_from_text(body)
    if level:
        return level

    level = extract_level_from_text(curriculum_text)
    if level:
        return level

    text = normalize_text(curriculum_text)
    years = []

    patterns = [
        r"(\d{1,2})\s*(?:anos|ano)\s+de\s+experi[eê]ncia",
        r"experi[eê]ncia\s+de\s+(\d{1,2})\s*(?:anos|ano)",
        r"mais\s+de\s+(\d{1,2})\s*(?:anos|ano)",
        r"(\d{1,2})\+\s*(?:anos|ano)",
    ]

    for pattern in patterns:
        for m in re.finditer(pattern, text):
            try:
                years.append(int(m.group(1)))
            except Exception:
                pass

    if years:
        max_years = max(years)
        if max_years < 2:
            return "Júnior"
        if max_years <= 5:
            return "Pleno"
        return "Sênior"

    return "Pleno"


def extract_job_title_from_subject(subject: str) -> str:
    if not subject:
        return ""

    text = decode_mime_words(subject)

    patterns = [
        r"new application:\s*(.+?)\s+from\s+.+$",
        r"application received:\s*(.+?)\s+from\s+.+$",
        r"candidate applied:\s*(.+?)\s+from\s+.+$",
        r"candidatura:\s*(.+?)\s+de\s+.+$",
        r"nova candidatura:\s*(.+?)\s+de\s+.+$",
    ]

    for pattern in patterns:
        m = re.search(pattern, text, flags=re.IGNORECASE)
        if m:
            return normalize_spaces(m.group(1))

    return ""


def extract_explicit_role_from_email(subject: str, body: str) -> str:
    linked_in_title = extract_job_title_from_subject(subject)
    if linked_in_title:
        role = detect_explicit_role_from_text(linked_in_title)
        if role:
            return role

    role = detect_explicit_role_from_text(subject)
    if role:
        return role

    # Só usa o corpo do email, não o currículo.
    # Procura em trechos objetivos de vaga/cargo.
    body_text = normalize_text(body or "")
    focused_patterns = [
        r"vaga\s*[:\-]\s*([^\n\r|;,.]{2,80})",
        r"cargo\s*[:\-]\s*([^\n\r|;,.]{2,80})",
        r"oportunidade\s*[:\-]\s*([^\n\r|;,.]{2,80})",
        r"candidatura\s*[:\-]\s*([^\n\r|;,.]{2,80})",
        r"aplicou\s+para\s+([^\n\r|;,.]{2,80})",
        r"aplicando\s+para\s+([^\n\r|;,.]{2,80})",
        r"interesse\s+na\s+vaga\s+de\s+([^\n\r|;,.]{2,80})",
    ]

    for pattern in focused_patterns:
        m = re.search(pattern, body_text, flags=re.IGNORECASE)
        if not m:
            continue

        role = detect_explicit_role_from_text(m.group(1))
        if role:
            return role

    return ""


# =========================
# AI EXTRACTION
# =========================


def extract_json_block(text: str) -> str:
    if not text:
        return ""

    text = text.strip().replace("```json", "").replace("```", "").strip()

    start = text.find("{")
    end = text.rfind("}")

    if start >= 0 and end > start:
        return text[start : end + 1].strip()

    return text


def clean_ni(value: Any) -> str:
    s = safe_str(value)
    if s.lower() in {
        "n/i",
        "ni",
        "não informado",
        "nao informado",
        "não encontrado",
        "nao encontrado",
        "null",
        "none",
    }:
        return ""
    return s


def extract_candidate_data_with_ai(file_path: Path) -> dict:
    text = extract_text_from_file(file_path)

    print(f"[EMAIL] Texto extraído de {file_path.name}: {len(text)} caracteres")

    empty = {
        "nome_completo": "",
        "idade": "",
        "localizacao": "",
        "cargo_pretendido": "",
        "habilidades": "",
        "formacoes": "",
        "email": "",
        "telefone": "",
        "portfolio": "",
    }

    if not text.strip():
        return empty

    email_regex = extract_email_regex(text)
    phone_regex = extract_phone_regex(text)
    portfolio_regex = extract_portfolio_regex(text)

    prompt = f"""
Extraia do currículo as informações abaixo e devolva APENAS JSON válido.

Campos:
{{
  "nome_completo": "",
  "idade": "",
  "localizacao": "",
  "cargo_pretendido": "",
  "habilidades": "",
  "formacoes": "",
  "email": "",
  "telefone": "",
  "portfolio": ""
}}

Regras:
- NÃO invente dados.
- Use apenas o currículo.
- Se não encontrar, deixe "".
- "nome_completo" deve ser nome de pessoa explicitamente presente.
- "idade" só se houver idade ou data de nascimento explícita.
- "localizacao" deve ser cidade/região/endereço explicitamente presente.
- "habilidades" deve ser string curta separada por "; ".
- "formacoes" deve ser string curta separada por "; ".
- Responda somente JSON puro.

Arquivo: {file_path.name}

Currículo:
{text[:15000]}
""".strip()

    try:
        resp = requests.post(
            f"{STARIA_API_BASE}/ask",
            json={
                "question": prompt,
                "use_rag": False,
                "model": STARIA_OLLAMA_MODEL,
            },
            timeout=180,
        )

        resp.raise_for_status()
        answer = resp.json().get("answer", "")
        parsed = json.loads(extract_json_block(answer))

    except Exception as e:
        print(f"[EMAIL] Falha IA: {e}")
        parsed = {}

    return {
        "nome_completo": clean_ni(parsed.get("nome_completo")),
        "idade": clean_ni(parsed.get("idade")),
        "localizacao": clean_ni(parsed.get("localizacao")),
        "cargo_pretendido": clean_ni(parsed.get("cargo_pretendido")),
        "habilidades": clean_ni(parsed.get("habilidades")),
        "formacoes": clean_ni(parsed.get("formacoes")),
        "email": clean_ni(parsed.get("email") or email_regex),
        "telefone": clean_ni(parsed.get("telefone") or phone_regex),
        "portfolio": clean_ni(parsed.get("portfolio") or portfolio_regex),
    }


def enrich_extracted_with_fallbacks(
    extracted: dict,
    file_path: Path,
    curriculum_text: str,
) -> dict:
    extracted = dict(extracted or {})

    if not extracted.get("email"):
        extracted["email"] = extract_email_regex(curriculum_text)

    if not extracted.get("telefone"):
        extracted["telefone"] = extract_phone_regex(curriculum_text)

    if not extracted.get("portfolio"):
        extracted["portfolio"] = extract_portfolio_regex(curriculum_text)

    if not extracted.get("nome_completo"):
        lines = curriculum_text.splitlines()
        for line in lines[:12]:
            line = normalize_spaces(line)
            if not line:
                continue
            normalized_line = normalize_text(line)
            if "curriculum" in normalized_line or "curriculo" in normalized_line:
                continue
            if len(line.split()) >= 2 and len(line) <= 80:
                extracted["nome_completo"] = line
                break

    return extracted


# =========================
# EXCEL / MATCHING
# =========================


def append_candidate_to_sheet(
    file_path: Path,
    sender: str,
    level: str,
    portfolio: str,
    extracted: dict,
    explicit_sheet: str,
    subject: str = "",
):
    if not explicit_sheet:
        print(
            f"[EMAIL] Sem vaga explícita confiável antes do currículo. "
            f"Linha NÃO será gravada: {file_path.name}"
        )
        return

    curriculum_text = extract_text_from_file(file_path)

    target_sheet = explicit_sheet
    cargo_canonico = sheet_display_title(target_sheet)

    extracted_for_match = dict(extracted)
    extracted_for_match["cargo_pretendido"] = cargo_canonico

    match_result = score_candidate_against_profiles(
        candidate=extracted_for_match,
        curriculum_text=curriculum_text,
        requested_role=cargo_canonico,
        extra_query=f"{subject} {cargo_canonico}".strip(),
    )

    best = match_result.get("best") or {}
    top_matches = match_result.get("top_matches") or []
    nota = int(match_result.get("nota", 0) or 0)

    if nota < MIN_SCORE:
        print(
            f"[EMAIL] Candidato descartado por nota abaixo de corte: "
            f"{extracted.get('nome_completo') or file_path.name} | "
            f"Vaga: {cargo_canonico} | Nota: {nota}"
        )

        if (
            REPROCESS_ALL_EMAILS
            and MOVE_REJECTED_UNDER_MIN_SCORE
            and file_path.exists()
        ):
            moved_to = move_rejected_file(file_path)
            print("[EMAIL] Currículo movido para rejeitados:", moved_to)

        return

    top_matches_text = "; ".join(
        [
            f"{m.get('title', '')} ({m.get('nota', m.get('score_pct', 0))})"
            for m in top_matches
            if m.get("title")
        ]
    )

    final_portfolio = extracted.get("portfolio") or portfolio or str(file_path)

    values = {
        "Nota": nota,
        "Nome completo": extracted.get("nome_completo", ""),
        "Localização": extracted.get("localizacao", ""),
        "Cargo pretendido": cargo_canonico,
        "Nível": level or "",
        "Portfólio": final_portfolio,
        "Habilidades": extracted.get("habilidades", ""),
        "Formação": extracted.get("formacoes", ""),
        "Email": extracted.get("email", ""),
        "Telefone": extracted.get("telefone", ""),
        "Caminho do currículo": str(file_path),
        "Nome do arquivo": file_path.name,
        "Data de entrada": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Origem": DEFAULT_ORIGEM,
        "Remetente do email": sender,
        "Status": DEFAULT_STATUS,
        "Observações": "",
        "Role ID sugerido": best.get("role_id", ""),
        "Título normalizado": best.get("title", ""),
        "Top 3 roles aderentes": top_matches_text,
        "Resumo de aderência": match_result.get("summary", ""),
        "Flags de risco": "; ".join(best.get("gaps", [])[:6]),
    }

    candidate_id = append_candidate_record(
        values=values,
        banco_path=BANCO_TALENTOS_XLSX,
        refined_path=CANDIDATOS_REFINADOS_XLSX,
        target_sheet=target_sheet,
    )

    if candidate_id == "DUPLICADO":
        print(
            f"[EMAIL] Candidato duplicado ignorado: "
            f"{values.get('Nome completo') or values.get('Email') or file_path.name} | "
            f"Aba: {target_sheet}"
        )
        return

    print(
        f"[EMAIL] Planilha atualizada com candidato {candidate_id}: "
        f"{file_path.name} | Aba: {target_sheet} | "
        f"Cargo: {cargo_canonico} | Nota: {nota} | "
        f"Perfil: {best.get('title', '')}"
    )


# =========================
# EMAIL FILTERS
# =========================

APPLICATION_KEYWORDS = [
    "new application",
    "application received",
    "candidate applied",
    "candidatura",
    "nova candidatura",
    "inscricao",
    "inscrição",
]

DIRECT_APPLICATION_HINTS = [
    "vaga",
    "curriculo",
    "currículo",
    "candidatura",
    "candidate",
    "application",
    "resume",
    "portfolio",
    "portfólio",
]


def is_linkedin_sender(sender: str) -> bool:
    return "jobs-listings@linkedin.com" in normalize_text(sender)


def has_application_signal(subject: str, body: str) -> bool:
    text = normalize_text(subject + " " + body)
    return any(k in text for k in APPLICATION_KEYWORDS)


def has_direct_application_signal(subject: str, body: str) -> bool:
    text = normalize_text(subject + " " + body)
    return any(k in text for k in DIRECT_APPLICATION_HINTS)


def email_has_cv_attachment(msg) -> bool:
    for part in msg.walk():
        content_disposition = str(part.get("Content-Disposition") or "")
        if "attachment" not in content_disposition.lower():
            continue

        filename = part.get_filename()
        if not filename:
            continue

        filename = decode_mime_words(filename)
        ext = Path(filename).suffix.lower()

        if ext in ALLOWED_EXTS:
            return True

    return False


def body_has_candidate_info(body: str) -> bool:
    text = normalize_text(body)

    candidate_signals = [
        "nome",
        "telefone",
        "email",
        "e-mail",
        "experiencia",
        "experiência",
        "formacao",
        "formação",
        "portfolio",
        "portfólio",
        "linkedin",
    ]

    return sum(1 for s in candidate_signals if s in text) >= 2


def should_process_email(msg, subject: str, sender: str, body: str) -> tuple[bool, str]:
    explicit_sheet = extract_explicit_role_from_email(subject, body)

    if not explicit_sheet:
        return False, ""

    has_cv = email_has_cv_attachment(msg)
    has_body_candidate = body_has_candidate_info(body)

    if has_cv or has_body_candidate:
        return True, explicit_sheet

    return False, explicit_sheet


# =========================
# PROCESS INBOX
# =========================


def process_inbox():
    mail = connect_mailbox()
    mail.select("inbox")

    if RESET_TALENT_BANK_BEFORE_REPROCESS:
        clear_talent_bank_data_rows()

    search_criteria = "ALL" if REPROCESS_ALL_EMAILS else "UNSEEN"
    status, messages = mail.search(None, search_criteria)

    if status != "OK":
        print("[EMAIL] Erro ao buscar emails.")
        return

    mail_ids = messages[0].split()
    print(f"[EMAIL] Critério IMAP: {search_criteria}")
    print("[EMAIL] Emails encontrados:", len(mail_ids))

    if EMAIL_LIMIT > 0:
        mail_ids = mail_ids[:EMAIL_LIMIT]
        print("[EMAIL] MODO TESTE/LIMITE. Processando apenas:", len(mail_ids))

    processed = 0
    ignored = 0
    with_resume = 0
    no_explicit_role = 0

    for idx, mail_id in enumerate(mail_ids, start=1):
        status, msg_data = mail.fetch(mail_id, "(RFC822)")

        if status != "OK":
            print("[EMAIL] Falha ao buscar email ID:", mail_id)
            continue

        raw_email = msg_data[0][1]
        msg = email.message_from_bytes(raw_email)

        subject = decode_mime_words(msg.get("Subject", ""))
        sender_raw = msg.get("From", "")
        sender = decode_sender(sender_raw)
        body = extract_email_body(msg)

        print(f"\n[EMAIL] {idx}/{len(mail_ids)} Processando email")
        print("[EMAIL] Assunto:", subject)
        print("[EMAIL] De:", sender)

        should_process, explicit_sheet = should_process_email(
            msg, subject, sender, body
        )

        if not should_process:
            if email_has_cv_attachment(msg) and not explicit_sheet:
                print(
                    "[EMAIL] Ignorado: tem currículo, mas não há vaga explícita confiável no assunto/corpo."
                )
                no_explicit_role += 1
            else:
                print("[EMAIL] Ignorado por não atender critérios de currículo")
            ignored += 1
            continue

        level = infer_candidate_level(subject=subject, body=body, curriculum_text="")
        portfolio = extract_portfolio(body)
        cargo_canonico = sheet_display_title(explicit_sheet)

        print("[EMAIL] Nível detectado:", level if level else "não informado")
        print(
            "[EMAIL] Portfólio detectado:", portfolio if portfolio else "não informado"
        )
        print("[EMAIL] Vaga explícita detectada:", cargo_canonico)
        print("[EMAIL] Aba destino:", explicit_sheet)

        saved = save_attachment(msg, subject=subject, explicit_role=cargo_canonico)

        if not saved:
            print("[EMAIL] Nenhum currículo válido encontrado")
            ignored += 1
            continue

        with_resume += 1
        print("[EMAIL] Total de currículos disponíveis:", len(saved))

        for file_path in saved:
            curriculum_text = extract_text_from_file(file_path)

            level_from_curriculum = infer_candidate_level(
                subject=subject,
                body=body,
                curriculum_text=curriculum_text,
            )

            if level_from_curriculum:
                level = level_from_curriculum

            extracted = extract_candidate_data_with_ai(file_path)
            extracted = enrich_extracted_with_fallbacks(
                extracted=extracted,
                file_path=file_path,
                curriculum_text=curriculum_text,
            )

            extracted["cargo_pretendido"] = cargo_canonico

            print("[EMAIL] Dados extraídos:")
            print("         Nome:", extracted.get("nome_completo") or "(vazio)")
            print("         Idade:", extracted.get("idade") or "(vazio)")
            print("         Localização:", extracted.get("localizacao") or "(vazio)")
            print(
                "         Cargo pretendido:",
                extracted.get("cargo_pretendido") or "(vazio)",
            )
            print("         Habilidades:", extracted.get("habilidades") or "(vazio)")
            print("         Formação:", extracted.get("formacoes") or "(vazio)")
            print("         Email:", extracted.get("email") or "(vazio)")
            print("         Telefone:", extracted.get("telefone") or "(vazio)")

            has_any_data = any(
                [
                    extracted.get("nome_completo"),
                    extracted.get("localizacao"),
                    extracted.get("habilidades"),
                    extracted.get("formacoes"),
                    extracted.get("email"),
                    extracted.get("telefone"),
                ]
            )

            if not has_any_data:
                print(
                    f"[EMAIL] Extração vazia para {file_path.name}. Linha NÃO será gravada."
                )
                continue

            append_candidate_to_sheet(
                file_path=file_path,
                sender=sender,
                level=level,
                portfolio=portfolio,
                extracted=extracted,
                explicit_sheet=explicit_sheet,
                subject=subject,
            )

        processed += 1

    print("\n[EMAIL] Processamento concluído.")
    print("[EMAIL] Emails avaliados:", len(mail_ids))
    print("[EMAIL] Emails processados:", processed)
    print("[EMAIL] Emails com currículo:", with_resume)
    print("[EMAIL] Emails ignorados:", ignored)
    print("[EMAIL] Emails com currículo sem vaga explícita:", no_explicit_role)


if __name__ == "__main__":
    process_inbox()
