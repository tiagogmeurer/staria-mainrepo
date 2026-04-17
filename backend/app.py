print("### StarIA Backend is ON. ###")

import os
import re
import random
import unicodedata
from pathlib import Path

from fastapi import FastAPI, HTTPException
from pydantic import BaseModel

from datasets.professional_profiles.loader import get_profiles_catalog_paths

from tools.ollama_client import ollama_chat
from tools.spreadsheets import read_excel_preview, compute_basic_stats
from tools.drive_sync import list_files
from tools.automations import create_folder, write_text_report
from rag.retriever import retrieve

from datasets.professional_profiles.loader import get_profiles_catalog_paths
from datasets.professional_profiles.matching_engine import (
    search_candidates_by_profile_query,
    format_match_summary,
)


APP_NAME = "StarIA"
MODEL_DEFAULT = os.getenv("STAR_OLLAMA_MODEL", "star-llama:latest")

# Raiz do Drive do StarIA
STARIA_DRIVE_ROOT_ENV = os.getenv("STARIA_DRIVE_ROOT")
DRIVE_SYNC_ROOT_ENV = os.getenv("DRIVE_SYNC_ROOT")
DRIVE_ROOT = STARIA_DRIVE_ROOT_ENV or r"G:\Drives compartilhados\STARMKT\StarIA"

SYSTEM_PROMPT = """Você é o StarIA, cérebro corporativo da StarMKT.
Regras:
- Seja objetivo, orientado a ação e seguro.
- Se não tiver contexto suficiente, peça o mínimo necessário mas não atue na suposição.
- Nunca invente números/valores de documentos: se não estiver no contexto, diga que não está.
- Quando houver CONTEXTO fornecido pelo sistema, você DEVE responder usando esse conteúdo.
- Não recuse pedidos de resumo/extração quando o texto já estiver no CONTEXTO.
"""

STRICT_RAG_SYSTEM_PROMPT = """
Você é um assistente factual corporativo baseado EXCLUSIVAMENTE no contexto documental fornecido.

REGRAS OBRIGATÓRIAS:
1. Use somente fatos explicitamente presentes no CONTEXTO.
2. Nunca invente números, nomes, datas, totais ou conclusões.
3. Nunca estime, arredonde, complete lacunas ou use conhecimento externo.
4. Se a resposta não estiver explícita no CONTEXTO, responda EXATAMENTE:
Não encontrei essa informação explicitamente nos documentos disponíveis.
5. Se houver valor aproximado no contexto, preserve a formulação original.
   Exemplo: se o texto diz "beira cem funcionários", NÃO converta para 100 ou 10.
6. Sempre ancore a resposta na evidência textual.
7. Nunca cite arquivos que não estejam no CONTEXTO recebido.
8. Nunca diga que encontrou informação em "algum arquivo não especificado".

FORMATO DE RESPOSTA:
- Se encontrou a resposta no contexto:
Resposta: <resposta curta e fiel ao texto>
Evidência: "<trecho literal mínimo que sustenta a resposta>"
Fonte: <caminho da fonte exatamente como aparece no contexto>

- Se NÃO encontrou:
Não encontrei essa informação explicitamente nos documentos disponíveis.
""".strip()

app = FastAPI(title=APP_NAME)


def _safe_base() -> Path:
    return Path(DRIVE_ROOT)


def _curriculos_base() -> Path:
    return _safe_base() / "curriculos"


def _profiles_base() -> Path:
    env_path = os.getenv("STARIA_PROFILES_DIR")
    if env_path:
        return Path(env_path)
    return _safe_base() / "banco_talentos" / "perfis"


def _profiles_catalog_xlsx() -> Path:
    env_path = os.getenv("STARIA_PROFILES_XLSX")
    if env_path:
        return Path(env_path)
    return _profiles_base() / "profiles_catalog.xlsx"


def _looks_like_profile_matching_intent(q: str) -> bool:
    ql = _search_norm(q)
    triggers = [
        "preciso de",
        "procuro",
        "quero encontrar",
        "quais candidatos",
        "melhores candidatos",
        "aderencia",
        "aderência",
        "perfil ideal",
        "mais aderente",
        "mais aderentes",
    ]
    return any(t in ql for t in triggers)

def _run_profile_matching(question: str) -> dict:
    return search_candidates_by_profile_query(
        query=question,
        limit=5,
        min_score=0.10,
    )

def _profiles_base() -> Path:
    env_path = os.getenv("STARIA_PROFILES_DIR")
    if env_path:
        return Path(env_path)
    return _safe_base() / "banco_talentos" / "perfis"


def _profiles_catalog_xlsx() -> Path:
    env_path = os.getenv("STARIA_PROFILES_XLSX")
    if env_path:
        return Path(env_path)
    return _profiles_base() / "profiles_catalog.xlsx"

@app.on_event("startup")
def startup_check():
    print("[StarIA] STARIA_DRIVE_ROOT env =", STARIA_DRIVE_ROOT_ENV)
    print("[StarIA] DRIVE_SYNC_ROOT env =", DRIVE_SYNC_ROOT_ENV)
    print("[StarIA] DRIVE_ROOT =", DRIVE_ROOT)
    print("[StarIA] CURRICULOS_PATH =", str(_curriculos_base()))
    print("[StarIA] CURRICULOS_EXISTS =", _curriculos_base().exists())

    print("[StarIA] PROFILES_DIR =", str(_profiles_base()))
    print("[StarIA] PROFILES_CATALOG_XLSX =", str(_profiles_catalog_xlsx()))
    print("[StarIA] PROFILES_EXISTS =", _profiles_catalog_xlsx().exists())

    try:
        print("[StarIA] PROFILES_PATHS =", get_profiles_catalog_paths())
    except Exception as e:
        print("[StarIA] PROFILES_PATHS_ERROR =", repr(e))

class AskRequest(BaseModel):
    question: str
    model: str | None = None
    use_rag: bool = True

class MatchRequest(BaseModel):
    query: str
    limit: int = 10
    min_score: float = 0.15


@app.get("/health")
def health():
    return {
        "service": "StarNodeV0",
        "status": "online"
    }

@app.get("/debug/paths")
def debug_paths():
    return {
        "drive_root": str(_safe_base()),
        "curriculos_path": str(_curriculos_base()),
        "curriculos_exists": _curriculos_base().exists(),
        "profiles_dir": str(_profiles_base()),
        "profiles_dir_exists": _profiles_base().exists(),
        "profiles_catalog_xlsx": str(_profiles_catalog_xlsx()),
        "profiles_catalog_exists": _profiles_catalog_xlsx().exists(),
    }


def _norm(s: str) -> str:
    return (s or "").strip().lower().replace("\\", "/")


def _strip_accents(text: str) -> str:
    text = text or ""
    return "".join(
        c for c in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(c)
    )


def _search_norm(text: str) -> str:
    text = _strip_accents(text or "").lower()
    text = re.sub(r"[^a-z0-9\s/+-]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _singularize_pt(text: str) -> str:
    t = _search_norm(text)
    words = []
    for w in t.split():
        if len(w) > 3 and w.endswith("s"):
            words.append(w[:-1])
        else:
            words.append(w)
    return " ".join(words).strip()


def _user_wants_sources(q: str) -> bool:
    q = (q or "").strip().lower()
    triggers = [
        "mostre as fontes",
        "mostrar fontes",
        "quais fontes",
        "fonte",
        "fontes",
        "source",
        "sources",
        "mostre o caminho dos arquivos",
        "mostrar o caminho dos arquivos",
        "caminho dos arquivos",
        "caminho do arquivo",
        "caminhos",
    ]
    return any(t in q for t in triggers)


def _list_curriculos(limit: int = 300) -> list[dict]:
    base = _curriculos_base()
    if not base.exists():
        return []

    exts = {".pdf", ".docx", ".doc", ".txt", ".rtf"}
    out: list[dict] = []

    for p in base.rglob("*"):
        if not p.is_file():
            continue

        ext = p.suffix.lower()
        if ext not in exts:
            continue

        try:
            st = p.stat()
            out.append(
                {
                    "name": p.name,
                    "path": str(p),
                    "ext": ext,
                    "size": st.st_size,
                    "mtime": int(st.st_mtime),
                }
            )
        except Exception:
            out.append(
                {
                    "name": p.name,
                    "path": str(p),
                    "ext": ext,
                }
            )

        if len(out) >= limit:
            break

    out.sort(key=lambda x: x.get("name", "").lower())
    return out


def _is_curriculos_scope(q: str) -> bool:
    q = q.strip().lower()
    return bool(
        re.search(r"\bcurr[ií]cul", q)
        or "curriculo" in q
        or "currículos" in q
        or "curriculos" in q
    )


def _is_list_curriculos_intent(q: str) -> bool:
    q = q.strip().lower()
    return _is_curriculos_scope(q) and bool(
        re.search(r"\b(quais|liste|listar|lista|mostre|mostrar|ver|nomes)\b", q)
    )


def _is_count_curriculos_intent(q: str) -> bool:
    q = q.strip().lower()
    return _is_curriculos_scope(q) and bool(
        re.search(r"\b(quantos|quantas|qtd|quantidade|total|n[uú]mero)\b", q)
    )


def _is_greeting(q: str) -> bool:
    q = (q or "").strip().lower()

    greetings = [
        "opa",
        "oi",
        "olá",
        "ola",
        "hey",
        "eai",
        "e aí",
        "e ai",
        "bom dia",
        "boa tarde",
        "boa noite",
        "como vai",
        "tudo bem",
        "fala aí",
        "fala ai",
        "salve",
    ]

    if q in greetings:
        return True

    return any(q.startswith(g) for g in greetings)


def _get_greeting_reply() -> str:
    options = [
        "Olá! Sou a StarIA, a inteligência artificial da StarMKT. Como posso ajudar você hoje?",
        "Opa! Fala aí — sou a StarIA. Como posso te ajudar agora?",
        "Hey! StarIA na área. Como posso ser útil?",
        "Oi! Sou a StarIA, assistente da StarMKT. Me diga como posso ajudar.",
        "Olá! Tudo certo? Sou a StarIA. O que você precisa agora?",
    ]
    return random.choice(options)


def _extract_forced_file(q: str) -> str | None:
    ql = (q or "").lower()

    markers = [
        "use apenas o arquivo ",
        "usar apenas o arquivo ",
        "somente o arquivo ",
        "apenas o arquivo ",
    ]

    for marker in markers:
        idx = ql.find(marker)
        if idx >= 0:
            tail = q[idx + len(marker):].strip()

            for sep in [":", ";", ",", "\n", "?", " qual", " quant", " responda", " diga"]:
                pos = tail.lower().find(sep)
                if pos > 0:
                    tail = tail[:pos].strip()

            tail = tail.strip(' "\'“”‘’')
            if tail:
                return tail
    return None


def _hit_path(hit: dict) -> str:
    meta = hit.get("meta") or {}
    return str(meta.get("path", hit.get("id", "")) or "")


def _basename(path: str) -> str:
    p = (path or "").replace("\\", "/").rstrip("/")
    if "/" in p:
        return p.rsplit("/", 1)[-1]
    return p


def _filter_hits_by_forced_file(hits: list[dict], forced_file: str) -> list[dict]:
    if not forced_file:
        return hits

    forced_norm = _norm(forced_file)
    out = []

    for h in hits:
        path = _hit_path(h)
        base = _basename(path)

        path_norm = _norm(path)
        base_norm = _norm(base)

        if forced_norm == path_norm or forced_norm == base_norm or forced_norm in path_norm:
            out.append(h)

    return out


def _filter_hits_for_curriculos_scope(hits: list[dict]) -> list[dict]:
    out = []
    for h in hits:
        meta = h.get("meta") or {}
        folder = str(meta.get("folder", "") or "").strip().lower()
        path = _hit_path(h).lower().replace("\\", "/")

        if folder == "curriculos" or "/curriculos/" in path or path.endswith("/curriculos") or "curriculos" in path:
            out.append(h)

    return out


def _dedupe_sources(paths: list[str]) -> list[str]:
    seen = set()
    out = []
    for p in paths:
        key = _norm(p)
        if key and key not in seen:
            seen.add(key)
            out.append(p)
    return out


def _candidate_name_from_path(path: str) -> str:
    base = _basename(path)
    stem = Path(base).stem
    stem = re.sub(r"[_\-]+", " ", stem)
    stem = re.sub(r"\s+", " ", stem).strip()
    return stem or base


def _clean_snippet(text: str, max_len: int = 220) -> str:
    t = re.sub(r"\s+", " ", (text or "")).strip()
    if len(t) <= max_len:
        return t
    return t[:max_len].rstrip() + "..."


def _looks_like_talent_search_intent(q: str) -> bool:
    ql = (q or "").strip().lower()

    triggers = [
        "candidato",
        "candidatos",
        "talento",
        "talentos",
        "perfil",
        "perfis",
        "vaga",
        "shortlist",
        "banco de talentos",
        "quem tem",
        "quem possui",
        "quem conhece",
        "experiência em",
        "experiencia em",
        "com experiência",
        "com experiencia",
        "que trabalhou com",
        "adequado para",
        "adequada para",
        "aderente à vaga",
        "aderente a vaga",
        "alguém com",
        "alguem com",
        "profissional com",
    ]

    if any(t in ql for t in triggers):
        return True

    if bool(re.search(r"\b(photoshop|excel|crm|varejo|atendimento|social media|tráfego pago|trafego pago|designer|design|marketing|comercial|rh|recrutamento)\b", ql)):
        return True

    return False


def _build_talent_bank_answer(question: str, hits: list[dict], max_candidates: int = 5) -> tuple[str, list[str], list[dict]]:
    grouped: dict[str, dict] = {}

    for h in hits:
        path = _hit_path(h)
        if not path:
            continue

        doc = (h.get("doc") or "").strip()
        if not doc:
            continue

        meta = h.get("meta") or {}
        chunk = meta.get("chunk", "?")

        if path not in grouped:
            grouped[path] = {
                "path": path,
                "name": _candidate_name_from_path(path),
                "evidences": [],
                "chunks": [],
                "score": 0,
            }

        snippet = _clean_snippet(doc)
        if snippet and snippet not in grouped[path]["evidences"]:
            grouped[path]["evidences"].append(snippet)

        grouped[path]["chunks"].append(chunk)
        grouped[path]["score"] += 1

    items = sorted(
        grouped.values(),
        key=lambda x: (-x["score"], x["name"].lower())
    )

    items = items[:max_candidates]
    sources = [item["path"] for item in items]

    if not items:
        answer = "Não encontrei candidatos com evidências explícitas para esse critério nos currículos disponíveis."
        return answer, sources, items

    lines = [f"Encontrei {len(items)} candidato(s) com indícios relevantes nos currículos para esse critério:"]
    for idx, item in enumerate(items, 1):
        lines.append(f"{idx}. {item['name']}")
        if item["evidences"]:
            lines.append(f"   Evidência: \"{item['evidences'][0]}\"")
        lines.append(f"   Fonte: {item['path']}")
        if len(item["evidences"]) > 1:
            lines.append(f"   Evidência adicional: \"{item['evidences'][1]}\"")

    answer = "\n".join(lines)
    return answer, sources, items


def _is_identity_question(q: str) -> bool:
    ql = (q or "").strip().lower()
    triggers = [
        "quem é você",
        "quem é vc",
        "o que você é",
        "oq vc é",
        "qual seu nome",
        "qual é seu nome",
        "você é quem",
        "vc é quem",
        "quem é a staria",
        "o que é a staria",
    ]
    return any(t in ql for t in triggers)


def _is_company_question(q: str) -> bool:
    ql = (q or "").strip().lower()
    return (
        "starmkt" in ql
        and any(
            t in ql
            for t in [
                "o que é",
                "quem é",
                "qual é",
                "me fale sobre",
                "fale sobre",
                "explique",
                "do que se trata",
                "qual a missão",
                "qual é a missão",
            ]
        )
    )


BANCO_TALENTOS_XLSX = _safe_base() / "banco_talentos" / "banco_talentos.xlsx"


def _is_banco_talentos_question(q: str) -> bool:
    ql = _search_norm(q)

    triggers = [
        "banco de talentos",
        "banco talentos",
        "banco de dados",
        "base de candidatos",
        "base de talento",
        "candidato do banco",
        "candidatos do banco",
        "no banco",
        "na base",
    ]

    return any(t in ql for t in triggers)



def _known_job_aliases() -> dict[str, list[str]]:
    return {
        "diretor de arte": [
            "diretor de arte",
            "diretores de arte",
            "direcao de arte",
            "direção de arte",
            "DA",
        ],
        "designer": [
            "designer",
            "designers",
            "design",
            "design visual",
        ],
        "social media": [
            "social media",
            "social medias",
        ],
        "copywriter": [
            "copywriter",
            "copywriters",
        ],
        "redator": [
            "redator",
            "redatores",
            "redacao",
            "redação",
        ],
        "motion designer": [
            "motion designer",
            "motion designers",
            "motion",
        ],
        "editor de video": [
            "editor de video",
            "editor de vídeos",
            "editor de videos",
            "video maker",
            "videomaker",
        ],
        "atendimento": [
            "atendimento",
            "atendimentos",
        ],
        "planejamento": [
            "planejamento",
            "planejamentos",
            "planner",
        ],
        "analista": [
            "analista",
            "analistas",
        ],
        "gerente": [
            "gerente",
            "gerentes",
        ],
        "trafego pago": [
            "trafego pago",
            "tráfego pago",
            "gestor de trafego",
            "gestor de tráfego",
            "midia paga",
            "mídia paga",
        ],
        "ui": [
            "ui",
            "ui design",
            "designer ui",
        ],
        "ux": [
            "ux",
            "ux design",
            "designer ux",
        ],
    }



def _load_banco_talentos_rows(limit: int = 1000) -> list[dict]:
    from openpyxl import load_workbook

    if not BANCO_TALENTOS_XLSX.exists():
        return []

    wb = load_workbook(BANCO_TALENTOS_XLSX, data_only=True)
    ws = wb.active

    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {}
        has_any = False

        for idx, val in enumerate(row):
            key = headers[idx] if idx < len(headers) else f"col_{idx+1}"
            item[key] = val
            if val not in (None, ""):
                has_any = True

        if has_any:
            rows.append(item)

    return rows[:limit]


def _extract_requested_job_title(question: str) -> str:
    q = _search_norm(question)

    # 1) tenta achar qualquer alias conhecido em qualquer lugar da frase
    for canonical, aliases in _known_job_aliases().items():
        for alias in aliases:
            if _search_norm(alias) in q:
                return canonical

    # 2) fallback por regex, caso o usuário escreva algo fora da lista
    patterns = [
        r"(?:quais sao os|quais sao as|quem sao os|quem sao as|liste os|liste as|listar os|listar as|mostre os|mostre as)\s+(.+?)\s+(?:do banco de talentos|da base de candidatos|do banco de dados|do banco|na base|no banco)",
        r"(?:tem|temos|existem|ha)\s+(.+?)\s+(?:no banco de talentos|no banco de dados|no banco|na base de candidatos|na base)",
        r"(?:quero ver|procure|buscar|busque)\s+(.+?)\s+(?:do banco de talentos|da base de candidatos|do banco de dados|do banco|na base)",
        r"(.+?)\s+(?:no banco de talentos|no banco de dados|no banco|na base de candidatos|na base)\??$",
    ]

    for pattern in patterns:
        m = re.search(pattern, q)
        if m:
            cargo = m.group(1).strip(" .,:;!?")
            cargo = re.sub(
                r"^(tem|temos|existem|ha|quais sao os|quais sao as|quem sao os|quem sao as|liste os|liste as|listar os|listar as|mostre os|mostre as)\s+",
                "",
                cargo,
            ).strip()
            if cargo:
                return cargo

    return ""



def _cargo_matches(cargo_planilha: str, cargo_busca: str) -> bool:
    cargo_plan = _search_norm(cargo_planilha)
    cargo_query = _search_norm(cargo_busca)

    if not cargo_plan or not cargo_query:
        return False

    # match direto
    if cargo_query in cargo_plan or cargo_plan in cargo_query:
        return True

    # match por aliases conhecidos
    aliases_map = _known_job_aliases()

    query_aliases = set(aliases_map.get(cargo_query, [cargo_query]))
    query_aliases = {_search_norm(x) for x in query_aliases}

    for alias in query_aliases:
        if alias and alias in cargo_plan:
            return True

    # match por tokens
    cargo_plan_tokens = set(_singularize_pt(cargo_plan).split())
    cargo_query_tokens = set(_singularize_pt(cargo_query).split())

    if cargo_query_tokens and cargo_query_tokens.issubset(cargo_plan_tokens):
        return True

    return False


def _format_banco_candidate_line(row: dict) -> str:
    nome = str(row.get("Nome completo") or "").strip() or "Sem nome"
    cargo = str(row.get("Cargo pretendido") or "").strip()
    nivel = str(row.get("Nível") or "").strip()
    localizacao = str(row.get("Localização") or "").strip()

    linha = f"- {nome}"
    if cargo:
        linha += f" | Cargo: {cargo}"
    if nivel:
        linha += f" | Nível: {nivel}"
    if localizacao:
        linha += f" | Local: {localizacao}"
    return linha


@app.post("/ask")
def ask(req: AskRequest):
    model = req.model or MODEL_DEFAULT
    question = (req.question or "").strip()

    # 0-A) Identidade do assistente
    if _is_identity_question(question):
        answer = ollama_chat(
            model=model,
            system=SYSTEM_PROMPT,
            user=(
                "Responda de forma curta, natural e institucional. "
                "Explique quem é a StarIA dentro da StarMKT. "
                "Não diga que não encontrou documentos. "
                "Fale como assistente corporativa da StarMKT."
            ),
            context=None,
            temperature=0.2,
            num_ctx=2048,
        ).strip()
        return {"answer": answer}

    # 0-B) Pergunta institucional sobre a StarMKT
    if _is_company_question(question):
        institutional_context = (
            "A StarMKT é uma empresa focada em campanhas promocionais impactantes "
            "para o setor de atacarejo, com objetivo de impulsionar crescimento de negócios "
            "e engajamento do cliente. O lema da empresa é 'Vender mais e melhor.'"
        )

        answer = ollama_chat(
            model=model,
            system=(
                "Você é o StarIA, cérebro corporativo da StarMKT. "
                "Responda com base no contexto institucional fornecido. "
                "Seja breve, claro e natural."
            ),
            user=question,
            context=institutional_context,
            temperature=0.2,
            num_ctx=2048,
        ).strip()
        return {"answer": answer}

    # 0-C) Saudação simples
    if _is_greeting(question):
        return {"answer": _get_greeting_reply()}
    
    # 0-D) intenção de matching de perfil profissional
    if _looks_like_profile_matching_intent(question):
        result = _run_profile_matching(question)

    # se encontrou candidatos aderentes, responde imediatamente
        if result.get("matches"):
            return {
               "answer": format_match_summary(result),
               "profile_match": result.get("profile"),
               "matches": result.get("matches", []),
            }


    # 1) INVENTÁRIO: contagem real de currículos
    if _is_count_curriculos_intent(question):
        files = _list_curriculos(limit=1000)
        base_path = str(_curriculos_base())

        response = {
            "answer": f"Existem {len(files)} currículo(s) na pasta de currículos.",
            "files": files,
        }
        if _user_wants_sources(question):
            response["sources"] = [base_path]
        return response

    # 2) INVENTÁRIO: listagem real de currículos
    if _is_list_curriculos_intent(question):
        files = _list_curriculos(limit=300)
        base_path = str(_curriculos_base())

        if not files:
            response = {
                "answer": "Ainda não há currículos na pasta de currículos.",
                "files": [],
            }
            if _user_wants_sources(question):
                response["sources"] = [base_path]
            return response

        lines = [f"- {f['name']}" for f in files]
        response = {
            "answer": f"Currículos encontrados ({len(files)}):\n" + "\n".join(lines),
            "files": files,
        }
        if _user_wants_sources(question):
            response["sources"] = [base_path]
        return response

   # 2.5) BANCO DE TALENTOS via planilha
    if _is_banco_talentos_question(question):
        rows = _load_banco_talentos_rows(limit=2000)

        if not rows:
            return {"answer": "O banco de talentos está vazio no momento."}

        cargo_solicitado = _extract_requested_job_title(question)

        print("\n[DEBUG BANCO] question =", question)
        print("[DEBUG BANCO] cargo_solicitado =", cargo_solicitado)

        if cargo_solicitado:
            filtrados = []

            for r in rows:
                cargo = str(r.get("Cargo pretendido") or "").strip()

                if _cargo_matches(cargo, cargo_solicitado):
                    filtrados.append(r)

            print("[DEBUG BANCO] total_filtrados =", len(filtrados))

            if not filtrados:
                return {
                    "answer": f"Não encontrei candidatos com cargo compatível com '{cargo_solicitado}' no banco de talentos."
                }

            linhas = [_format_banco_candidate_line(r) for r in filtrados[:30]]

            answer = (
                f"Encontrei {len(filtrados)} candidato(s) com cargo compatível com '{cargo_solicitado}' no banco de talentos:\n\n"
                + "\n".join(linhas)
            )

            return {"answer": answer}

        total = len(rows)
        nomes = [_format_banco_candidate_line(r) for r in rows[:10]]

        answer = (
            f"Nosso banco de talentos possui {total} candidato(s) registrado(s).\n\n"
            f"Primeiros registros:\n" + "\n".join(nomes)
        )

        return {"answer": answer}\

    # 3) BANCO DE TALENTOS: busca por perfil/habilidade/cargo com evidência nos currículos
    if req.use_rag and _looks_like_talent_search_intent(question):
        hits = retrieve(question, k=12, where={"folder": "curriculos"}) or []
        hits = _filter_hits_for_curriculos_scope(hits)

        print("\n[DEBUG TALENT BANK] question =", question)
        print("[DEBUG TALENT BANK] hits count =", len(hits))

        if hits:
            for i, h in enumerate(hits, 1):
                meta = h.get("meta") or {}
                print(f"[DEBUG TALENT BANK] hit {i} path =", meta.get("path"))
                print(f"[DEBUG TALENT BANK] hit {i} chunk =", meta.get("chunk"))
                print(f"[DEBUG TALENT BANK] hit {i} doc preview =", (h.get("doc") or "")[:300])

        answer, sources, ranked_items = _build_talent_bank_answer(question, hits, max_candidates=5)

        response = {
            "answer": answer,
            "matches": [
                {
                    "candidate": item["name"],
                    "path": item["path"],
                    "evidences": item["evidences"][:2],
                    "score": item["score"],
                }
                for item in ranked_items
            ],
        }

        if _user_wants_sources(question):
            response["sources"] = sources

        return response

    # 4) RAG factual controlado
    context = None
    sources: list[str] = []
    forced_file = _extract_forced_file(question)

    if req.use_rag:
        where = {"folder": "curriculos"} if _is_curriculos_scope(question) else None
        hits = retrieve(question, k=6, where=where) or []

        print("\n[DEBUG] question =", question)
        print("[DEBUG] forced_file =", forced_file)
        print("[DEBUG] initial hits count =", len(hits))

        if _is_curriculos_scope(question):
            hits = _filter_hits_for_curriculos_scope(hits)
            print("[DEBUG] hits after curriculos filter =", len(hits))

        if forced_file:
            hits = _filter_hits_by_forced_file(hits, forced_file)
            print("[DEBUG] hits after forced_file filter =", len(hits))

        if hits:
            for i, h in enumerate(hits, 1):
                meta = h.get("meta") or {}
                print(f"[DEBUG] hit {i} path =", meta.get("path"))
                print(f"[DEBUG] hit {i} chunk =", meta.get("chunk"))
                print(f"[DEBUG] hit {i} doc preview =", (h.get("doc") or "")[:300])

        if hits:
            parts = []
            collected_sources = []

            for h in hits:
                meta = h.get("meta") or {}
                path = meta.get("path", h.get("id"))
                chunk = meta.get("chunk", "?")
                doc = (h.get("doc") or "").strip()

                if not doc:
                    continue

                collected_sources.append(path)
                parts.append(f"[FONTE: {path} | chunk {chunk}]\n{doc}")

            sources = _dedupe_sources(collected_sources)
            context = "\n\n---\n\n".join(parts) if parts else None

            print("[DEBUG] context preview =", (context or "")[:1200])

    # 5) Sem contexto válido
    if req.use_rag and not context:
        non_documental = (
            not _is_curriculos_scope(question)
            and not _looks_like_talent_search_intent(question)
            and not _user_wants_sources(question)
        )

        if non_documental:
            answer = ollama_chat(
                model=model,
                system=SYSTEM_PROMPT,
                user=question,
                context=None,
                temperature=0.2,
                num_ctx=2048,
            ).strip()
            return {"answer": answer}

        response = {
            "answer": "Não encontrei essa informação explicitamente nos documentos disponíveis."
        }
        if _user_wants_sources(question):
            response["sources"] = sources
        return response

    # 6) Resposta factual rígida
    answer = ollama_chat(
        model=model,
        system=STRICT_RAG_SYSTEM_PROMPT,
        user=question,
        context=context,
        temperature=0.0,
        num_ctx=4096,
    ).strip()

    # 7) Pós-blindagem
    if not answer:
        answer = "Não encontrei essa informação explicitamente nos documentos disponíveis."

    lower_answer = answer.lower()
    has_not_found = "não encontrei essa informação explicitamente nos documentos disponíveis." in lower_answer
    has_evidence = "evidência:" in lower_answer
    has_source = "fonte:" in lower_answer

    if req.use_rag and not has_not_found and (not has_evidence or not has_source):
        answer = "Não encontrei essa informação explicitamente nos documentos disponíveis."

    response = {"answer": answer}
    if _user_wants_sources(question):
        response["sources"] = sources

    return response


class ListFilesRequest(BaseModel):
    rel_path: str = ""
    exts: list[str] | None = None
    limit: int = 200


@app.post("/files/list")
def files_list(req: ListFilesRequest):
    try:
        files = list_files(req.rel_path, req.exts, req.limit)
        return {"files": files}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


class ExcelPreviewRequest(BaseModel):
    path: str
    sheet_name: str | int | None = None
    n: int = 30


@app.post("/excel/preview")
def excel_preview(req: ExcelPreviewRequest):
    try:
        return read_excel_preview(req.path, req.sheet_name, req.n)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


class ExcelStatsRequest(BaseModel):
    path: str
    sheet_name: str | int | None = None


@app.post("/excel/stats")
def excel_stats(req: ExcelStatsRequest):
    try:
        return compute_basic_stats(req.path, req.sheet_name)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


class AutomationCreateFolder(BaseModel):
    path: str


@app.post("/automation/create-folder")
def automation_create_folder(req: AutomationCreateFolder):
    try:
        r = create_folder(req.path)
        return {"ok": r.ok, "message": r.message, "data": r.data}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


class AutomationWriteReport(BaseModel):
    path: str
    content: str


@app.post("/automation/write-report")
def automation_write_report(req: AutomationWriteReport):
    try:
        r = write_text_report(req.path, req.content)
        return {"ok": r.ok, "message": r.message, "data": r.data}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/match")
def match_candidates(req: MatchRequest):
    result = search_candidates_by_profile_query(
        query=req.query,
        limit=req.limit,
        min_score=req.min_score,
    )
    return {
        "answer": format_match_summary(result),
        "profile_match": result.get("profile"),
        "matches": result.get("matches", []),
        "total_candidates": result.get("total_candidates", 0),
    }


@app.get("/debug/paths")
def debug_paths():
    return {
        "drive_root": str(_safe_base()),
        "curriculos_path": str(_curriculos_base()),
        "curriculos_exists": _curriculos_base().exists(),
        "profiles_dir": str(_profiles_base()),
        "profiles_dir_exists": _profiles_base().exists(),
        "profiles_catalog_xlsx": str(_profiles_catalog_xlsx()),
        "profiles_catalog_exists": _profiles_catalog_xlsx().exists(),
    }

@app.get("/debug/curriculos-path")
def debug_curriculos_path():
    base = _curriculos_base()

    print("[StarIA] Listing curriculos from:", str(base))
    print("[StarIA] Exists:", base.exists())

    files = []

    if base.exists():
        try:
            for p in base.rglob("*"):
                if p.is_file():
                    files.append(str(p))
                    if len(files) >= 20:
                        break
        except Exception as e:
            return {
                "drive_root": str(_safe_base()),
                "curriculos_path": str(base),
                "exists": base.exists(),
                "error": str(e),
            }

    print("[StarIA] Sample files found:", len(files))

    return {
        "drive_root": str(_safe_base()),
        "curriculos_path": str(base),
        "exists": base.exists(),
        "sample_files": files,
    }