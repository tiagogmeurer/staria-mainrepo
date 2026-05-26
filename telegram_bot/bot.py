import os
import re
import shutil
import asyncio
import sys
from pathlib import Path
from datetime import datetime
from typing import Any

from dotenv import load_dotenv
import requests
from openpyxl import load_workbook

from telegram import Update
from telegram.error import NetworkError
from telegram.ext import (
    Application,
    MessageHandler,
    CommandHandler,
    ContextTypes,
    filters,
)


BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env", override=True)

BACKEND_DIR = Path(os.getenv("STARIA_BACKEND_DIR", r"C:\AI\backend"))
if str(BACKEND_DIR) not in sys.path:
    sys.path.append(str(BACKEND_DIR))

from rh.talent_bank_workbook import (
    CANONICAL_SHEETS,
    normalize_role_to_sheet_name,
    sheet_display_title,
    safe_str as rh_safe_str,
)

from workers.gmail_worker import (
    append_candidate_to_sheet,
    convert_to_pdf_if_possible,
    enrich_extracted_with_fallbacks,
    extract_candidate_data_with_ai,
    extract_explicit_role_from_email,
    extract_portfolio,
    extract_text_from_file,
    infer_candidate_level,
)


# ========= CONFIG =========
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()

STAR_API_BASE = os.getenv("STAR_API_BASE", "http://127.0.0.1:8088").strip()
STAR_USE_RAG_DEFAULT = os.getenv("STAR_USE_RAG", "true").lower() == "true"

DRIVE_ROOT = Path(
    os.getenv("STARIA_DRIVE_ROOT", r"G:\Drives compartilhados\STARMKT\StarIA")
).resolve()

BANCO_TALENTOS_DIR = DRIVE_ROOT / "banco_talentos"
CURRICULOS_DIR = BANCO_TALENTOS_DIR / "curriculos"
BANCO_TALENTOS_XLSX = BANCO_TALENTOS_DIR / "banco_talentos.xlsx"

INDEXER_SCRIPT = Path(os.getenv("INDEXER_SCRIPT", r"C:\AI\backend\index_once.py"))

ALLOWED_CHAT_IDS = os.getenv("ALLOWED_CHAT_IDS", "").strip()
ALLOWED = set()
if ALLOWED_CHAT_IDS:
    for x in ALLOWED_CHAT_IDS.split(","):
        x = x.strip()
        if x:
            ALLOWED.add(int(x))


# ========= HELPERS =========
def normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "")).strip()


def safe_str(value: Any) -> str:
    if value is None:
        return ""
    return normalize_spaces(str(value))


def now_stamp() -> str:
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


def safe_filename(name: str) -> str:
    name = (name or "").strip().replace("\u0000", "")
    name = re.sub(r'[<>:"/\\|?*]+', "_", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name[:180] if len(name) > 180 else name


def ensure_bank_paths():
    CURRICULOS_DIR.mkdir(parents=True, exist_ok=True)
    BANCO_TALENTOS_DIR.mkdir(parents=True, exist_ok=True)

    if not BANCO_TALENTOS_XLSX.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {BANCO_TALENTOS_XLSX}")


def is_allowed(update: Update) -> bool:
    chat_id = update.effective_chat.id if update.effective_chat else None
    return not ALLOWED or chat_id in ALLOWED


def is_probably_question(text: str) -> bool:
    if not text:
        return False

    t = text.strip().lower()

    if t.startswith("/"):
        return False

    if "?" in t:
        return True

    return bool(
        re.match(
            r"^(quais|qual|me diga|liste|listar|lista|mostre|mostrar|resuma|procure|busque|tem|existe|verifique|explique|defina|como|quem|onde|quando|por que|porque|fala|me fale|quantos|quantas)\b",
            t,
        )
    )


def looks_like_greeting(text: str) -> bool:
    if not text:
        return False

    t = text.strip().lower()

    greetings = [
        "opa",
        "epa",
        "oi",
        "olá",
        "ola",
        "hey",
        "eai",
        "e aí",
        "e ai",
        "e ae",
        "bom dia",
        "boa tarde",
        "boa noite",
        "como vai",
        "tudo bem",
        "fala aí",
        "fala ai",
        "fala ae",
        "salve",
    ]

    if t in greetings:
        return True

    return any(t.startswith(g) for g in greetings)


def call_star_api(question: str, use_rag: bool = True) -> dict:
    payload = {
        "question": question,
        "use_rag": use_rag,
    }

    print("[BOT] DEBUG PAYLOAD:", payload)

    r = requests.post(
        f"{STAR_API_BASE}/ask",
        json=payload,
        timeout=60,
    )
    r.raise_for_status()
    data = r.json()
    print("[BOT] JSON da API:", data)
    return data


# ========= CURRÍCULOS =========
def route_folder(user_text: str) -> Path | None:
    t = (user_text or "").strip().lower()

    if re.search(r"\bcurr[ií]cul", t) or "curriculos" in t or "currículos" in t:
        return CURRICULOS_DIR

    return None


def looks_like_curriculos_inventory(text: str) -> bool:
    s = (text or "").strip().lower()

    if not re.search(r"\bcurr[ií]cul", s):
        return False

    patterns = [
        r"\bquant(os|as)\b.*\bexist(em|e)?\b",
        r"\bquantidade\b",
        r"\bconta(r)?\b",
        r"\blistar?\b",
        r"\bquais\b.*\bexist(em|e)?\b",
        r"\btem\b.*\bcurr",
        r"\bver\b.*\bcurr",
        r"\bn[uú]mero\b.*\bcurr",
        r"\btotal\b.*\bcurr",
    ]
    return any(re.search(p, s) for p in patterns)


def list_curriculos_files(limit: int = 80) -> list[dict]:
    exts = {".pdf", ".docx", ".doc", ".txt", ".rtf"}

    if not CURRICULOS_DIR.exists():
        return []

    items = []
    for p in CURRICULOS_DIR.rglob("*"):
        if p.is_file() and p.suffix.lower() in exts:
            try:
                st = p.stat()
                items.append(
                    {
                        "name": p.name,
                        "path": str(p),
                        "size": st.st_size,
                        "mtime": int(st.st_mtime),
                    }
                )
            except Exception:
                items.append({"name": p.name, "path": str(p)})

    items.sort(key=lambda x: x.get("mtime", 0), reverse=True)
    return items[:limit]


# ========= BANCO DE TALENTOS =========
def is_banco_talentos_question(text: str) -> bool:
    t = (text or "").strip().lower()
    triggers = [
        "banco de talentos",
        "banco talentos",
        "como está nosso banco de talentos",
        "como está o banco de talentos",
        "quantos candidatos temos",
        "quantos candidatos há",
        "quantos candidatos existem",
        "liste o banco de talentos",
        "listar banco de talentos",
        "me mostre o banco de talentos",
        "resuma o banco de talentos",
    ]
    return any(x in t for x in triggers)


def is_add_to_talent_bank_intent(text: str) -> bool:
    t = (text or "").strip().lower()
    triggers = [
        "adicione o candidato em anexo no nosso banco de talentos",
        "adicionar candidato no banco de talentos",
        "adicione no banco de talentos",
        "coloque no banco de talentos",
        "adicione o anexo no banco de talentos",
        "adicionar no banco de talentos",
        "salve no banco de talentos",
    ]
    return any(x in t for x in triggers)


def _load_banco_talentos_rows(limit: int = 9999) -> list[dict]:
    ensure_bank_paths()

    wb = load_workbook(BANCO_TALENTOS_XLSX, data_only=True)
    rows = []

    for ws in wb.worksheets:
        headers = [safe_str(c.value) for c in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            item = {"_sheet": ws.title}
            has_any = False

            for idx, val in enumerate(row):
                key = headers[idx] if idx < len(headers) else f"col_{idx+1}"
                item[key] = val
                if val not in (None, ""):
                    has_any = True

            if has_any:
                rows.append(item)

            if len(rows) >= limit:
                return rows

    return rows


def parse_score(value: Any) -> float:
    try:
        return float(str(value).replace(",", "."))
    except Exception:
        return 0.0


def build_banco_talentos_summary() -> str:
    rows = _load_banco_talentos_rows(limit=20000)

    if not rows:
        return "📂 O banco de talentos está vazio no momento."

    total = len(rows)

    by_sheet = {}
    by_level = {}

    for r in rows:
        sheet = safe_str(r.get("_sheet")) or "Sem aba"
        level = safe_str(r.get("Nível")) or "Sem nível"

        by_sheet[sheet] = by_sheet.get(sheet, 0) + 1
        by_level[level] = by_level.get(level, 0) + 1

    lines = [f"📌 Banco de talentos: {total} candidato(s) registrado(s).\n"]

    lines.append("📂 Por vaga/aba:")
    for sheet in CANONICAL_SHEETS:
        if sheet in by_sheet:
            lines.append(f"- {sheet_display_title(sheet)}: {by_sheet[sheet]}")

    extra_sheets = sorted(set(by_sheet) - set(CANONICAL_SHEETS))
    for sheet in extra_sheets:
        lines.append(f"- {sheet}: {by_sheet[sheet]}")

    lines.append("\n🎚 Por nível:")
    for level, count in sorted(by_level.items(), key=lambda x: x[0]):
        lines.append(f"- {level}: {count}")

    lines.append("\n🏅 Últimos/top registros:")
    sorted_rows = sorted(rows, key=lambda r: parse_score(r.get("Nota")), reverse=True)

    for r in sorted_rows[:10]:
        nome = safe_str(r.get("Nome completo")) or "Sem nome"
        cargo = safe_str(r.get("Cargo pretendido")) or safe_str(r.get("_sheet"))
        nivel = safe_str(r.get("Nível")) or "Sem nível"
        nota = safe_str(r.get("Nota")) or "0"
        loc = safe_str(r.get("Localização"))

        line = f"- {nome} | {cargo} | {nivel} | Nota {nota}"
        if loc:
            line += f" | {loc}"

        lines.append(line)

    return "\n".join(lines)


def parse_top_query(text: str) -> tuple[int, str] | None:
    t = (text or "").strip().lower()

    m = re.search(r"\btop\s*(\d+)?\s+(.+)", t)
    if not m:
        return None

    limit = int(m.group(1) or 5)
    query = safe_str(m.group(2))

    if not query:
        return None

    return max(1, min(limit, 20)), query


def resolve_top_query_to_sheet(query: str) -> str:
    q = (query or "").strip().lower()

    if "comunicação" in q or "comunicacao" in q:
        return "COORDENADOR DE COMUNICAÇÃO"

    return normalize_role_to_sheet_name(query)



def get_top_candidates(query: str, limit: int = 5) -> str:
    rows = _load_banco_talentos_rows(limit=30000)

    if not rows:
        return "📂 O banco de talentos está vazio."

    target_sheet = resolve_top_query_to_sheet(query)

    # 🔥 busca EXCLUSIVAMENTE pela aba correta
    filtered = [
        r for r in rows
        if safe_str(r.get("_sheet")) == target_sheet
    ]

    if not filtered:
        return (
            f"Não encontrei candidatos para "
            f"{sheet_display_title(target_sheet)}."
        )

    filtered.sort(
        key=lambda r: parse_score(r.get("Nota")),
        reverse=True,
    )

    lines = [
        f"🏆 Top {min(limit, len(filtered))} candidato(s) para "
        f"{sheet_display_title(target_sheet)}:\n"
    ]

    for idx, r in enumerate(filtered[:limit], start=1):
        nome = safe_str(r.get("Nome completo")) or "Sem nome"
        nota = safe_str(r.get("Nota")) or "0"
        nivel = safe_str(r.get("Nível")) or "Sem nível"
        loc = safe_str(r.get("Localização")) or "Sem localização"
        email = safe_str(r.get("Email"))
        tel = safe_str(r.get("Telefone"))
        portfolio = safe_str(r.get("Portfólio"))

        lines.append(f"{idx}. {nome}")
        lines.append(
            f"   Nota: {nota} | "
            f"Nível: {nivel} | "
            f"Local: {loc}"
        )

        if email:
            lines.append(f"   Email: {email}")

        if tel:
            lines.append(f"   Telefone: {tel}")

        if portfolio:
            lines.append(f"   Portfólio/CV: {portfolio}")

        lines.append("")

    return "\n".join(lines).strip()


async def run_indexer_and_notify(msg):
    try:
        import subprocess
        import sys

        p = subprocess.run(
            [sys.executable, str(INDEXER_SCRIPT)],
            cwd=str(INDEXER_SCRIPT.parent),
            capture_output=True,
            text=True,
            timeout=600,
        )

        if p.returncode == 0:
            await msg.reply_text("📚 Indexação concluída.")
        else:
            err = (p.stderr or p.stdout or "").strip()
            await msg.reply_text(f"⚠️ Indexação falhou:\n{err[-1200:]}")
    except Exception as e:
        await msg.reply_text(f"⚠️ Não consegui rodar a indexação: {e}")


async def process_candidate_from_telegram(
    msg,
    context: ContextTypes.DEFAULT_TYPE,
    caption_or_text: str,
    chat_id: int,
):
    ensure_bank_paths()

    explicit_sheet = extract_explicit_role_from_email(caption_or_text, caption_or_text)

    if not explicit_sheet:
        await msg.reply_text(
            "⚠️ Recebi o currículo, mas não identifiquei a vaga de forma objetiva.\n\n"
            "Envie novamente com uma legenda incluindo a vaga, por exemplo:\n"
            "'adicione no banco de talentos - vaga diretor de arte sênior'\n"
            "'adicione no banco de talentos - vaga atendimento pleno'"
        )
        return

    doc = msg.document
    file_name = safe_filename(doc.file_name or f"arquivo_{now_stamp()}")
    tmp_dir = Path(r"C:\AI\runtime\tmp_telegram")
    tmp_dir.mkdir(parents=True, exist_ok=True)

    tmp_path = tmp_dir / f"{now_stamp()}__{file_name}"
    final_path = CURRICULOS_DIR / file_name

    tg_file = await context.bot.get_file(doc.file_id)
    await tg_file.download_to_drive(custom_path=str(tmp_path))

    if final_path.exists():
        stem = Path(file_name).stem
        suf = Path(file_name).suffix
        n = 2
        while True:
            cand = CURRICULOS_DIR / f"{stem}__{n}{suf}"
            if not cand.exists():
                final_path = cand
                break
            n += 1

    shutil.move(str(tmp_path), str(final_path))
    final_path = convert_to_pdf_if_possible(final_path)

    print("[BOT] Currículo salvo no banco:", final_path)

    curriculum_text = extract_text_from_file(final_path)

    extracted = await asyncio.to_thread(extract_candidate_data_with_ai, final_path)
    extracted = enrich_extracted_with_fallbacks(
        extracted=extracted,
        file_path=final_path,
        curriculum_text=curriculum_text,
    )

    cargo_canonico = sheet_display_title(explicit_sheet)
    extracted["cargo_pretendido"] = cargo_canonico

    level = infer_candidate_level(
        subject=caption_or_text,
        body=caption_or_text,
        curriculum_text=curriculum_text,
        explicit_sheet=explicit_sheet,
    )

    portfolio = extract_portfolio(caption_or_text) or extracted.get("portfolio", "")

    sender = f"Telegram chat {chat_id}"

    await asyncio.to_thread(
        append_candidate_to_sheet,
        final_path,
        sender,
        level,
        portfolio,
        extracted,
        explicit_sheet,
        caption_or_text,
    )

    confirm = (
        "✅ Candidato processado para o banco de talentos.\n\n"
        f"👤 Nome: {extracted.get('nome_completo') or 'Não identificado'}\n"
        f"💼 Vaga: {cargo_canonico}\n"
        f"🎚 Nível: {level}\n"
        f"📎 Arquivo: {final_path.name}"
    )

    await msg.reply_text(confirm[:3500])
    await run_indexer_and_notify(msg)


# ========= HANDLERS =========
async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        await update.message.reply_text("Acesso não autorizado para este bot.")
        return

    msg = update.message
    if not msg or not msg.document:
        return

    chat_id = update.effective_chat.id if update.effective_chat else 0
    caption_or_text = (msg.caption or msg.text or "").strip()

    if is_add_to_talent_bank_intent(caption_or_text):
        try:
            await msg.reply_text("📥 Recebi o currículo. Processando para o banco de talentos...")
            await process_candidate_from_telegram(msg, context, caption_or_text, chat_id)
            return
        except Exception as e:
            print("[BOT] Erro ao adicionar no banco de talentos:", repr(e))
            await msg.reply_text(f"⚠️ Erro ao adicionar no banco de talentos: {e}")
            return

    target_dir = route_folder(caption_or_text)

    if target_dir is None:
        await msg.reply_text(
            "Não entendi a pasta de destino.\n\n"
            "Exemplos:\n"
            "- 'Coloque na pasta de currículos'\n"
            "- 'adicione no banco de talentos - vaga diretor de arte sênior'"
        )
        return

    target_dir.mkdir(parents=True, exist_ok=True)

    doc = msg.document
    file_name = safe_filename(doc.file_name or f"arquivo_{now_stamp()}")
    tmp_dir = Path(r"C:\AI\runtime\tmp_telegram")
    tmp_dir.mkdir(parents=True, exist_ok=True)

    tmp_path = tmp_dir / f"{now_stamp()}__{file_name}"
    final_path = target_dir / file_name

    tg_file = await context.bot.get_file(doc.file_id)
    await tg_file.download_to_drive(custom_path=str(tmp_path))

    if final_path.exists():
        stem = Path(file_name).stem
        suf = Path(file_name).suffix
        n = 2
        while True:
            cand = target_dir / f"{stem}__{n}{suf}"
            if not cand.exists():
                final_path = cand
                break
            n += 1

    shutil.move(str(tmp_path), str(final_path))

    await msg.reply_text(f"✅ Recebido e salvo em:\n{final_path}")
    await run_indexer_and_notify(msg)


async def handle_top_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        await update.message.reply_text("Acesso não autorizado para este bot.")
        return

    query = " ".join(context.args or []).strip()

    if not query:
        await update.message.reply_text(
            "Use assim:\n/top atendimento\n/top 3 diretor de arte\n/top 5 redator"
        )
        return

    m = re.match(r"^(\d+)\s+(.+)$", query)
    if m:
        limit = int(m.group(1))
        query = m.group(2)
    else:
        limit = 5

    out = get_top_candidates(query=query, limit=limit)
    await update.message.reply_text(out[:3500])


async def handle_resumo_banco_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        await update.message.reply_text("Acesso não autorizado para este bot.")
        return

    try:
        out = build_banco_talentos_summary()
        await update.message.reply_text(out[:3500])
    except Exception as e:
        await update.message.reply_text(f"⚠️ Erro ao consultar banco de talentos: {e}")


async def handle_text_only(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        try:
            await update.message.reply_text("Acesso não autorizado para este bot.")
        except Exception as send_err:
            print("[BOT] Falha ao enviar acesso não autorizado:", repr(send_err))
        return

    msg = update.message
    t = (msg.text or "").strip() if msg and msg.text else ""
    if not t:
        return

    print("[BOT] Mensagem recebida:", t)

    parsed_top = parse_top_query(t)
    if parsed_top:
        limit, query = parsed_top
        try:
            out = get_top_candidates(query=query, limit=limit)
            await msg.reply_text(out[:3500])
            return
        except Exception as e:
            await msg.reply_text(f"⚠️ Erro ao consultar top candidatos: {e}")
            return

    if is_banco_talentos_question(t):
        try:
            out = build_banco_talentos_summary()
            await msg.reply_text(out[:3500])
            return
        except Exception as e:
            print("[BOT] Erro ao consultar banco de talentos:", repr(e))
            await msg.reply_text(f"⚠️ Erro ao consultar banco de talentos: {e}")
            return

    if looks_like_curriculos_inventory(t):
        try:
            files = list_curriculos_files(limit=80)
            total = len(files)

            if total == 0:
                await msg.reply_text(
                    "📂 Pasta de currículos está vazia.\n"
                    f"Destino: {CURRICULOS_DIR}"
                )
                return

            names = [f["name"] for f in files if f.get("name")]
            preview = names[:25]

            out = (
                f"📌 Currículos encontrados: {total}\n"
                f"📂 Pasta: {CURRICULOS_DIR}\n\n"
                "🗂 Últimos arquivos:\n- " + "\n- ".join(preview)
            )
            if total > len(preview):
                out += f"\n... (+{total - len(preview)})"

            await msg.reply_text(out[:3500])
            return

        except Exception as e:
            print("[BOT] Erro ao listar currículos:", repr(e))
            await msg.reply_text(f"⚠️ Erro ao listar currículos: {e}")
            return

    if route_folder(t) is not None and not is_probably_question(t):
        await msg.reply_text(
            "Entendi a pasta — agora me envie o arquivo junto com essa legenda.\n"
            "Ex: envie o PDF com: 'Coloque na pasta de currículos'."
        )
        return

    if is_probably_question(t) or looks_like_greeting(t):
        try:
            print("[BOT] Chamando StarIA em:", STAR_API_BASE)

            data = await asyncio.to_thread(call_star_api, t, STAR_USE_RAG_DEFAULT)

            answer = (data.get("answer") or "").strip() or "(sem resposta)"
            sources = data.get("sources") or []
            files = data.get("files") or []

            out = answer

            if files:
                names = [f.get("name") for f in files if isinstance(f, dict) and f.get("name")]
                if names:
                    out += "\n\n📎 Arquivos:\n- " + "\n- ".join(names[:20])

            if sources:
                out += "\n\n🧾 Fontes:\n- " + "\n- ".join(sources[:6])

            await msg.reply_text(out[:3500])
            return

        except Exception as e:
            print("[BOT] Erro ao consultar StarIA:", repr(e))
            await msg.reply_text(f"⚠️ Erro ao consultar StarIA: {e}")
            return

    await msg.reply_text(
        "Me envie o arquivo com uma legenda indicando a ação.\n\n"
        "Exemplos:\n"
        "- 'Coloque na pasta de currículos'\n"
        "- 'adicione no banco de talentos - vaga diretor de arte sênior'\n\n"
        "Consultas:\n"
        "- /resumo_banco\n"
        "- /top atendimento\n"
        "- /top 3 diretor de arte"
    )


async def on_error(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    err = context.error

    if isinstance(err, NetworkError):
        print("[BOT] NetworkError transitório (reconectando):", err)
        return

    print("[BOT] Exceção não tratada:", repr(err))


def main():
    if not TELEGRAM_BOT_TOKEN:
        raise SystemExit("Defina TELEGRAM_BOT_TOKEN no ambiente (.env)")

    app = (
        Application.builder()
        .token(TELEGRAM_BOT_TOKEN)
        .read_timeout(30)
        .write_timeout(30)
        .connect_timeout(30)
        .pool_timeout(30)
        .build()
    )

    app.add_handler(CommandHandler("top", handle_top_command))
    app.add_handler(CommandHandler("resumo_banco", handle_resumo_banco_command))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text_only))
    app.add_error_handler(on_error)

    print("[ONLINE] StarIA Telegram bot running in polling mode")
    app.run_polling(
        allowed_updates=Update.ALL_TYPES,
        drop_pending_updates=True,
    )


if __name__ == "__main__":
    main()